import cv2
from openpyxl import load_workbook

from bubble_scanner.auto_cli import classify_inputs, main
from bubble_scanner.template import Template
from tests.score_report_synth import write_score_report_pdf
from tests.synth import render_sheet


def make_bubble_template_yaml(tmp_path):
    yaml_text = """
page:
  width: 900
  height: 700
sections:
  - name: Answers
    columns:
      - {first_question: 1, last_question: 4, x_start: 150, y_start: 100, row_height: 80}
bubble_spacing_x: 60
bubble_radius: 18
choices:
  even: [A, B, C, D]
  odd: [F, G, H, J]
thresholds:
  fill_ratio_min: 0.35
  relative_margin: 0.15
"""
    path = tmp_path / "template.yaml"
    path.write_text(yaml_text)
    return path


def test_classify_inputs_splits_images_pdfs_and_score_reports(tmp_path):
    template_path = make_bubble_template_yaml(tmp_path)
    template = Template.from_yaml(template_path)

    image_path = tmp_path / "sheet.png"
    cv2.imwrite(str(image_path), render_sheet(template, {1: ["F"]}))

    score_pdf_path = tmp_path / "score.pdf"
    write_score_report_pdf(score_pdf_path, [(1, "Math", "A", "A", "Correct")])

    bubble_paths, score_rows = classify_inputs([image_path, score_pdf_path])

    assert bubble_paths == [image_path]
    assert len(score_rows) == 1
    assert score_rows[0].your_answer == "A"


def test_classify_inputs_expands_directory_with_mixed_contents(tmp_path):
    template_path = make_bubble_template_yaml(tmp_path)
    template = Template.from_yaml(template_path)

    batch = tmp_path / "batch"
    batch.mkdir()
    cv2.imwrite(str(batch / "sheet.png"), render_sheet(template, {1: ["F"]}))
    write_score_report_pdf(batch / "score.pdf", [(1, "Math", "A", "A", "Correct")])
    (batch / "notes.txt").write_text("ignore me")

    bubble_paths, score_rows = classify_inputs([batch])

    assert len(bubble_paths) == 1
    assert len(score_rows) == 1


def test_auto_cli_produces_two_tabs_for_mixed_input(tmp_path):
    template_path = make_bubble_template_yaml(tmp_path)
    template = Template.from_yaml(template_path)

    image_path = tmp_path / "sheet.png"
    cv2.imwrite(str(image_path), render_sheet(template, {1: ["F"], 2: ["A"]}))

    score_pdf_path = tmp_path / "score.pdf"
    write_score_report_pdf(
        score_pdf_path,
        [(1, "Reading and Writing", "D", "D", "Correct"), (2, "Reading and Writing", "B", "D", "Incorrect")],
    )

    output_path = tmp_path / "combined.xlsx"
    exit_code = main(
        [
            "--input",
            str(image_path),
            str(score_pdf_path),
            "--template",
            str(template_path),
            "--output",
            str(output_path),
            "--no-refresh-keys",
        ]
    )

    assert exit_code == 0
    wb = load_workbook(output_path)
    assert set(wb.sheetnames) == {"sheet", "Score Report Answers"}
    assert wb.active.title == "sheet"

    bubble_ws = wb["sheet"]
    assert [c.value for c in bubble_ws[1]] == ["Question", "Answers"]
    assert bubble_ws.cell(row=2, column=2).value == "F"  # Q1
    assert bubble_ws.cell(row=3, column=2).value == "A"  # Q2

    score_ws = wb["Score Report Answers"]
    score_rows = list(score_ws.iter_rows(min_row=2, values_only=True))
    # No matching reference key for this synthetic report -> plain "Module N" labels.
    assert score_rows == [
        ("Unknown", "Reading and Writing - Module 1", 1, "D"),
        ("Unknown", "Reading and Writing - Module 1", 2, "D"),
    ]


def test_auto_cli_handles_bubble_only_input_without_requiring_score_reports(tmp_path):
    template_path = make_bubble_template_yaml(tmp_path)
    template = Template.from_yaml(template_path)
    image_path = tmp_path / "sheet.png"
    cv2.imwrite(str(image_path), render_sheet(template, {1: ["G"]}))

    output_path = tmp_path / "out.xlsx"
    exit_code = main(
        ["--input", str(image_path), "--template", str(template_path), "--output", str(output_path)]
    )

    assert exit_code == 0
    wb = load_workbook(output_path)
    assert wb.sheetnames == ["sheet"]
    assert wb.active.title == "sheet"


def test_auto_cli_handles_score_report_only_input_without_needing_a_valid_template(tmp_path):
    score_pdf_path = tmp_path / "score.pdf"
    write_score_report_pdf(score_pdf_path, [(1, "Math", "A", "A", "Correct")])

    output_path = tmp_path / "out.xlsx"
    exit_code = main(
        [
            "--input",
            str(score_pdf_path),
            "--template",
            str(tmp_path / "nonexistent_template.yaml"),
            "--output",
            str(output_path),
            "--no-refresh-keys",
        ]
    )

    assert exit_code == 0
    wb = load_workbook(output_path)
    assert wb.sheetnames == ["Score Report Answers"]
