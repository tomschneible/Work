from openpyxl import load_workbook

from bubble_scanner.score_report_cli import main
from tests.score_report_synth import write_score_report_pdf


def test_score_report_cli_end_to_end(tmp_path):
    input_path = tmp_path / "report.pdf"
    write_score_report_pdf(
        input_path,
        [
            (1, "Reading and Writing", "D", "D", "Correct"),
            (2, "Reading and Writing", "B", "D", "Incorrect"),
        ],
    )
    output_path = tmp_path / "answers.xlsx"

    exit_code = main(["--input", str(input_path), "--output", str(output_path), "--no-refresh-keys"])

    assert exit_code == 0
    assert output_path.exists()
    wb = load_workbook(output_path)
    ws = wb.active
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    assert len(rows) == 2
    # No matching reference key for this synthetic report -> plain "Module N" labels.
    assert rows[0][2:] == ("Module 1", 1, "Reading and Writing", "D")
    assert rows[1][2:] == ("Module 1", 2, "Reading and Writing", "D")


def test_score_report_cli_reports_missing_input(tmp_path, capsys):
    exit_code = main(
        [
            "--input",
            str(tmp_path / "does_not_exist.pdf"),
            "--output",
            str(tmp_path / "out.xlsx"),
        ]
    )
    assert exit_code == 1
    assert "does not exist" in capsys.readouterr().err
