"""Write parsed score-report rows to an .xlsx spreadsheet."""
from __future__ import annotations

from pathlib import Path
from typing import List

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

from .score_report import ScoreReportRow


def add_score_report_answers_sheet(
    wb: Workbook, rows: List[ScoreReportRow], title: str = "Score Report Answers"
) -> None:
    """Add a worksheet with score-report answers to an existing workbook
    (used both standalone by write_score_report_xlsx and alongside a
    bubble-sheet answers sheet when a batch mixes both input types)."""
    if not rows:
        raise ValueError("No rows to export")

    ws = wb.create_sheet(title=title)

    header = ["Test", "Section / Module", "Question", "Your Answer"]
    ws.append(header)
    for cell in ws[1]:
        cell.font = Font(bold=True)

    for row in rows:
        module_label = row.module_label or f"Module {row.module}"
        test = row.test or "Unknown"
        section_module = f"{row.section} - {module_label}"
        ws.append([test, section_module, row.question, row.your_answer])

    column_widths = [18, 34, 10, 14]
    for col_index, width in enumerate(column_widths, start=1):
        ws.column_dimensions[get_column_letter(col_index)].width = width


def write_score_report_xlsx(rows: List[ScoreReportRow], output_path: str | Path) -> None:
    wb = Workbook()
    del wb["Sheet"]
    add_score_report_answers_sheet(wb, rows)
    wb.save(str(output_path))
