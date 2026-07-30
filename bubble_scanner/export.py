"""Write sheet results to an .xlsx spreadsheet."""
from __future__ import annotations

from pathlib import Path
from typing import List, Tuple

from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from .pipeline import SheetResult

BLANK_FILL = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")
MULTIPLE_FILL = PatternFill(start_color="F8D7DA", end_color="F8D7DA", fill_type="solid")
LOW_CONFIDENCE_FONT = Font(italic=True, color="808080")

FIXED_COLUMNS = ["Sheet", "Alignment", "Needs Review"]


def _column_order(results: List[SheetResult]) -> List[Tuple[str, int]]:
    """(section, question) keys in the order they should appear as columns,
    following the order questions were produced in (i.e. the template's
    declared section/column order), taking the union across all sheets in
    case sheets were scored against slightly different templates."""
    seen = set()
    order = []
    for result in results:
        for q in result.questions:
            key = (q.section, q.question)
            if key not in seen:
                seen.add(key)
                order.append(key)
    return order


def write_xlsx(results: List[SheetResult], output_path: str | Path) -> None:
    if not results:
        raise ValueError("No results to export")

    columns = _column_order(results)

    wb = Workbook()
    ws = wb.active
    ws.title = "Answers"

    header = FIXED_COLUMNS + [f"{section}_Q{question}" for section, question in columns]
    ws.append(header)
    for cell in ws[1]:
        cell.font = Font(bold=True)

    num_fixed = len(FIXED_COLUMNS)

    for result in results:
        by_key = {(q.section, q.question): q for q in result.questions}
        row = [
            result.label,
            "contour" if result.used_contour_alignment else "resized (no border found)",
            "YES" if result.has_review_items else "",
        ]
        for key in columns:
            q = by_key.get(key)
            row.append(q.answer if q else "")
        ws.append(row)

        row_index = ws.max_row
        for col_offset, key in enumerate(columns, start=1):
            q = by_key.get(key)
            if q is None:
                continue
            cell = ws.cell(row=row_index, column=num_fixed + col_offset)
            if q.answer == "MULTIPLE":
                cell.fill = MULTIPLE_FILL
                cell.comment = Comment(", ".join(q.candidates), "bubble_scanner")
            elif q.answer == "":
                cell.fill = BLANK_FILL
            if q.low_confidence:
                cell.font = LOW_CONFIDENCE_FONT

    for col_index in range(1, len(header) + 1):
        ws.column_dimensions[get_column_letter(col_index)].width = 12

    wb.save(str(output_path))
