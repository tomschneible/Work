"""Write parsed score-report rows to an .xlsx spreadsheet."""
from __future__ import annotations

from pathlib import Path
from typing import List

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

from .score_report import ScoreReportRow


def add_score_report_answers_sheet(
    wb: Workbook, rows: List[ScoreReportRow], title: str = "Score Report Answers"
) -> None:
    """Add a worksheet with score-report answers to an existing workbook
    (used both standalone by write_score_report_xlsx and alongside a
    bubble-sheet answers sheet when a batch mixes both input types)."""
    if not rows:
        raise ValueError("No rows to export")

    ws = wb.create_sheet(title=title)

    header = ["Source", "Module", "Question", "Section", "Your Answer"]
    ws.append(header)
    for cell in ws[1]:
        cell.font = Font(bold=True)

    for row in rows:
        ws.append([row.source, row.module, row.question, row.section, row.your_answer])

    for col_index in range(1, len(header) + 1):
        ws.column_dimensions[get_column_letter(col_index)].width = 18


def write_score_report_xlsx(rows: List[ScoreReportRow], output_path: str | Path) -> None:
    wb = Workbook()
    del wb["Sheet"]
    add_score_report_answers_sheet(wb, rows)
    wb.save(str(output_path))
