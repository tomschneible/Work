import cv2
from openpyxl import load_workbook

from bubble_scanner.cli import main
from bubble_scanner.template import Template
from tests.synth import render_sheet


def make_template_yaml(tmp_path):
    yaml_text = """
page:
  width: 900
  height: 700
sections:
  - name: Answers
    columns:
      - {first_question: 1, last_question: 4, x_start: 150, y_start: 100, row_height: 80}
bubble_spacing_x: 60
bubble_radius: 18
choices:
  even: [A, B, C, D]
  odd: [F, G, H, J]
thresholds:
  fill_ratio_min: 0.35
  relative_margin: 0.15
"""
    path = tmp_path / "template.yaml"
    path.write_text(yaml_text)
    return path


def test_cli_accepts_multiple_input_paths(tmp_path):
    template_path = make_template_yaml(tmp_path)
    template = Template.from_yaml(template_path)

    sheet_a = tmp_path / "a.png"
    sheet_b = tmp_path / "b.png"
    cv2.imwrite(str(sheet_a), render_sheet(template, {1: ["F"]}))
    cv2.imwrite(str(sheet_b), render_sheet(template, {1: ["G"], 2: ["A"]}))

    output_path = tmp_path / "results.xlsx"
    exit_code = main(
        [
            "--input",
            str(sheet_a),
            str(sheet_b),
            "--template",
            str(template_path),
            "--output",
            str(output_path),
        ]
    )

    assert exit_code == 0
    assert output_path.exists()

    wb = load_workbook(output_path)
    assert set(wb.sheetnames) == {"Overview", "a", "b"}

    ws_a = wb["a"]
    assert [c.value for c in ws_a[1]] == ["Question", "Answers"]
    assert ws_a.cell(row=2, column=2).value == "F"  # Q1

    ws_b = wb["b"]
    assert ws_b.cell(row=2, column=2).value == "G"  # Q1
    assert ws_b.cell(row=3, column=2).value == "A"  # Q2


def test_cli_reports_missing_input(tmp_path, capsys):
    template_path = make_template_yaml(tmp_path)
    exit_code = main(
        [
            "--input",
            str(tmp_path / "does_not_exist.png"),
            "--template",
            str(template_path),
            "--output",
            str(tmp_path / "out.xlsx"),
        ]
    )
    assert exit_code == 1
    assert "does not exist" in capsys.readouterr().err
