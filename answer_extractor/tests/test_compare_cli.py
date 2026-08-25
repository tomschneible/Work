"""Tests for compare_cli.main -- exercises it against both a .xlsx-vs-.xlsx
pair (the tool's original shape) and a .pdf-vs-.pdf pair (see
scoresheet_check.py's load_reference_answers/load_our_answers), since the
CLI itself is just argument-parsing glue over those two functions plus
scoresheet_check.compare/write_comparison_report/summarize, which have
their own thorough tests in test_scoresheet_check.py."""
from __future__ import annotations

import os

import openpyxl
import pytest

from answer_extractor.compare_cli import main
from tests.scoresheet_pdf_synth import MARK_FONT_PATH, Block, Row, write_scoresheet_pdf
from tests.test_scoresheet_check import _write_our_output, _write_reference_scoresheet


def test_main_compares_two_xlsx_files_and_writes_a_report(tmp_path, capsys):
    ours_path = tmp_path / "ours.xlsx"
    reference_path = tmp_path / "reference.xlsx"
    output_path = tmp_path / "comparison.xlsx"
    _write_our_output(ours_path)
    _write_reference_scoresheet(reference_path)

    exit_code = main(
        ["--ours", str(ours_path), "--reference", str(reference_path), "--output", str(output_path)]
    )

    assert exit_code == 0
    assert output_path.exists()
    wb = openpyxl.load_workbook(output_path)
    assert "Comparison" in wb.sheetnames
    out = capsys.readouterr().out
    assert "questions compared" in out


@pytest.mark.skipif(
    not os.path.exists(MARK_FONT_PATH),
    reason=f"synthetic PDF fixtures need a Unicode-capable font at {MARK_FONT_PATH} (fonts-dejavu-core)",
)
def test_main_compares_two_pdf_files(tmp_path, capsys):
    ours_path = tmp_path / "ours.pdf"
    reference_path = tmp_path / "reference.pdf"
    output_path = tmp_path / "comparison.xlsx"
    write_scoresheet_pdf(ours_path, [Block("English", "Math", [[Row(1, "A", "A")], [Row(1, "B", "C")]])])
    write_scoresheet_pdf(reference_path, [Block("English", "Math", [[Row(1, "A", "A")], [Row(1, "B", "B")]])])

    exit_code = main(
        ["--ours", str(ours_path), "--reference", str(reference_path), "--output", str(output_path)]
    )

    assert exit_code == 0
    wb = openpyxl.load_workbook(output_path)
    ws = wb["Comparison"]
    rows_by_section_and_question = {(row[0].value, row[1].value): row for row in ws.iter_rows(min_row=2)}
    assert rows_by_section_and_question[("English", 1)][4].value == "✔"  # match
    # mathematics 1: a mismatch against a PDF reference is an unflagged
    # "silent miss" (see scoresheet_check.py's module docstring).
    assert rows_by_section_and_question[("Mathematics", 1)][4].value == "✘"
    out = capsys.readouterr().out
    assert "questions compared" in out


@pytest.mark.skipif(
    not os.path.exists(MARK_FONT_PATH),
    reason=f"synthetic PDF fixtures need a Unicode-capable font at {MARK_FONT_PATH} (fonts-dejavu-core)",
)
def test_main_compares_our_pdf_against_a_reference_xlsx(tmp_path):
    """The two sides needn't match format -- e.g. checking this run's own
    freshly generated PDF report against an existing vendor .xlsx."""
    ours_path = tmp_path / "ours.pdf"
    reference_path = tmp_path / "reference.xlsx"
    output_path = tmp_path / "comparison.xlsx"
    write_scoresheet_pdf(ours_path, [Block("English", "Math", [[Row(1, "A", "A")], [Row(1, "F", "F")]])])
    _write_reference_scoresheet(reference_path)

    exit_code = main(
        ["--ours", str(ours_path), "--reference", str(reference_path), "--output", str(output_path)]
    )

    assert exit_code == 0
    assert output_path.exists()
