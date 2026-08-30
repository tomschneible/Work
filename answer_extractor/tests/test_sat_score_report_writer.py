import datetime as dt
from pathlib import Path

import openpyxl
import pytest
from openpyxl.styles import Border, PatternFill, Side
from openpyxl.utils.cell import column_index_from_string, coordinate_from_string

from answer_extractor.google_sheets_export import CellWrite, FillResult
from answer_extractor.sat_score_report_writer import (
    _HIDDEN_COLUMN_SHRINK_FACTOR,
    _MARK_HEADER_NON_BLANK,
    _TABLE_COLUMN_NARROW_FACTOR,
    ReferenceQuestion,
    blocks_to_clear,
    columns_to_hide,
    fill_sat_score_report,
    header_bar_extension,
    hidden_columns_to_shrink,
    locate_sat_blocks,
    read_reference_questions,
    trailing_rows_to_delete,
    visible_table_columns_to_narrow,
)

_MISSING = object()  # distinguishes "never written" from an explicitly-written None (an omitted answer)


def _at(result: FillResult, a1: str, sheet: str = "Student Responses"):
    col_letter, row = coordinate_from_string(a1)
    col = column_index_from_string(col_letter)
    for w in result.cell_writes:
        if w.sheet == sheet and w.row == row and w.column == col:
            return w.value
    return _MISSING


def _block(ws, title_col: int, title: str, flagged: bool, answers, title_row: int = 4) -> None:
    """One SAT block: a title (`title_row`), header labels (the next row:
    Correct Answer/Your Answer/Domain/Skill), and answer-key rows starting
    two rows below the title -- the shape _write_template composes into a
    full subject's worth of blocks (see there), reused directly by tests
    that only need one or two standalone blocks. `title_row` defaults to
    4 (module1's own row); pass a later row to place a second subject's
    blocks below the first's, like a real template."""
    header_row = title_row + 1
    ws.cell(row=title_row, column=title_col, value=title)
    if flagged:
        ws.cell(row=header_row, column=title_col, value=False)
    ws.cell(row=header_row, column=title_col + 1, value="Correct Answer")
    ws.cell(row=header_row, column=title_col + 2, value="Your Answer")
    ws.cell(row=header_row, column=title_col + 4, value="Domain")
    ws.cell(row=header_row, column=title_col + 5, value="Skill")
    for i, (q, correct) in enumerate(answers, start=header_row + 1):
        ws.cell(row=i, column=title_col, value=q)
        ws.cell(row=i, column=title_col + 1, value=correct)


def _write_template(path: Path) -> None:
    """A miniature version of the real DSAT templates: one subject
    (Reading and Writing), Module 1 plus two same-difficulty pairs of
    Module 2 blocks (Higher x2, Lower x2), and the single shared row of
    flag cells every subject's score formulas actually reference (see
    module docstring) -- confirmed against a real blank template where a
    *second* subject's blocks reuse the *first* subject's flag row rather
    than having their own; this fixture only needs one subject to prove
    the same mechanism, since a second subject would just repeat it.

    Each block is 6 columns wide (question, correct, your-answer, mark,
    Domain, Skill -- see _CLEAR_BLOCK_WIDTH) plus one blank spacer column
    before the next block's own title, matching a real template's spacing
    exactly (confirmed against a real filled report) -- narrower spacing
    would let blocks_to_clear's clear range bleed into the next block's
    own title column, a fixture-only collision that doesn't happen
    against the real layout this is modeling.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Student Responses"

    ws["A1"] = "Type name here, date below"
    ws["A2"] = dt.datetime(2024, 1, 1)  # dummy placeholder date, like the real templates'

    # Module 1: columns A-F (question=A, correct=B, your=C, mark=D, Domain=E,
    # Skill=F). No flag.
    _block(ws, 1, "Reading and Writing Module 1", flagged=False, answers=[(1, "A"), (2, "B")])
    # Module 2 Higher, canonical copy: columns H-M, flag at H5.
    _block(ws, 8, "R & W Module 2 - Higher Difficulty", flagged=True, answers=[(1, "C"), (2, "D")])
    # Module 2 Lower, canonical copy: columns O-T, flag at O5.
    _block(ws, 15, "R & W Module 2 - Lower Difficulty", flagged=True, answers=[(1, "F"), (2, "G")])
    # Module 2 Higher, duplicate copy (byte-identical key): columns V-AA, flag at V5.
    _block(ws, 22, "R & W Module 2 - Higher Difficulty", flagged=True, answers=[(1, "C"), (2, "D")])
    # Module 2 Lower, duplicate copy: columns AC-AH, flag at AC5.
    _block(ws, 29, "R & W Module 2 - Lower Difficulty", flagged=True, answers=[(1, "F"), (2, "G")])

    # Scaled section score: value cell sits a couple rows *above* its own
    # label, like the real templates' "Total Score"/section-score cells --
    # well clear of every block's own columns (through AI).
    ws["AN10"] = 200
    ws["AN12"] = "Reading\n& Writing\nScore"

    wb.save(str(path))


def test_locate_sat_blocks_dedupes_duplicate_pairs_to_the_leftmost(tmp_path):
    path = tmp_path / "template.xlsx"
    _write_template(path)
    ws = openpyxl.load_workbook(path)["Student Responses"]

    blocks = locate_sat_blocks(ws)
    by_key = {(b.subject, b.module_slot): b for b in blocks}

    assert set(by_key) == {
        ("reading and writing", "module1"),
        ("reading and writing", "harder"),
        ("reading and writing", "easier"),
    }
    assert by_key[("reading and writing", "harder")].question_col == 8  # H, not the V duplicate
    assert by_key[("reading and writing", "easier")].question_col == 15  # O, not the AC duplicate
    assert by_key[("reading and writing", "module1")].flag_cell is None
    assert by_key[("reading and writing", "harder")].flag_cell == (5, 8)
    assert by_key[("reading and writing", "easier")].flag_cell == (5, 15)


def test_read_reference_questions_reads_correct_answer_domain_and_skill(tmp_path):
    """_block only ever writes the Domain/Skill *header* labels, not
    per-question values (no test needed them before this) -- add a
    couple directly here rather than changing the shared fixture."""
    path = tmp_path / "template.xlsx"
    _write_template(path)
    wb = openpyxl.load_workbook(path)
    ws = wb["Student Responses"]
    ws["E6"], ws["F6"] = "Craft and Structure", "Words in Context"  # Module 1 Q1
    ws["E7"], ws["F7"] = "Information and Ideas", "Central Ideas"  # Module 1 Q2
    ws["L6"], ws["M6"] = "Expression of Ideas", "Rhetorical Synthesis"  # Harder canonical Q1
    wb.save(path)
    ws = openpyxl.load_workbook(path)["Student Responses"]

    module1 = read_reference_questions(ws, "reading and writing", "module1")
    assert module1[1] == ReferenceQuestion(
        correct_answer="A", domain="Craft and Structure", skill="Words in Context"
    )
    assert module1[2] == ReferenceQuestion(
        correct_answer="B", domain="Information and Ideas", skill="Central Ideas"
    )

    harder = read_reference_questions(ws, "reading and writing", "harder")
    assert harder[1] == ReferenceQuestion(
        correct_answer="C", domain="Expression of Ideas", skill="Rhetorical Synthesis"
    )
    # Q2 never had Domain/Skill poked above -- still reads cleanly as None,
    # not an error; only a missing *block* raises (see below).
    assert harder[2] == ReferenceQuestion(correct_answer="D", domain=None, skill=None)


def test_read_reference_questions_raises_for_a_block_that_does_not_exist(tmp_path):
    path = tmp_path / "template.xlsx"
    _write_template(path)
    ws = openpyxl.load_workbook(path)["Student Responses"]

    with pytest.raises(ValueError, match="math.*module1"):
        read_reference_questions(ws, "math", "module1")


def test_fill_sat_score_report_writes_name_date_and_active_variant_only(tmp_path):
    path = tmp_path / "template.xlsx"
    _write_template(path)

    writes = fill_sat_score_report(
        path,
        answers={
            ("reading and writing", "module1", 1): "A",
            ("reading and writing", "module1", 2): "B",
            ("reading and writing", "harder", 1): "C",
            ("reading and writing", "harder", 2): "D",
        },
        active_variants={"reading and writing": "harder"},
        student_name="Jane Student",
        test_date=dt.date(2026, 3, 8),
    )

    assert _at(writes, "A1") == "Jane Student"
    assert _at(writes, "A2") == "2026-03-08"  # ISO-formatted for the Sheets API
    # Module 1 (always active).
    assert _at(writes, "C6") == "A"
    assert _at(writes, "C7") == "B"
    # Harder, canonical (H) copy -- filled.
    assert _at(writes, "J6") == "C"
    assert _at(writes, "J7") == "D"
    assert _at(writes, "H5") is True
    # Every active block's own mark_col header cell (blank in the
    # template) gets a zero-width space so it's no longer blank -- see
    # _MARK_HEADER_NON_BLANK's own comment for why.
    assert _at(writes, "D5") == _MARK_HEADER_NON_BLANK  # Module 1's own mark_col
    assert _at(writes, "K5") == _MARK_HEADER_NON_BLANK  # canonical Module 2's own mark_col
    # Harder, duplicate (V) copy -- left completely untouched (not even a blank write).
    assert _at(writes, "X6") is _MISSING
    assert _at(writes, "X7") is _MISSING
    assert _at(writes, "V5") is _MISSING
    # Easier -- not active, left untouched, flag never written (stays whatever the template had).
    assert _at(writes, "Q6") is _MISSING
    assert _at(writes, "O5") is _MISSING
    # The report should only show the filled-in module -- every other Module 2
    # column (Easier's canonical block, plus both duplicates), title through
    # Domain/Skill, cleared for the sheet's entire height (see
    # test_blocks_to_clear_hides_every_module_2_block_but_the_canonical_one for
    # these same rectangles' own column-letter breakdown) -- and the same
    # columns hidden outright too, so their now-blank width doesn't still
    # count toward the exported PDF's own print-area sizing (see
    # test_columns_to_hide_hides_every_module_2_block_but_the_canonical_one).
    max_row = 12  # this fixture's own last row (the "AN12" score label)
    assert writes.cleared_ranges == [
        ("Student Responses", 0, max_row, 14, 20),
        ("Student Responses", 0, max_row, 21, 27),
        ("Student Responses", 0, max_row, 28, 34),
    ]
    # One single contiguous range now -- covers the spacer columns between
    # occurrences too, not just each occurrence's own 6-column block (see
    # test_columns_to_hide_hides_every_module_2_block_but_the_canonical_one).
    assert writes.hidden_column_ranges == [("Student Responses", 13, 34)]
    # Module 1's own title-through-answer (A-C) and domain-through-spacer
    # (E-G), then the canonical Module 2 block's own title-through-answer
    # (H-J) and domain-through-skill (L-M) -- mark_col (D, K) excluded
    # from each block (see
    # test_visible_table_columns_to_narrow_spans_module1_through_the_canonical_block)
    # -- followed by the same non-canonical columns hide_columns already
    # hides, narrowed to near-zero too (see
    # test_hidden_columns_to_shrink_shrinks_the_same_columns_columns_to_hide_hides).
    assert writes.narrowed_column_ranges == [
        ("Student Responses", 0, 3, _TABLE_COLUMN_NARROW_FACTOR),
        ("Student Responses", 4, 7, _TABLE_COLUMN_NARROW_FACTOR),
        ("Student Responses", 7, 10, _TABLE_COLUMN_NARROW_FACTOR),
        ("Student Responses", 11, 13, _TABLE_COLUMN_NARROW_FACTOR),
        ("Student Responses", 13, 34, _HIDDEN_COLUMN_SHRINK_FACTOR),
    ]
    # This fixture's own row 1 has no decorative fill at all -- nothing
    # for header_bar_extension to find or extend.
    assert writes.header_bar_extension == []
    # Module 1's own title cell (A4) and the canonical Module 2 block's
    # own title cell (H4, already canonical here -- no repositioning) --
    # only active blocks' titles, 0-indexed.
    assert writes.overflow_title_cells == [("Student Responses", 3, 0), ("Student Responses", 3, 7)]
    # This fixture's own last row already matches its last real content
    # (no stray trailing formatting the way a real template has -- see
    # test_trailing_rows_to_delete_finds_a_sheets_own_trailing_blank_rows),
    # so there's nothing for trailing_rows_to_delete to find here.
    assert writes.deleted_row_ranges == []


def test_blocks_to_clear_hides_every_module_2_block_but_the_canonical_one(tmp_path):
    path = tmp_path / "template.xlsx"
    _write_template(path)
    ws = openpyxl.load_workbook(path)["Student Responses"]

    # Higher's canonical (H) is always the one column every subject's real
    # answers get consolidated into (see fill_sat_score_report), regardless
    # of active_variants -- blocks_to_clear no longer takes that mapping at
    # all: it just clears every *other* Module 2 column, for the sheet's
    # entire height, unconditionally.
    ranges = blocks_to_clear(ws)

    max_row = ws.max_row
    assert ranges == [
        (0, max_row, 14, 20),  # O -- Lower, canonical
        (0, max_row, 21, 27),  # V -- Higher, duplicate
        (0, max_row, 28, 34),  # AC -- Lower, duplicate
    ]


def test_columns_to_hide_hides_every_module_2_block_but_the_canonical_one(tmp_path):
    """One contiguous range spanning every non-canonical occurrence *and*
    the spacer columns between/after them, starting right where the
    canonical block's own last column (M) ends -- i.e. at N, the spacer
    between the canonical block and the next occurrence (O) -- through
    AH, the last occurrence's own last column. Starting at the canonical
    block's own end rather than at O itself is what sweeps in that one
    spacer too, not just the ones between/after the non-canonical
    occurrences -- see columns_to_hide's own docstring for why it was
    the one column previously left out."""
    path = tmp_path / "template.xlsx"
    _write_template(path)
    ws = openpyxl.load_workbook(path)["Student Responses"]

    ranges = columns_to_hide(ws)

    assert ranges == [(13, 34)]  # N (the spacer after canonical) through AH's own end


def test_columns_to_hide_is_empty_without_any_non_canonical_occurrence():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Student Responses"

    assert columns_to_hide(ws) == []


def test_visible_table_columns_to_narrow_spans_module1_through_the_canonical_block(tmp_path):
    """Module 1's own title-through-answer (A-C) and domain-through-
    spacer (E-G), then the canonical Module 2 block's own title-through-
    answer (H-J) and domain-through-skill (L-M) -- each block's own
    mark_col (D, K) is deliberately skipped (see this function's own
    docstring for why), and the sidebar (never modeled in this fixture)
    and every non-canonical Module 2 occurrence are excluded either
    way."""
    path = tmp_path / "template.xlsx"
    _write_template(path)
    ws = openpyxl.load_workbook(path)["Student Responses"]

    assert visible_table_columns_to_narrow(ws) == [
        (0, 3, _TABLE_COLUMN_NARROW_FACTOR),
        (4, 7, _TABLE_COLUMN_NARROW_FACTOR),
        (7, 10, _TABLE_COLUMN_NARROW_FACTOR),
        (11, 13, _TABLE_COLUMN_NARROW_FACTOR),
    ]


def test_visible_table_columns_to_narrow_is_empty_without_a_module1_block():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Student Responses"

    assert visible_table_columns_to_narrow(ws) == []


def test_hidden_columns_to_shrink_shrinks_the_same_columns_columns_to_hide_hides(tmp_path):
    path = tmp_path / "template.xlsx"
    _write_template(path)
    ws = openpyxl.load_workbook(path)["Student Responses"]

    assert hidden_columns_to_shrink(ws) == [(13, 34, _HIDDEN_COLUMN_SHRINK_FACTOR)]


def test_header_bar_extension_extends_a_solid_row1_fill_to_the_narrowed_table_width(tmp_path):
    path = tmp_path / "template.xlsx"
    _write_template(path)
    wb = openpyxl.load_workbook(path)
    ws = wb["Student Responses"]

    # A real template's own accent bar: solid fill on row 1, stopping
    # short of the canonical Module 2 block (matching the real file's
    # own shape -- confirmed against it, the bar stops right at Module
    # 1's own end, not partway through it).
    for col in range(1, 8):  # A-G
        ws.cell(row=1, column=col).fill = PatternFill(fill_type="solid", fgColor="FF0497D4")
    wb.save(str(path))
    ws = openpyxl.load_workbook(path)["Student Responses"]

    assert header_bar_extension(ws) == (0, "FF0497D4", 7, 13)


def test_header_bar_extension_is_none_without_a_row1_fill(tmp_path):
    path = tmp_path / "template.xlsx"
    _write_template(path)
    ws = openpyxl.load_workbook(path)["Student Responses"]

    assert header_bar_extension(ws) is None


def test_trailing_rows_to_delete_finds_a_sheets_own_trailing_blank_rows(tmp_path):
    """The bug confirmed live against the real "Student Responses" tab: it
    carries stray formatting (border/fill, no value) all the way out to
    row 996, even though its actual content -- every block, every score
    cell, the footer -- ends at row 64. openpyxl's own `ws.max_row`
    reflects that stray formatting, not the real content boundary, which
    is exactly why trailing_rows_to_delete can't just use it directly as
    "the last row" -- it has to scan for the last row that actually holds
    a value and report everything past *that* for deletion."""
    path = tmp_path / "template.xlsx"
    _write_template(path)
    wb = openpyxl.load_workbook(path)
    ws = wb["Student Responses"]
    # Stray formatting with no real content -- the same shape the real
    # template has: a border applied to cells well past the last row
    # anything was ever written to, with no value of their own -- this is
    # what actually makes openpyxl (and Sheets' own export) treat those
    # rows as part of the sheet's used range; a bare row height alone
    # (tried first) turned out not to reproduce it.
    for row in range(13, 51):
        ws.cell(row=row, column=1).border = Border(bottom=Side(style="thin"))
    wb.save(str(path))
    ws = openpyxl.load_workbook(path)["Student Responses"]

    assert ws.max_row == 50  # confirms the fixture actually reproduces the gap
    ranges = trailing_rows_to_delete(ws)

    assert ranges == [(12, 50)]  # 0-indexed: 1-indexed rows 13-50, end exclusive


def test_trailing_rows_to_delete_is_empty_when_there_is_no_trailing_gap(tmp_path):
    path = tmp_path / "template.xlsx"
    _write_template(path)
    ws = openpyxl.load_workbook(path)["Student Responses"]

    assert trailing_rows_to_delete(ws) == []


def test_fill_sat_score_report_consolidates_a_non_canonical_active_variant(tmp_path):
    """Reading & Writing's own occurrences sit at rows 4-7, Math's own
    occurrences of the exact same columns sit at rows 24-27. R&W's active
    variant (Higher) already lives at the canonical column (H) -- written
    in place, same as always. Math's active variant (Lower) doesn't (O) --
    its title, correct-answer key, your answer, and Domain/Skill all get
    copied into Math's own rows at the *canonical* column (H) instead,
    overwriting whatever Math's own Higher occurrence held there, so both
    subjects' real Module 2 tables end up at the same column position
    instead of visibly offset from each other (confirmed live against a
    real filled report where this showed up as exactly that offset)."""
    path = tmp_path / "template.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Student Responses"
    ws["A1"] = "Type name here, date below"
    ws["A2"] = dt.datetime(2024, 1, 1)
    _block(ws, 8, "R & W Module 2 - Higher Difficulty", flagged=True, answers=[(1, "C"), (2, "D")])
    _block(ws, 15, "R & W Module 2 - Lower Difficulty", flagged=True, answers=[(1, "F"), (2, "G")])
    # Math's own occurrences of the same columns, well below R&W's -- no
    # flag cell of its own (flagged=False): it reuses R&W's shared row (see
    # module docstring), which R&W's own blocks above already established.
    _block(ws, 8, "Math Module 2 - Higher Difficulty", flagged=False, answers=[(1, "A"), (2, "B")], title_row=24)
    _block(ws, 15, "Math Module 2 - Lower Difficulty", flagged=False, answers=[(1, "D"), (2, "C")], title_row=24)
    # Domain/Skill values for Math's real (Lower, column O) pick -- what
    # should get copied over to the canonical column.
    ws["S26"], ws["T26"] = "PSD", "RPU"
    ws["S27"], ws["T27"] = "ALG", "LE1"
    wb.save(str(path))

    writes = fill_sat_score_report(
        path,
        answers={
            ("reading and writing", "harder", 1): "C",
            ("reading and writing", "harder", 2): "D",
            ("math", "easier", 1): "D",
            ("math", "easier", 2): "C",
        },
        active_variants={"reading and writing": "harder", "math": "easier"},
        student_name="Jane Student",
        test_date=dt.date(2026, 3, 8),
    )

    # R&W's own pick was already canonical -- written in place as always.
    assert _at(writes, "J6") == "C"
    assert _at(writes, "J7") == "D"
    assert _at(writes, "H5") is True
    # Both subjects' own canonical mark_col header cell (K, same column
    # position for both, different rows) gets the same zero-width-space
    # treatment -- see _MARK_HEADER_NON_BLANK's own comment.
    assert _at(writes, "K5") == _MARK_HEADER_NON_BLANK
    assert _at(writes, "K25") == _MARK_HEADER_NON_BLANK

    # Math's own pick (Lower, column O) gets copied into the canonical
    # column (H) at Math's own rows (26-27), not left at its native column.
    # Lower's own correct-answer key (from _block's answers=[(1,"D"),(2,"C")])
    # is "D"/"C" -- coincidentally the same letters the student answered
    # with, but at different destination columns (I = correct-answer, J =
    # your-answer), so this still confirms both got copied independently.
    assert _at(writes, "H24") == "Math Module 2 - Lower Difficulty"  # title moved too
    assert _at(writes, "I26") == "D"  # correct-answer key, copied from Lower's own key
    assert _at(writes, "J26") == "D"  # your answer
    assert _at(writes, "L26") == "PSD"  # Domain
    assert _at(writes, "M26") == "RPU"  # Skill
    assert _at(writes, "I27") == "C"
    assert _at(writes, "J27") == "C"
    assert _at(writes, "L27") == "ALG"
    assert _at(writes, "M27") == "LE1"

    # Both subjects' own canonical title cells get OVERFLOW_CELL forced,
    # regardless of whether repositioning happened for that one -- R&W's
    # own (H4, written in place) and Math's own (H24, repositioned here) --
    # since a real export showed *only* an untouched, non-repositioned
    # title cell could still end up truncated (a template inconsistency
    # this doesn't try to distinguish between block by block).
    assert writes.overflow_title_cells == [("Student Responses", 3, 7), ("Student Responses", 23, 7)]

    # Math's own native Lower occurrence (O, rows 24-27) is never written to
    # at all -- nothing in this whole run ever targets column O (15),
    # including its own flag cell (there's no second, subject-specific flag
    # to set -- see module docstring): it's left untouched for
    # blocks_to_clear to remove entirely.
    assert not any(w.column == 15 for w in writes.cell_writes)


def test_fill_sat_score_report_leaves_a_missing_answer_blank(tmp_path):
    path = tmp_path / "template.xlsx"
    _write_template(path)

    writes = fill_sat_score_report(
        path,
        answers={("reading and writing", "module1", 1): "A"},  # question 2 omitted
        active_variants={},
        student_name="Jane Student",
        test_date=dt.date(2026, 3, 8),
    )

    assert _at(writes, "C6") == "A"
    assert _at(writes, "C7") is None  # explicitly blanked, not just absent


def test_fill_sat_score_report_raises_for_an_answer_in_the_inactive_variant(tmp_path):
    path = tmp_path / "template.xlsx"
    _write_template(path)

    with pytest.raises(ValueError, match="active variant"):
        fill_sat_score_report(
            path,
            answers={("reading and writing", "easier", 1): "F"},  # but harder is active
            active_variants={"reading and writing": "harder"},
            student_name="Jane Student",
            test_date=dt.date(2026, 3, 8),
        )


def test_fill_sat_score_report_raises_on_unmatched_answer_key(tmp_path):
    path = tmp_path / "template.xlsx"
    _write_template(path)

    with pytest.raises(ValueError, match="math"):
        fill_sat_score_report(
            path,
            answers={("math", "module1", 1): "A"},  # no Math block in this fixture
            active_variants={},
            student_name="Jane Student",
            test_date=dt.date(2026, 3, 8),
        )


def test_find_score_value_cells_locates_the_value_above_its_label():
    path_ws = openpyxl.Workbook()
    ws = path_ws.active
    ws.title = "Student Responses"
    ws["A1"] = "Type name here, date below"
    ws["A2"] = dt.datetime(2024, 1, 1)
    ws["Z10"] = 200
    ws["Z12"] = "Reading\n& Writing\nScore"

    from answer_extractor.sat_score_report_writer import _find_score_value_cells

    assert _find_score_value_cells(ws) == {"reading and writing": (10, 26)}


def test_fill_sat_score_report_writes_a_given_section_score(tmp_path):
    path = tmp_path / "template.xlsx"
    _write_template(path)

    writes = fill_sat_score_report(
        path,
        answers={},
        active_variants={},
        student_name="Jane Student",
        test_date=dt.date(2026, 3, 8),
        section_scores={"reading and writing": 590},
    )

    assert _at(writes, "AN10") == 590


def test_fill_sat_score_report_leaves_score_cell_alone_when_not_given(tmp_path):
    path = tmp_path / "template.xlsx"
    _write_template(path)

    writes = fill_sat_score_report(
        path,
        answers={},
        active_variants={},
        student_name="Jane Student",
        test_date=dt.date(2026, 3, 8),
        # section_scores omitted entirely
    )

    assert _at(writes, "AN10") is _MISSING  # never written -- template's own default (200) stands


def test_fill_sat_score_report_raises_for_an_unrecognized_score_subject(tmp_path):
    path = tmp_path / "template.xlsx"
    _write_template(path)

    with pytest.raises(ValueError, match="science"):
        fill_sat_score_report(
            path,
            answers={},
            active_variants={},
            student_name="Jane Student",
            test_date=dt.date(2026, 3, 8),
            section_scores={"science": 500},
        )
