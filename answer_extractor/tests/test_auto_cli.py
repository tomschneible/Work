from unittest.mock import MagicMock, patch

import cv2
from openpyxl import load_workbook

from answer_extractor.auto_cli import classify_inputs, main
from answer_extractor.detect import QuestionResult
from answer_extractor.pipeline import SheetResult
from answer_extractor.score_report import ScoreReportRow
from answer_extractor.score_report_pipeline import ExportOutcome
from answer_extractor.template import Template
from tests.score_report_synth import write_score_report_pdf
from tests.synth import render_sheet


def make_bubble_template_yaml(tmp_path):
    yaml_text = """
page:
  width: 900
  height: 700
sections:
  - name: Answers
    columns:
      - {first_question: 1, last_question: 4, x_start: 150, y_start: 100, row_height: 80}
bubble_spacing_x: 60
bubble_radius: 18
choices:
  even: [A, B, C, D]
  odd: [F, G, H, J]
thresholds:
  fill_ratio_min: 0.35
  relative_margin: 0.15
"""
    path = tmp_path / "template.yaml"
    path.write_text(yaml_text)
    return path


def test_classify_inputs_splits_images_pdfs_and_score_reports(tmp_path):
    template_path = make_bubble_template_yaml(tmp_path)
    template = Template.from_yaml(template_path)

    image_path = tmp_path / "sheet.png"
    cv2.imwrite(str(image_path), render_sheet(template, {1: ["F"]}))

    score_pdf_path = tmp_path / "score.pdf"
    write_score_report_pdf(score_pdf_path, [(1, "Math", "A", "A", "Correct")])

    bubble_paths, score_rows = classify_inputs([image_path, score_pdf_path])

    assert bubble_paths == [image_path]
    assert len(score_rows) == 1
    assert score_rows[0].your_answer == "A"


def test_classify_inputs_expands_directory_with_mixed_contents(tmp_path):
    template_path = make_bubble_template_yaml(tmp_path)
    template = Template.from_yaml(template_path)

    batch = tmp_path / "batch"
    batch.mkdir()
    cv2.imwrite(str(batch / "sheet.png"), render_sheet(template, {1: ["F"]}))
    write_score_report_pdf(batch / "score.pdf", [(1, "Math", "A", "A", "Correct")])
    (batch / "notes.txt").write_text("ignore me")

    bubble_paths, score_rows = classify_inputs([batch])

    assert len(bubble_paths) == 1
    assert len(score_rows) == 1


def test_auto_cli_produces_two_tabs_for_mixed_input(tmp_path):
    template_path = make_bubble_template_yaml(tmp_path)
    template = Template.from_yaml(template_path)

    image_path = tmp_path / "sheet.png"
    cv2.imwrite(str(image_path), render_sheet(template, {1: ["F"], 2: ["A"]}))

    score_pdf_path = tmp_path / "score.pdf"
    write_score_report_pdf(
        score_pdf_path,
        [(1, "Reading and Writing", "D", "D", "Correct"), (2, "Reading and Writing", "B", "D", "Incorrect")],
    )

    output_path = tmp_path / "combined.xlsx"
    exit_code = main(
        [
            "--input",
            str(image_path),
            str(score_pdf_path),
            "--template",
            str(template_path),
            "--output",
            str(output_path),
            "--no-refresh-keys",
        ]
    )

    assert exit_code == 0
    wb = load_workbook(output_path)
    assert set(wb.sheetnames) == {"sheet", "Score Report Answers"}
    assert wb.active.title == "sheet"

    bubble_ws = wb["sheet"]
    assert [c.value for c in bubble_ws[1]] == ["Question", "Answers"]
    assert bubble_ws.cell(row=2, column=2).value == "F"  # Q1
    assert bubble_ws.cell(row=3, column=2).value == "A"  # Q2

    score_ws = wb["Score Report Answers"]
    score_rows = list(score_ws.iter_rows(min_row=2, values_only=True))
    # No matching reference key for this synthetic report -> plain "Module N" labels.
    assert score_rows == [
        ("Unknown", "Reading and Writing - Module 1", 1, "D"),
        ("Unknown", "Reading and Writing - Module 1", 2, "D"),
    ]


def test_auto_cli_handles_bubble_only_input_without_requiring_score_reports(tmp_path):
    template_path = make_bubble_template_yaml(tmp_path)
    template = Template.from_yaml(template_path)
    image_path = tmp_path / "sheet.png"
    cv2.imwrite(str(image_path), render_sheet(template, {1: ["G"]}))

    output_path = tmp_path / "out.xlsx"
    exit_code = main(
        ["--input", str(image_path), "--template", str(template_path), "--output", str(output_path)]
    )

    assert exit_code == 0
    wb = load_workbook(output_path)
    assert wb.sheetnames == ["sheet"]
    assert wb.active.title == "sheet"


def test_auto_cli_handles_score_report_only_input_without_needing_a_valid_template(tmp_path):
    score_pdf_path = tmp_path / "score.pdf"
    write_score_report_pdf(score_pdf_path, [(1, "Math", "A", "A", "Correct")])

    output_path = tmp_path / "out.xlsx"
    exit_code = main(
        [
            "--input",
            str(score_pdf_path),
            "--template",
            str(tmp_path / "nonexistent_template.yaml"),
            "--output",
            str(output_path),
            "--no-refresh-keys",
        ]
    )

    assert exit_code == 0
    wb = load_workbook(output_path)
    assert wb.sheetnames == ["Score Report Answers"]


def _act_result(label="Student, Jane 2027 ACT 25MC1 January 17 2026"):
    """An auto-detected-looking SheetResult matching one of the ACT
    templates wired to the Drive score-report path -- for tests of
    auto_cli's export-vs-combine dispatch, without needing real
    template-detection machinery or a real Drive connection."""
    questions = [QuestionResult("English", 1, "A", ["A"], {}, low_confidence=False)]
    return SheetResult(
        label=label,
        source="test",
        used_contour_alignment=False,
        questions=questions,
        template_name="act_answer_sheet",
    )


def _fake_image(tmp_path, name="sheet.png"):
    path = tmp_path / name
    path.write_bytes(b"not a real image, but scan_bubble_sheets is mocked in these tests")
    return path


def test_auto_cli_exports_an_act_sheet_to_a_report_instead_of_the_combined_xlsx(tmp_path):
    image_path = _fake_image(tmp_path)
    output_path = tmp_path / "combined.xlsx"
    report_dir = tmp_path / "reports"
    outcome = ExportOutcome(pdf_path=report_dir / "Jane Student - January 17, 2026.pdf", xlsx_path=None)

    with patch("answer_extractor.auto_cli.scan_bubble_sheets", return_value=[_act_result()]), \
         patch("answer_extractor.auto_cli.build_services", return_value=(MagicMock(), MagicMock())), \
         patch("answer_extractor.auto_cli.export_sheet_report", return_value=outcome) as export_mock:
        exit_code = main(
            [
                "--input", str(image_path),
                "--output", str(output_path),
                "--report-output-dir", str(report_dir),
            ]
        )

    assert exit_code == 0
    assert not output_path.exists()  # nothing left to combine -- no .xlsx written at all
    export_mock.assert_called_once()
    assert export_mock.call_args[0][2] == "1hzDrOzqBymstYHdTqjdLxKOmdlbKqSSt"  # default templates root
    assert export_mock.call_args[0][4] == report_dir


def test_auto_cli_respects_a_custom_templates_root_folder_id(tmp_path):
    image_path = _fake_image(tmp_path)
    outcome = ExportOutcome(pdf_path=tmp_path / "r.pdf", xlsx_path=None)

    with patch("answer_extractor.auto_cli.scan_bubble_sheets", return_value=[_act_result()]), \
         patch("answer_extractor.auto_cli.build_services", return_value=(MagicMock(), MagicMock())), \
         patch("answer_extractor.auto_cli.export_sheet_report", return_value=outcome) as export_mock:
        main(
            [
                "--input", str(image_path),
                "--output", str(tmp_path / "combined.xlsx"),
                "--report-output-dir", str(tmp_path),
                "--templates-root-folder-id", "CUSTOM_ROOT",
            ]
        )

    assert export_mock.call_args[0][2] == "CUSTOM_ROOT"


def test_auto_cli_still_combines_act_sheets_when_a_fixed_template_is_given(tmp_path):
    """--template forces one fixed template for everything, which also
    means opting out of the Sheets-report path -- even for a result whose
    template_name happens to say "act_answer_sheet"."""
    template_path = make_bubble_template_yaml(tmp_path)
    image_path = tmp_path / "sheet.png"
    cv2.imwrite(str(image_path), render_sheet(Template.from_yaml(template_path), {1: ["F"]}))
    output_path = tmp_path / "combined.xlsx"

    with patch("answer_extractor.auto_cli.export_sheet_report") as export_mock:
        exit_code = main(
            [
                "--input", str(image_path),
                "--template", str(template_path),
                "--output", str(output_path),
            ]
        )

    assert exit_code == 0
    export_mock.assert_not_called()
    wb = load_workbook(output_path)
    assert wb.sheetnames == ["sheet"]


def test_auto_cli_falls_back_to_the_combined_xlsx_when_google_auth_is_unavailable(tmp_path):
    image_path = _fake_image(tmp_path)
    output_path = tmp_path / "combined.xlsx"

    with patch("answer_extractor.auto_cli.scan_bubble_sheets", return_value=[_act_result()]), \
         patch("answer_extractor.auto_cli.build_services", side_effect=FileNotFoundError("no client secret")):
        exit_code = main(["--input", str(image_path), "--output", str(output_path)])

    assert exit_code == 0
    wb = load_workbook(output_path)
    assert wb.sheetnames == ["Student, Jane 2027 ACT 25MC1 Ja"]  # openpyxl truncates sheet titles to 31 chars


def test_auto_cli_falls_back_to_the_combined_xlsx_for_one_sheet_that_fails_to_export(tmp_path):
    image_path = _fake_image(tmp_path)
    output_path = tmp_path / "combined.xlsx"

    with patch("answer_extractor.auto_cli.scan_bubble_sheets", return_value=[_act_result()]), \
         patch("answer_extractor.auto_cli.build_services", return_value=(MagicMock(), MagicMock())), \
         patch("answer_extractor.auto_cli.export_sheet_report", side_effect=ValueError("no matching template")):
        exit_code = main(
            ["--input", str(image_path), "--output", str(output_path), "--report-output-dir", str(tmp_path)]
        )

    assert exit_code == 0
    wb = load_workbook(output_path)
    assert wb.sheetnames == ["Student, Jane 2027 ACT 25MC1 Ja"]  # openpyxl truncates sheet titles to 31 chars


def test_auto_cli_prints_a_traceback_when_a_sheet_fails_to_export(tmp_path, capsys):
    """The one-line summary alone (e.g. "The read operation timed out")
    doesn't say *which* call inside export_sheet_report failed --
    copy_template? export_xlsx? write_cells? export_pdf? The full
    traceback does, and (via mac_droplet.sh's own stderr capture) always
    lands in "Answer Extractor - Last Run Warnings.txt" even when the
    droplet's own dialog only has room for a short preview."""
    image_path = _fake_image(tmp_path)
    output_path = tmp_path / "combined.xlsx"

    with patch("answer_extractor.auto_cli.scan_bubble_sheets", return_value=[_act_result()]), \
         patch("answer_extractor.auto_cli.build_services", return_value=(MagicMock(), MagicMock())), \
         patch(
             "answer_extractor.auto_cli.export_sheet_report",
             side_effect=TimeoutError("The read operation timed out"),
         ):
        exit_code = main(
            ["--input", str(image_path), "--output", str(output_path), "--report-output-dir", str(tmp_path)]
        )

    assert exit_code == 0
    err = capsys.readouterr().err
    assert "Warning: couldn't export" in err and "The read operation timed out" in err
    assert "Traceback (most recent call last):" in err


def test_auto_cli_mixes_exported_and_combined_sheets_in_one_run(tmp_path):
    exported_result = _act_result("Student, Jane 2027 ACT 25MC1 January 17 2026")
    combined_result = SheetResult(
        label="Smith, John 2026 SAT 1234 March 2026",
        source="test",
        used_contour_alignment=False,
        questions=[QuestionResult("Math", 1, "B", ["B"], {}, low_confidence=False)],
        template_name="default_template",  # not wired to the Drive path
    )
    image_path = _fake_image(tmp_path)
    output_path = tmp_path / "combined.xlsx"
    outcome = ExportOutcome(pdf_path=tmp_path / "r.pdf", xlsx_path=None)

    with patch("answer_extractor.auto_cli.scan_bubble_sheets", return_value=[exported_result, combined_result]), \
         patch("answer_extractor.auto_cli.build_services", return_value=(MagicMock(), MagicMock())), \
         patch("answer_extractor.auto_cli.export_sheet_report", return_value=outcome) as export_mock:
        exit_code = main(
            ["--input", str(image_path), "--output", str(output_path), "--report-output-dir", str(tmp_path)]
        )

    assert exit_code == 0
    export_mock.assert_called_once()
    assert export_mock.call_args[0][3] is exported_result
    wb = load_workbook(output_path)
    assert wb.sheetnames == ["Smith, John 2026 SAT 1234 March"]  # truncated to 31 chars


def _sat_rows(source="Student, Jane 2027 DSAT 8 March 8 2026"):
    return [
        ScoreReportRow(
            source=source, module=1, question=1, section="Math", your_answer="B", module_label="Module 1"
        )
    ]


def test_auto_cli_exports_identified_sat_rows_to_a_report_instead_of_the_combined_xlsx(tmp_path):
    score_pdf_path = tmp_path / "score.pdf"
    write_score_report_pdf(score_pdf_path, [(1, "Math", "A", "A", "Correct")])
    output_path = tmp_path / "combined.xlsx"
    report_dir = tmp_path / "reports"
    sat_rows = _sat_rows()
    pdf_path = report_dir / "Jane Student - March 8, 2026.pdf"

    with patch("answer_extractor.auto_cli.classify_inputs", return_value=([], sat_rows)), \
         patch("answer_extractor.auto_cli.annotate_rows", return_value=sat_rows), \
         patch("answer_extractor.auto_cli.build_services", return_value=(MagicMock(), MagicMock())), \
         patch("answer_extractor.auto_cli.export_sat_report", return_value=pdf_path) as export_mock:
        exit_code = main(
            ["--input", str(score_pdf_path), "--output", str(output_path), "--report-output-dir", str(report_dir)]
        )

    assert exit_code == 0
    assert not output_path.exists()  # nothing left to combine
    export_mock.assert_called_once()
    assert export_mock.call_args[0][3] == sat_rows
    assert export_mock.call_args[0][4] == report_dir


def test_auto_cli_falls_back_to_the_combined_xlsx_for_a_sat_report_that_fails_to_export(tmp_path):
    score_pdf_path = tmp_path / "score.pdf"
    write_score_report_pdf(score_pdf_path, [(1, "Math", "A", "A", "Correct")])
    output_path = tmp_path / "combined.xlsx"
    sat_rows = _sat_rows()

    with patch("answer_extractor.auto_cli.classify_inputs", return_value=([], sat_rows)), \
         patch("answer_extractor.auto_cli.annotate_rows", return_value=sat_rows), \
         patch("answer_extractor.auto_cli.build_services", return_value=(MagicMock(), MagicMock())), \
         patch("answer_extractor.auto_cli.export_sat_report", side_effect=ValueError("cancelled")):
        exit_code = main(
            ["--input", str(score_pdf_path), "--output", str(output_path), "--report-output-dir", str(tmp_path)]
        )

    assert exit_code == 0
    wb = load_workbook(output_path)
    assert wb.sheetnames == ["Score Report Answers"]
    rows = list(wb["Score Report Answers"].iter_rows(min_row=2, values_only=True))
    assert rows == [("Unknown", "Math - Module 1", 1, "B")]


def test_auto_cli_prints_a_traceback_when_a_sat_report_fails_to_export(tmp_path, capsys):
    """See the matching ACT test's own docstring above -- same reasoning,
    the SAT export loop's own except block."""
    score_pdf_path = tmp_path / "score.pdf"
    write_score_report_pdf(score_pdf_path, [(1, "Math", "A", "A", "Correct")])
    output_path = tmp_path / "combined.xlsx"
    sat_rows = _sat_rows()

    with patch("answer_extractor.auto_cli.classify_inputs", return_value=([], sat_rows)), \
         patch("answer_extractor.auto_cli.annotate_rows", return_value=sat_rows), \
         patch("answer_extractor.auto_cli.build_services", return_value=(MagicMock(), MagicMock())), \
         patch(
             "answer_extractor.auto_cli.export_sat_report",
             side_effect=TimeoutError("The read operation timed out"),
         ):
        exit_code = main(
            ["--input", str(score_pdf_path), "--output", str(output_path), "--report-output-dir", str(tmp_path)]
        )

    assert exit_code == 0
    err = capsys.readouterr().err
    assert "Warning: couldn't export" in err and "The read operation timed out" in err
    assert "Traceback (most recent call last):" in err


def test_auto_cli_never_attempts_sat_export_when_identification_itself_fails(tmp_path):
    score_pdf_path = tmp_path / "score.pdf"
    write_score_report_pdf(score_pdf_path, [(1, "Math", "A", "A", "Correct")])
    output_path = tmp_path / "combined.xlsx"

    with patch("answer_extractor.auto_cli.load_answer_keys", side_effect=RuntimeError("network down")), \
         patch("answer_extractor.auto_cli.export_sat_report") as export_mock, \
         patch("answer_extractor.auto_cli.build_services") as build_mock:
        exit_code = main(["--input", str(score_pdf_path), "--output", str(output_path)])

    assert exit_code == 0
    export_mock.assert_not_called()
    build_mock.assert_not_called()  # no reason to even attempt Google auth
    wb = load_workbook(output_path)
    assert wb.sheetnames == ["Score Report Answers"]
