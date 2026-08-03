from answer_extractor.score_report import parse_score_report, parse_score_reports
from answer_extractor.score_report_export import write_score_report_xlsx
from tests.score_report_synth import write_score_report_pdf


def test_parse_single_module():
    rows_in = [
        (1, "Reading and Writing", "D", "D", "Correct"),
        (2, "Reading and Writing", "B", "D", "Incorrect"),
        (3, "Reading and Writing", "C", "C", "Correct"),
    ]

    def make(tmp_path):
        path = tmp_path / "report.pdf"
        write_score_report_pdf(path, rows_in)
        return path

    import tempfile
    from pathlib import Path

    with tempfile.TemporaryDirectory() as tmp:
        path = make(Path(tmp))
        rows = parse_score_report(path)

    assert len(rows) == 3
    assert [r.question for r in rows] == [1, 2, 3]
    assert [r.your_answer for r in rows] == ["D", "D", "C"]
    assert all(r.module == 1 for r in rows)
    assert all(r.section == "Reading and Writing" for r in rows)


def test_parse_detects_module_boundaries_when_numbering_restarts(tmp_path):
    rows_in = [
        (1, "Reading and Writing", "D", "D", "Correct"),
        (2, "Reading and Writing", "B", "B", "Correct"),
        (1, "Reading and Writing", "A", "A", "Correct"),  # module 2 restarts at 1
        (2, "Reading and Writing", "C", "D", "Incorrect"),
        (1, "Math", "B", "B", "Correct"),  # module 3
    ]
    path = tmp_path / "report.pdf"
    write_score_report_pdf(path, rows_in)

    rows = parse_score_report(path)
    assert [r.module for r in rows] == [1, 1, 2, 2, 3]
    assert [r.question for r in rows] == [1, 2, 1, 2, 1]
    assert [r.your_answer for r in rows] == ["D", "B", "A", "D", "B"]


def test_parse_handles_grid_in_numeric_and_fraction_answers(tmp_path):
    rows_in = [
        (17, "Math", ".3928, .3929, 11/28", "11/28", "Correct"),
        (19, "Math", "54", "54", "Correct"),
        (21, "Math", "79", "73", "Incorrect"),
    ]
    path = tmp_path / "report.pdf"
    write_score_report_pdf(path, rows_in)

    rows = parse_score_report(path)
    answers = {r.question: r.your_answer for r in rows}
    assert answers == {17: "11/28", 19: "54", 21: "73"}


def test_parse_spans_multiple_pages(tmp_path):
    rows_in = [(i, "Math", "A", "A", "Correct") for i in range(1, 41)]
    path = tmp_path / "report.pdf"
    write_score_report_pdf(path, rows_in)

    rows = parse_score_report(path)
    assert len(rows) == 40
    assert [r.question for r in rows] == list(range(1, 41))


def test_parse_score_reports_combines_multiple_files_with_source_labels(tmp_path):
    path_a = tmp_path / "student_a.pdf"
    path_b = tmp_path / "student_b.pdf"
    write_score_report_pdf(path_a, [(1, "Math", "A", "A", "Correct")])
    write_score_report_pdf(path_b, [(1, "Math", "B", "B", "Correct")])

    rows = parse_score_reports([path_a, path_b])
    assert len(rows) == 2
    by_source = {r.source: r.your_answer for r in rows}
    assert by_source == {"student_a": "A", "student_b": "B"}


def test_parse_score_reports_expands_directory(tmp_path):
    write_score_report_pdf(tmp_path / "one.pdf", [(1, "Math", "A", "A", "Correct")])
    write_score_report_pdf(tmp_path / "two.pdf", [(1, "Math", "B", "B", "Correct")])
    (tmp_path / "not_a_pdf.txt").write_text("ignore me")

    rows = parse_score_reports([tmp_path])
    assert len(rows) == 2


def test_write_score_report_xlsx(tmp_path):
    rows_in = [
        (1, "Reading and Writing", "D", "D", "Correct"),
        (2, "Reading and Writing", "B", "D", "Incorrect"),
    ]
    path = tmp_path / "report.pdf"
    write_score_report_pdf(path, rows_in)
    rows = parse_score_report(path)

    out_path = tmp_path / "answers.xlsx"
    write_score_report_xlsx(rows, out_path)

    from openpyxl import load_workbook

    wb = load_workbook(out_path)
    ws = wb.active
    header = [c.value for c in ws[1]]
    assert header == ["Test", "Section / Module", "Question", "Your Answer"]
    data_rows = list(ws.iter_rows(min_row=2, values_only=True))
    # No annotate_rows call here -- falls back to plain "Module N" / "Unknown".
    assert data_rows[0] == ("Unknown", "Reading and Writing - Module 1", 1, "D")
    assert data_rows[1] == ("Unknown", "Reading and Writing - Module 1", 2, "D")
