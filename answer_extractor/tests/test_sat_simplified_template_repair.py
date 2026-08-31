import openpyxl
import pytest

from answer_extractor.google_sheets_export import CellWrite
from answer_extractor.sat_simplified_template_repair import (
    repair_calculations_writes,
    repaired_formula,
)


def test_repaired_formula_drops_dead_branches_and_unwraps_the_remaining_flag():
    """Real shape from a current-format template's own "Student
    Responses" tab -- same-sheet references, no sheet qualifier."""
    formula = (
        '=COUNTIF($K$9:$K$35, "✔") + \n'
        'if($O$8=TRUE,COUNTIF($R$9:$R$35, "✔"),0) + \n'
        'if($V$8=TRUE,COUNTIF($Y$9:$Y$35, "✔"),0) +\n'
        'if($AC$8=TRUE,countif($AF$9:AF$35,"✔"),0) +\n'
        'if($AJ$8=TRUE,countif($AM$9:AM$35,"✔"),0)'
    )

    result = repaired_formula(formula)

    assert result == '=COUNTIF($K$9:$K$35, "✔")+COUNTIF($R$9:$R$35, "✔")'


def test_repaired_formula_handles_sheet_qualified_references():
    """Real shape from "Calculations" -- cross-sheet references to
    'Student Responses', with an "*wrapped*" wildcard match (Math's own
    domain/skill rows) that repaired_formula doesn't need to understand,
    just preserve untouched inside the parts it keeps."""
    formula = (
        "=countifs('Student Responses'!K41:K62,\"✔\",'Student Responses'!L41:L62,\"*ALG*\")+\n"
        "if('Student Responses'!O8=TRUE,countifs('Student Responses'!R41:R62,\"✔\","
        "'Student Responses'!S41:S62,\"*ALG*\"),0)+\n"
        "if('Student Responses'!V8=TRUE,countifs('Student Responses'!Y41:Y62,\"✔\","
        "'Student Responses'!Z41:Z62,\"*ALG*\"),0)+\n"
        "if('Student Responses'!AC8=TRUE,countifs('Student Responses'!AF41:AF62,\"✔\","
        "'Student Responses'!AG41:AG62,\"*ALG*\"),0)+\n"
        "if('Student Responses'!AJ8=TRUE,countifs('Student Responses'!AM41:AM62,\"✔\","
        "'Student Responses'!AN41:AN62,\"*ALG*\"),0)"
    )

    result = repaired_formula(formula)

    assert result == (
        "=countifs('Student Responses'!K41:K62,\"✔\",'Student Responses'!L41:L62,\"*ALG*\")+"
        "countifs('Student Responses'!R41:R62,\"✔\",'Student Responses'!S41:S62,\"*ALG*\")"
    )


def test_repaired_formula_raises_if_the_result_still_references_a_flag_cell():
    """A formula this repair doesn't actually understand (e.g. a
    genuinely different shape) should fail loudly, not silently write a
    still-broken formula into a live template."""
    with pytest.raises(ValueError, match="still references a flag cell"):
        repaired_formula('=IF($O$8=TRUE,1,2)+IF($AK$8=TRUE,3,4)')  # $AK$8 -- not one this repair knows


def test_repair_calculations_writes_finds_and_repairs_every_matching_cell(tmp_path):
    path = tmp_path / "reference.xlsx"
    wb = openpyxl.Workbook()
    sr = wb.active
    sr.title = "Student Responses"
    sr["D26"] = (
        '=COUNTIF($K$9:$K$35, "✔") + \n'
        'if($O$8=TRUE,COUNTIF($R$9:$R$35, "✔"),0) + \n'
        'if($V$8=TRUE,COUNTIF($Y$9:$Y$35, "✔"),0)'
    )
    sr["A1"] = "not a formula at all"
    sr["A2"] = "=SUM(A3:A4)"  # a real formula, but not this pattern -- must be left alone

    calc = wb.create_sheet("Calculations")
    calc["B2"] = (
        "=countifs('Student Responses'!K9:K35,\"✔\",'Student Responses'!L9:L35,\"I&I\")+\n"
        "if('Student Responses'!O8=TRUE,countifs('Student Responses'!R9:R35,\"✔\","
        "'Student Responses'!S9:S35,\"I&I\"),0)"
    )
    wb.save(str(path))

    reference_wb = openpyxl.load_workbook(path, data_only=False)
    writes = repair_calculations_writes(reference_wb)

    assert len(writes) == 2
    by_coord = {(w.sheet, w.row, w.column): w.value for w in writes}
    assert by_coord[("Student Responses", 26, 4)] == '=COUNTIF($K$9:$K$35, "✔")+COUNTIF($R$9:$R$35, "✔")'
    assert by_coord[("Calculations", 2, 2)] == (
        "=countifs('Student Responses'!K9:K35,\"✔\",'Student Responses'!L9:L35,\"I&I\")+"
        "countifs('Student Responses'!R9:R35,\"✔\",'Student Responses'!S9:S35,\"I&I\")"
    )
    # A1 (not a formula) and A2 (a real formula, but not this pattern)
    # never produced a write -- len(writes) == 2 above already confirms
    # only the two matching cells did.
