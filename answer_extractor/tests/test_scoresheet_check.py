import os
from pathlib import Path

import openpyxl
import pytest

from answer_extractor.detect import QuestionResult
from answer_extractor.export import write_xlsx
from answer_extractor.pipeline import SheetResult
from answer_extractor.scoresheet_check import (
    OurAnswer,
    compare,
    load_our_answers,
    load_reference_answers,
    ours_from_results,
    parse_program_output,
    parse_reference_scoresheet,
)
from tests.sat_scoresheet_pdf_synth import SatGroup, write_sat_scoresheet_pdf
from tests.scoresheet_pdf_synth import MARK_FONT_PATH, Block, Row, write_scoresheet_pdf


def _write_reference_scoresheet(path: Path) -> None:
    """A miniature version of the real vendor layout: two sections
    (English, Math), each with its own single "Correct Answer"/"Your
    Answer" column-group -- same title-then-header shape as the real file,
    just one block per section instead of the several side-by-side blocks
    a section with more questions needs (see
    test_parse_reference_scoresheet_assigns_every_block_under_one_title_to_the_same_section
    for that case, confirmed against a real reference file whose English
    and Math sections are each split across two blocks)."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "ScoreSheet"

    # Row 1: section titles. "English" governs columns A-F, "Math" columns H-M.
    ws["A1"] = "English"
    ws["H1"] = "Math"
    # Row 2: header pairs for each of the two column-groups.
    ws["B2"], ws["C2"] = "Correct Answer", "Your Answer"
    ws["I2"], ws["J2"] = "Correct Answer", "Your Answer"

    # English block: Q1-3, one correct/wrong/omitted.
    english_rows = [
        (1, "A", "A", "✔"),
        (2, "B", "C", "✘"),
        (3, "C", None, "ø"),
    ]
    for i, (q, correct, yours, mark) in enumerate(english_rows, start=3):
        ws.cell(row=i, column=1, value=q)
        ws.cell(row=i, column=2, value=correct)
        ws.cell(row=i, column=3, value=yours)
        ws.cell(row=i, column=4, value=mark)

    # Math block: Q1-2.
    math_rows = [
        (1, "F", "F", "✔"),
        (2, "G", "H", "✘"),
    ]
    for i, (q, correct, yours, mark) in enumerate(math_rows, start=3):
        ws.cell(row=i, column=8, value=q)
        ws.cell(row=i, column=9, value=correct)
        ws.cell(row=i, column=10, value=yours)
        ws.cell(row=i, column=11, value=mark)

    wb.save(str(path))


def test_parse_reference_scoresheet_reads_both_sections(tmp_path):
    path = tmp_path / "reference.xlsx"
    _write_reference_scoresheet(path)

    result = parse_reference_scoresheet(path)

    # "Math" in the source is normalized to "mathematics" to match our
    # template's section naming.
    assert result == {
        ("english", 1): "A",
        ("english", 2): "C",
        ("english", 3): "",  # omitted
        ("mathematics", 1): "F",
        ("mathematics", 2): "H",
    }


def test_parse_reference_scoresheet_missing_tab_raises(tmp_path):
    path = tmp_path / "reference.xlsx"
    _write_reference_scoresheet(path)

    with pytest.raises(ValueError, match="No 'Nope'"):
        parse_reference_scoresheet(path, sheet_name="Nope")


def _write_multi_block_reference_scoresheet(path: Path) -> None:
    """A section with more questions than fit in one block gets split
    across two side-by-side "Correct Answer"/"Your Answer" column-groups,
    both governed by the single title cell to their left -- confirmed
    against a real reference file (English 1-75 and Math 1-60, each split
    into two ~half-sized blocks this way) that this tool's own comparison
    read a perfect 215/215 against once the legacy 6-column bubble-sheet
    template existed to scan the matching answer sheet with. Mirrors that
    shape at a miniature scale: one "English" title governing two blocks,
    the second continuing the question numbering where the first left off."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "ScoreSheet"

    ws["A1"] = "English"
    ws["B2"], ws["C2"] = "Correct Answer", "Your Answer"  # first block
    ws["F2"], ws["G2"] = "Correct Answer", "Your Answer"  # second block, same title

    first_block = [(1, "A", "A"), (2, "B", "C")]
    for i, (q, correct, yours) in enumerate(first_block, start=3):
        ws.cell(row=i, column=1, value=q)
        ws.cell(row=i, column=2, value=correct)
        ws.cell(row=i, column=3, value=yours)

    second_block = [(3, "F", "F"), (4, "G", "G")]
    for i, (q, correct, yours) in enumerate(second_block, start=3):
        ws.cell(row=i, column=5, value=q)
        ws.cell(row=i, column=6, value=correct)
        ws.cell(row=i, column=7, value=yours)

    wb.save(str(path))


def test_parse_reference_scoresheet_assigns_every_block_under_one_title_to_the_same_section(tmp_path):
    path = tmp_path / "reference.xlsx"
    _write_multi_block_reference_scoresheet(path)

    result = parse_reference_scoresheet(path)

    assert result == {
        ("english", 1): "A",
        ("english", 2): "C",
        ("english", 3): "F",
        ("english", 4): "G",
    }


def _write_our_output(path: Path) -> None:
    questions = [
        QuestionResult("English", 1, "A", ["A"], {}, low_confidence=False),
        QuestionResult("English", 2, "B", ["B"], {}, low_confidence=False),  # will mismatch, unflagged
        QuestionResult("English", 3, "", [], {}, low_confidence=False),  # ordinary blank
        QuestionResult("Mathematics", 1, "F", ["F"], {}, low_confidence=False),
        QuestionResult(
            "Mathematics", 2, "", [], {}, low_confidence=False, unreadable=True
        ),  # flagged blank
    ]
    result = SheetResult(label="sheet1", source="test", used_contour_alignment=False, questions=questions)
    write_xlsx([result], path)


def test_parse_program_output_reads_answers_and_flags_back_off_the_cell_styling(tmp_path):
    path = tmp_path / "ours.xlsx"
    _write_our_output(path)

    result = parse_program_output(path)

    assert result[("english", 1)].answer == "A"
    assert result[("english", 1)].flag is None
    assert result[("english", 3)].answer == ""
    assert result[("english", 3)].flag == "blank"
    assert result[("mathematics", 2)].answer == ""
    assert result[("mathematics", 2)].flag == "unreadable"


def test_compare_categorizes_by_severity(tmp_path):
    ref_path = tmp_path / "reference.xlsx"
    ours_path = tmp_path / "ours.xlsx"
    _write_reference_scoresheet(ref_path)
    _write_our_output(ours_path)

    reference = parse_reference_scoresheet(ref_path)
    ours = parse_program_output(ours_path)
    rows = compare(reference, ours)
    by_key = {(r.section, r.question): r for r in rows}

    assert by_key[("english", 1)].severity == "match"
    # We read "B" where the reference says "C" -- wrong, and we never
    # flagged it, so this is the worst case.
    assert by_key[("english", 2)].severity == "silent_miss"
    # Reference also has this one omitted -- matches our ordinary blank.
    assert by_key[("english", 3)].severity == "match"
    assert by_key[("mathematics", 1)].severity == "match"
    # Reference says "H" was marked, but we flagged this one unreadable --
    # still wrong, but not a *silent* miss since it was already surfaced.
    assert by_key[("mathematics", 2)].severity == "flagged"


def test_ours_from_results_matches_parse_program_output(tmp_path):
    """The in-memory path (used by auto_compare_cli, which has the
    QuestionResult objects on hand from the same run) must classify flags
    identically to the round-trip-through-a-written-file path (used by
    compare_cli on two pre-existing spreadsheets) -- both ultimately go
    through export.flag_for, but this pins that they actually agree."""
    questions = [
        QuestionResult("English", 1, "A", ["A"], {}, low_confidence=False),
        QuestionResult("English", 2, "B", ["B"], {}, low_confidence=False),
        QuestionResult("English", 3, "", [], {}, low_confidence=False),
        QuestionResult("Mathematics", 1, "F", ["F"], {}, low_confidence=False),
        QuestionResult("Mathematics", 2, "", [], {}, low_confidence=False, unreadable=True),
    ]
    path = tmp_path / "ours.xlsx"
    write_xlsx(
        [SheetResult(label="sheet1", source="test", used_contour_alignment=False, questions=questions)],
        path,
    )

    from_memory = ours_from_results(questions)
    from_file = parse_program_output(path)

    assert from_memory == from_file


def test_ours_from_results_classifies_multiple_and_pattern_inferred():
    questions = [
        QuestionResult("English", 1, "MULTIPLE", ["A", "B"], {}, low_confidence=True),
        QuestionResult("English", 2, "C", ["C"], {}, low_confidence=True, pattern_inferred=True),
    ]
    result = ours_from_results(questions)
    assert result[("english", 1)].flag == "MULTIPLE"
    assert result[("english", 2)].flag == "pattern_inferred"


pdf_pytestmark = pytest.mark.skipif(
    not os.path.exists(MARK_FONT_PATH),
    reason=f"synthetic PDF fixtures need a Unicode-capable font at {MARK_FONT_PATH} (fonts-dejavu-core)",
)


def _write_scoresheet_pdf(path: Path) -> None:
    write_scoresheet_pdf(
        path,
        [
            Block("English", "Math", [
                [Row(1, "A", "A"), Row(2, "B", "C"), Row(3, "C", None)],
                [Row(1, "F", "F"), Row(2, "G", "H")],
            ]),
        ],
    )


@pdf_pytestmark
def test_load_reference_answers_picks_the_pdf_parser_by_extension(tmp_path):
    path = tmp_path / "reference.pdf"
    _write_scoresheet_pdf(path)

    assert load_reference_answers(path) == {
        ("english", 1): "A",
        ("english", 2): "C",
        ("english", 3): "",
        ("mathematics", 1): "F",
        ("mathematics", 2): "H",
    }


def test_load_reference_answers_still_reads_xlsx(tmp_path):
    path = tmp_path / "reference.xlsx"
    _write_reference_scoresheet(path)

    assert load_reference_answers(path) == parse_reference_scoresheet(path)


@pdf_pytestmark
def test_load_our_answers_wraps_a_pdf_as_unflagged_answers(tmp_path):
    path = tmp_path / "ours.pdf"
    _write_scoresheet_pdf(path)

    result = load_our_answers(path)

    assert result[("english", 1)] == OurAnswer(answer="A", flag=None, low_confidence=False)
    assert result[("english", 3)] == OurAnswer(answer="", flag=None, low_confidence=False)


def test_load_our_answers_still_reads_xlsx(tmp_path):
    path = tmp_path / "ours.xlsx"
    _write_our_output(path)

    assert load_our_answers(path) == parse_program_output(path)


def _write_sat_scoresheet_pdf(path: Path) -> None:
    write_sat_scoresheet_pdf(
        path,
        [
            [
                SatGroup(
                    "Reading and Writing Module 1",
                    [Row(1, "A", "A"), Row(2, "B", "C"), Row(3, "C", None)],
                ),
                SatGroup("Reading and Writing Module 2", [Row(1, "F", "F"), Row(2, "G", "H")]),
            ],
        ],
    )


@pdf_pytestmark
def test_load_reference_answers_also_picks_the_sat_pdf_parser(tmp_path):
    """load_reference_answers doesn't know in advance which PDF shape a
    given file is -- _load_pdf_answers tries the ACT reader first, then
    falls back to the SAT/DSAT one, same as _is_scoresheet_pdf does in
    auto_compare_cli.py."""
    path = tmp_path / "reference.pdf"
    _write_sat_scoresheet_pdf(path)

    assert load_reference_answers(path) == {
        ("reading and writing module 1", 1): "A",
        ("reading and writing module 1", 2): "C",
        ("reading and writing module 1", 3): "",
        ("reading and writing module 2", 1): "F",
        ("reading and writing module 2", 2): "H",
    }


@pdf_pytestmark
def test_load_our_answers_also_wraps_a_sat_pdf_as_unflagged_answers(tmp_path):
    path = tmp_path / "ours.pdf"
    _write_sat_scoresheet_pdf(path)

    result = load_our_answers(path)

    assert result[("reading and writing module 1", 1)] == OurAnswer(answer="A", flag=None, low_confidence=False)
    assert result[("reading and writing module 1", 3)] == OurAnswer(answer="", flag=None, low_confidence=False)


@pdf_pytestmark
def test_load_reference_answers_raises_naming_both_readers_when_neither_matches(tmp_path):
    import fitz

    path = tmp_path / "blank.pdf"
    doc = fitz.open()
    doc.new_page()
    doc.save(str(path))
    doc.close()

    with pytest.raises(ValueError, match="ACT ScoreSheet.*SAT/DSAT Question-Level Feedback"):
        load_reference_answers(path)


@pdf_pytestmark
def test_comparing_two_pdfs_end_to_end_reports_mismatches_as_silent_miss(tmp_path):
    """A PDF side carries no flag data, so a wrong answer against it always
    comes out as an unflagged 'silent miss' -- this is what compare_cli
    actually produces end to end when both --ours and --reference are
    PDFs (see that module's docstring)."""
    ours_path = tmp_path / "ours.pdf"
    reference_path = tmp_path / "reference.pdf"
    write_scoresheet_pdf(ours_path, [Block("English", "Math", [[Row(1, "A", "A")], [Row(1, "B", "C")]])])
    write_scoresheet_pdf(reference_path, [Block("English", "Math", [[Row(1, "A", "A")], [Row(1, "B", "B")]])])

    rows = compare(load_reference_answers(reference_path), load_our_answers(ours_path))
    by_key = {(r.section, r.question): r for r in rows}

    assert by_key[("english", 1)].severity == "match"
    assert by_key[("mathematics", 1)].severity == "silent_miss"
