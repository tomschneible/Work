import datetime as dt
from pathlib import Path
from typing import Dict, List, Tuple

import openpyxl
import pytest

from answer_extractor.google_sheets_export import CellWrite
from answer_extractor.score_report_writer import fill_score_report


def _write_template(path: Path) -> None:
    """A miniature version of the real per-test score-report templates:
    name/date placeholders, two sections (English, Math) each with their
    own question/correct-answer/your-answer block, and an unrelated
    scoring formula elsewhere on the sheet that filling in answers must
    never touch.

    The real templates' later blocks number their questions with a
    formula continuing the previous row's (e.g. '=A64+1', see
    scoresheet_grid.iter_block_questions) -- not reproduced here, since
    openpyxl never computes or caches a formula's result on save (proven
    against a real template+filled-example pair in this feature's manual
    validation instead, where those cached values already existed)."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "ScoreSheet"

    ws["A1"] = "Name:"
    ws["D1"] = "Enter Name on 'ScoreSheet' Tab"
    ws["A2"] = "Test Date:"
    ws["D2"] = "Enter Date on 'ScoreSheet' Tab"

    # English block: header at C5 ("Your Answer") -> question_col=A,
    # correct_col=B, answer_col=C. Rows 6-8, questions 1-3.
    ws["A4"] = "English"
    ws["B5"], ws["C5"] = "Correct Answer", "Your Answer"
    ws["A6"], ws["B6"] = 1, "A"
    ws["A7"], ws["B7"] = 2, "B"
    ws["A8"], ws["B8"] = 3, "C"
    ws["D8"] = '=IF(B8=C8,"match","no")'  # untouched by the writer either way

    # Math block: header at J5 ("Your Answer") -> question_col=H,
    # correct_col=I, answer_col=J. Rows 6-7, questions 1-2.
    ws["H4"] = "Math"
    ws["I5"], ws["J5"] = "Correct Answer", "Your Answer"
    ws["H6"], ws["I6"] = 1, "F"
    ws["H7"], ws["I7"] = 2, "G"

    wb.save(str(path))


def _by_cell(writes: List[CellWrite]) -> Dict[Tuple[str, int, int], object]:
    return {(w.sheet, w.row, w.column): w.value for w in writes}


def test_fill_score_report_writes_name_date_and_answers(tmp_path):
    path = tmp_path / "template.xlsx"
    _write_template(path)

    writes = fill_score_report(
        path,
        answers={
            ("english", 1): "A",
            ("english", 2): "C",
            # english 3 intentionally omitted -- left blank below.
            ("mathematics", 1): "F",
            ("mathematics", 2): "H",
        },
        student_name="Jane Student",
        test_date=dt.date(2026, 3, 4),
    )
    cells = _by_cell(writes.cell_writes)

    assert cells[("ScoreSheet", 1, 4)] == "Jane Student"  # D1
    assert cells[("ScoreSheet", 2, 4)] == "2026-03-04"  # D2, ISO-formatted for the Sheets API
    assert cells[("ScoreSheet", 6, 3)] == "A"  # C6
    assert cells[("ScoreSheet", 7, 3)] == "C"  # C7
    assert cells[("ScoreSheet", 8, 3)] is None  # C8, omitted
    assert cells[("ScoreSheet", 6, 10)] == "F"  # J6
    assert cells[("ScoreSheet", 7, 10)] == "H"  # J7
    assert writes.cleared_ranges == ()  # nothing conditionally-administered to clear, unlike SAT


def test_fill_score_report_never_writes_cells_it_should_leave_alone(tmp_path):
    """Only the name/date placeholders and each block's own answer_col
    cells are ever written -- not the unrelated scoring formula or the
    pre-baked correct-answer key, both on the same sheet."""
    path = tmp_path / "template.xlsx"
    _write_template(path)

    writes = fill_score_report(
        path,
        answers={
            ("english", 1): "A", ("english", 2): "B", ("english", 3): "C",
            ("mathematics", 1): "F", ("mathematics", 2): "G",
        },
        student_name="Jane Student",
        test_date=dt.date(2026, 3, 4),
    )
    cells = _by_cell(writes.cell_writes)

    assert ("ScoreSheet", 8, 4) not in cells  # D8, the match formula
    assert ("ScoreSheet", 6, 2) not in cells  # B6, a correct answer
    assert ("ScoreSheet", 7, 2) not in cells  # B7, a correct answer


def test_fill_score_report_raises_on_unmatched_answer_key(tmp_path):
    path = tmp_path / "template.xlsx"
    _write_template(path)

    with pytest.raises(ValueError, match="reading 1"):
        fill_score_report(
            path,
            answers={("english", 1): "A", ("reading", 1): "A"},  # no Reading block in this template
            student_name="Jane Student",
            test_date=dt.date(2026, 3, 4),
        )


def test_fill_score_report_raises_when_name_placeholder_missing(tmp_path):
    path = tmp_path / "template.xlsx"
    _write_template(path)
    wb = openpyxl.load_workbook(path)
    wb["ScoreSheet"]["D1"] = None  # simulate a template missing its marker
    wb.save(path)

    with pytest.raises(ValueError, match="enter name"):
        fill_score_report(path, answers={}, student_name="Jane Student", test_date=dt.date(2026, 3, 4))


def test_fill_score_report_does_not_mutate_the_template_file(tmp_path):
    path = tmp_path / "template.xlsx"
    _write_template(path)
    before = path.read_bytes()

    fill_score_report(
        path,
        answers={
            ("english", 1): "A", ("english", 2): "B", ("english", 3): "C",
            ("mathematics", 1): "F", ("mathematics", 2): "G",
        },
        student_name="Jane Student",
        test_date=dt.date(2026, 3, 4),
    )

    assert path.read_bytes() == before


def test_fill_score_report_passes_through_a_string_test_date_unchanged(tmp_path):
    """No day known (see scan_filename.ScanFilename.day_known) -- the
    caller passes a plain formatted string instead of a date, which must
    reach the write as-is, not run through ISO formatting."""
    path = tmp_path / "template.xlsx"
    _write_template(path)

    writes = fill_score_report(
        path, answers={}, student_name="Jane Student", test_date="January 2026"
    )
    cells = _by_cell(writes.cell_writes)

    assert cells[("ScoreSheet", 2, 4)] == "January 2026"
