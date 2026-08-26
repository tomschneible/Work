"""Orchestration tests for export_filled_report -- every step it calls is
mocked (each has its own dedicated tests: template_lookup's in
test_template_lookup.py, the Drive calls' in test_google_sheets_export.py),
so this only checks that export_filled_report calls them in the right
order, with the right arguments passed between them, calls `fill_fn` with
the downloaded local path, keeps (by default) or deletes (opted out via
keep_working_copy=False) the working Drive copy on success, always
cleans up a *failed* attempt's working copy regardless, and always
cleans up the local temp file."""
from unittest.mock import MagicMock, patch

import httplib2
import pytest
from googleapiclient.errors import HttpError

from openpyxl import Workbook, load_workbook

from answer_extractor.google_report_export_common import _tighten_print_areas, export_filled_report

_MODULE = "answer_extractor.google_report_export_common"


def _http_error(status: int) -> HttpError:
    return HttpError(httplib2.Response({"status": str(status)}), b'{"error": {"message": "nope"}}')


def _patch_all(**overrides):
    defaults = dict(
        resolve_template_folder=MagicMock(return_value="FOLDER_ID"),
        find_template_file=MagicMock(return_value={"id": "TEMPLATE_ID", "name": "ACT 25MC1"}),
        copy_template=MagicMock(return_value="COPY_ID"),
        export_xlsx=MagicMock(return_value=b"raw xlsx bytes"),
        replace_content=MagicMock(),
        export_pdf=MagicMock(return_value=b"%PDF-final"),
        delete_file=MagicMock(),
    )
    defaults.update(overrides)
    patchers = {name: patch(f"{_MODULE}.{name}", fn) for name, fn in defaults.items()}
    for p in patchers.values():
        p.start()
    return defaults, patchers


def _stop_all(patchers):
    for p in patchers.values():
        p.stop()


def test_export_filled_report_runs_every_step_in_order_and_returns_the_pdf():
    mocks, patchers = _patch_all()
    fill_fn = MagicMock(return_value=MagicMock())
    try:
        result = export_filled_report(
            drive=MagicMock(),
            templates_root_folder_id="ROOT",
            category_path=["ACT", "Enhanced"],
            test_code="25MC1",
            output_name="Jane Student - 2026-03-04",
            fill_fn=fill_fn,
        )
    finally:
        _stop_all(patchers)

    assert result == b"%PDF-final"
    mocks["resolve_template_folder"].assert_called_once()
    assert mocks["resolve_template_folder"].call_args[0][1] == "ROOT"
    assert mocks["resolve_template_folder"].call_args[0][2] == ["ACT", "Enhanced"]
    mocks["find_template_file"].assert_called_once()
    assert mocks["find_template_file"].call_args[0][2] == "25MC1"
    mocks["copy_template"].assert_called_once()
    assert mocks["copy_template"].call_args[0][1] == "TEMPLATE_ID"
    assert mocks["copy_template"].call_args[0][2] == "Jane Student - 2026-03-04"
    assert mocks["copy_template"].call_args.kwargs["parent_folder_id"] is None  # temp_folder_id omitted

    mocks["export_xlsx"].assert_called_once()
    assert mocks["export_xlsx"].call_args[0][1] == "COPY_ID"

    fill_fn.assert_called_once()  # called with the local temp path -- the fill_fn owns its own args otherwise

    mocks["replace_content"].assert_called_once()
    assert mocks["replace_content"].call_args[0][1] == "COPY_ID"

    mocks["export_pdf"].assert_called_once()
    assert mocks["export_pdf"].call_args[0][1] == "COPY_ID"


def test_export_filled_report_keeps_the_working_copy_on_success_by_default():
    """keep_working_copy defaults to True (this org's own choice, so the
    Sheet behind a generated report is available for review/editing and
    for debugging one that came out wrong) -- a successful run must not
    delete it."""
    mocks, patchers = _patch_all()
    try:
        export_filled_report(
            drive=MagicMock(),
            templates_root_folder_id="ROOT",
            category_path=["ACT", "Enhanced"],
            test_code="25MC1",
            output_name="Jane Student",
            fill_fn=MagicMock(return_value=MagicMock()),
        )
    finally:
        _stop_all(patchers)

    mocks["delete_file"].assert_not_called()


def test_export_filled_report_deletes_the_working_copy_on_success_if_asked_not_to_keep_it():
    mocks, patchers = _patch_all()
    try:
        export_filled_report(
            drive=MagicMock(),
            templates_root_folder_id="ROOT",
            category_path=["ACT", "Enhanced"],
            test_code="25MC1",
            output_name="Jane Student",
            fill_fn=MagicMock(return_value=MagicMock()),
            keep_working_copy=False,
        )
    finally:
        _stop_all(patchers)

    mocks["delete_file"].assert_called_once()
    assert mocks["delete_file"].call_args[0][1] == "COPY_ID"


def test_export_filled_report_places_the_working_copy_in_the_given_temp_folder():
    mocks, patchers = _patch_all()
    try:
        export_filled_report(
            drive=MagicMock(),
            templates_root_folder_id="ROOT",
            category_path=["SAT"],
            test_code="8",
            output_name="report",
            fill_fn=MagicMock(return_value=MagicMock()),
            temp_folder_id="TEMP_FOLDER_ID",
        )
    finally:
        _stop_all(patchers)

    assert mocks["copy_template"].call_args.kwargs["parent_folder_id"] == "TEMP_FOLDER_ID"


def test_export_filled_report_deletes_the_working_copy_even_if_a_later_step_fails():
    mocks, patchers = _patch_all(export_pdf=MagicMock(side_effect=RuntimeError("boom")))
    try:
        with pytest.raises(RuntimeError, match="boom"):
            export_filled_report(
                drive=MagicMock(),
                templates_root_folder_id="ROOT",
                category_path=["SAT"],
                test_code="1234",
                output_name="report",
                fill_fn=MagicMock(return_value=MagicMock()),
            )
    finally:
        _stop_all(patchers)

    mocks["delete_file"].assert_called_once()
    assert mocks["delete_file"].call_args[0][1] == "COPY_ID"


def test_export_filled_report_returns_the_pdf_even_if_cleanup_afterward_fails(capsys):
    """keep_working_copy=False -- the fill/export sequence itself fully
    succeeded -- a delete failure on the now-unneeded working copy must
    not throw away that already-obtained PDF (previously it did: a plain
    `finally: delete_file(...)` raising there replaced the successful
    return entirely). A non-404 delete failure (here, a permission error)
    plausibly means the copy is still sitting there, so it's still worth
    a warning."""
    mocks, patchers = _patch_all(delete_file=MagicMock(side_effect=_http_error(403)))
    try:
        result = export_filled_report(
            drive=MagicMock(),
            templates_root_folder_id="ROOT",
            category_path=["SAT"],
            test_code="1234",
            output_name="Jane Student",
            fill_fn=MagicMock(return_value=MagicMock()),
            keep_working_copy=False,
        )
    finally:
        _stop_all(patchers)

    assert result == b"%PDF-final"
    assert "couldn't clean up" in capsys.readouterr().err


def test_export_filled_report_returns_the_pdf_silently_if_the_copy_is_already_gone(capsys):
    """keep_working_copy=False -- a 404 on the cleanup delete means Drive
    no longer has the working copy at all -- already gone (by something
    else -- a retention policy, Shared Drive eventual consistency, ...),
    nothing left to clean up or warn about."""
    mocks, patchers = _patch_all(delete_file=MagicMock(side_effect=_http_error(404)))
    try:
        result = export_filled_report(
            drive=MagicMock(),
            templates_root_folder_id="ROOT",
            category_path=["SAT"],
            test_code="1234",
            output_name="Jane Student",
            fill_fn=MagicMock(return_value=MagicMock()),
            keep_working_copy=False,
        )
    finally:
        _stop_all(patchers)

    assert result == b"%PDF-final"
    assert capsys.readouterr().err == ""


def test_export_filled_report_preserves_the_real_error_when_cleanup_also_fails():
    """The fill/export sequence raised for its own reason -- a *second*,
    unrelated failure from the best-effort delete attempted afterward must
    not replace that original exception (previously it did: a plain
    `finally: delete_file(...)` that itself raises discards whatever was
    already propagating out of the try block)."""
    mocks, patchers = _patch_all(
        export_pdf=MagicMock(side_effect=RuntimeError("boom")),
        delete_file=MagicMock(side_effect=RuntimeError("404 not found")),
    )
    try:
        with pytest.raises(RuntimeError, match="boom"):
            export_filled_report(
                drive=MagicMock(),
                templates_root_folder_id="ROOT",
                category_path=["SAT"],
                test_code="1234",
                output_name="report",
                fill_fn=MagicMock(return_value=MagicMock()),
            )
    finally:
        _stop_all(patchers)

    mocks["delete_file"].assert_called_once()


def test_tighten_print_areas_bounds_an_untouched_sheet_to_its_real_content():
    """The bug confirmed live: a sheet whose real content is a small,
    bounded range (like this org's own "Cover Page" tab) but which
    carries stray row-height formatting out to row 1000 and no print
    area of its own exports with all ~1000 rows -- mostly blank,
    gridline-covered space below the real content."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Cover Page"
    ws["B7"] = "NAME:"
    ws["C7"] = "Jane Student"
    for row in range(1, 1000):
        ws.row_dimensions[row].height = 15.75  # the stray formatting itself

    _tighten_print_areas(wb)

    assert ws.print_area == "'Cover Page'!$B$7:$C$7"
    assert ws.sheet_view.showGridLines is False


def test_tighten_print_areas_leaves_a_sheet_with_its_own_print_area_alone():
    wb = Workbook()
    ws = wb.active
    ws["A1"] = "content"
    ws.print_area = "A1:Z50"  # deliberately configured some other way
    configured = ws.print_area  # openpyxl normalizes this on assignment

    _tighten_print_areas(wb)

    assert ws.print_area == configured


def test_export_filled_report_tightens_print_areas_before_uploading():
    """End to end through export_filled_report: the Workbook fill_fn hands
    back gets its print areas tightened before being saved and pushed
    back to Drive -- checked by inspecting the actual local file
    replace_content is handed, since that's the last point before the
    local temp file is cleaned up."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Cover Page"
    ws["B1"] = "untouched by fill_fn, but still on the sheet fill_fn returns"

    captured_path = {}

    def _capture_replace_content(drive, file_id, local_path):
        captured_path["path"] = local_path
        loaded = load_workbook(local_path)
        assert loaded["Cover Page"].print_area == "'Cover Page'!$B$1"
        assert loaded["Cover Page"].sheet_view.showGridLines is False

    mocks, patchers = _patch_all(replace_content=MagicMock(side_effect=_capture_replace_content))
    try:
        export_filled_report(
            drive=MagicMock(),
            templates_root_folder_id="ROOT",
            category_path=["ACT", "Enhanced"],
            test_code="25MC1",
            output_name="report",
            fill_fn=MagicMock(return_value=wb),
        )
    finally:
        _stop_all(patchers)

    assert captured_path  # the assertions inside _capture_replace_content actually ran


def test_export_filled_report_cleans_up_the_local_temp_file():
    written_paths = []
    real_open = open

    def _tracking_open(path, mode="r", *args, **kwargs):
        if "xlsx" in str(path) and "b" in mode:
            written_paths.append(str(path))
        return real_open(path, mode, *args, **kwargs)

    mocks, patchers = _patch_all()
    try:
        with patch("builtins.open", side_effect=_tracking_open):
            export_filled_report(
                drive=MagicMock(),
                templates_root_folder_id="ROOT",
                category_path=["SAT"],
                test_code="1234",
                output_name="report",
                fill_fn=MagicMock(return_value=MagicMock()),
            )
    finally:
        _stop_all(patchers)

    assert len(written_paths) == 1
    import os

    assert not os.path.exists(written_paths[0])
