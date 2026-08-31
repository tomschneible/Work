import datetime as dt
from pathlib import Path

import openpyxl
import pytest
from openpyxl.styles import Font
from openpyxl.utils.cell import column_index_from_string, coordinate_from_string

from answer_extractor.google_sheets_export import FillResult
from answer_extractor.sat_simplified_score_report_writer import (
    fill_simple_sat_score_report,
    locate_simple_sat_blocks,
)

_MISSING = object()  # distinguishes "never written" from an explicitly-written None


def _at(result: FillResult, a1: str, sheet: str = "Student Responses"):
    col_letter, row = coordinate_from_string(a1)
    col = column_index_from_string(col_letter)
    for w in result.cell_writes:
        if w.sheet == sheet and w.row == row and w.column == col:
            return w.value
    return _MISSING


def _simple_block(ws, title_col: int, title: str, questions, title_row: int = 4) -> None:
    """One block on the *simplified* template: a title with no
    difficulty suffix, header labels, and pre-numbered question rows --
    no correct-answer/Domain/Skill values (those come from reference_ws
    at fill time, never from this template itself)."""
    header_row = title_row + 1
    ws.cell(row=title_row, column=title_col, value=title)
    ws.cell(row=header_row, column=title_col + 1, value="Correct Answer")
    ws.cell(row=header_row, column=title_col + 2, value="Your Answer")
    ws.cell(row=header_row, column=title_col + 4, value="Domain")
    ws.cell(row=header_row, column=title_col + 5, value="Skill")
    for i, q in enumerate(questions, start=header_row + 1):
        ws.cell(row=i, column=title_col, value=q)


def _write_simple_template(path: Path) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Student Responses"
    ws["A1"] = "Type name here, date below"
    ws["A2"] = dt.datetime(2024, 1, 1)
    # Module 1: columns A-F. Module 2: columns H-M -- one slot, undecorated title.
    _simple_block(ws, 1, "Reading and Writing Module 1", questions=[1, 2])
    _simple_block(ws, 8, "Reading and Writing Module 2", questions=[1, 2])
    ws["AN10"] = 200
    ws["AN12"] = "Reading\n& Writing\nScore"
    score_report = wb.create_sheet("Score Report")
    score_report["V3"] = "Test:"
    score_report["X3"] = "Digital SAT (test number)"  # confirmed against a real template's own placeholder
    wb.save(str(path))


def _reference_block(ws, title_col: int, title: str, rows, title_row: int = 4) -> None:
    """One block on a *current-format* template standing in for
    reference_ws -- `rows` is (question, correct_answer, domain, skill)."""
    header_row = title_row + 1
    ws.cell(row=title_row, column=title_col, value=title)
    ws.cell(row=header_row, column=title_col + 1, value="Correct Answer")
    ws.cell(row=header_row, column=title_col + 2, value="Your Answer")
    ws.cell(row=header_row, column=title_col + 4, value="Domain")
    ws.cell(row=header_row, column=title_col + 5, value="Skill")
    for i, (q, correct, domain, skill) in enumerate(rows, start=header_row + 1):
        ws.cell(row=i, column=title_col, value=q)
        ws.cell(row=i, column=title_col + 1, value=correct)
        ws.cell(row=i, column=title_col + 4, value=domain)
        ws.cell(row=i, column=title_col + 5, value=skill)


def _write_reference_template(path: Path) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Student Responses"
    _reference_block(
        ws,
        1,
        "Reading and Writing Module 1",
        [
            (1, "A", "Craft and Structure", "Words in Context"),
            (2, "B", "Information and Ideas", "Central Ideas"),
        ],
    )
    _reference_block(
        ws,
        8,
        "R & W Module 2 - Higher Difficulty",
        [
            (1, "C", "Expression of Ideas", "Rhetorical Synthesis"),
            (2, "D", "Standard English Conventions", "Boundaries"),
        ],
    )
    wb.save(str(path))


def test_locate_simple_sat_blocks_finds_one_module1_and_one_module2(tmp_path):
    path = tmp_path / "template.xlsx"
    _write_simple_template(path)
    ws = openpyxl.load_workbook(path)["Student Responses"]

    blocks = locate_simple_sat_blocks(ws)
    by_slot = {b.module_slot: b for b in blocks}

    assert set(by_slot) == {"module1", "module2"}
    assert by_slot["module1"].subject == "reading and writing"
    assert by_slot["module1"].question_col == 1
    assert by_slot["module2"].question_col == 8
    # No difficulty is encoded in module_slot here -- unlike SatBlock,
    # there's only ever one Module 2 occurrence to find.


def test_fill_simple_sat_score_report_writes_name_date_answers_and_reference_data(tmp_path):
    template_path = tmp_path / "simple_template.xlsx"
    _write_simple_template(template_path)
    reference_path = tmp_path / "reference_template.xlsx"
    _write_reference_template(reference_path)
    reference_ws = openpyxl.load_workbook(reference_path)["Student Responses"]

    result = fill_simple_sat_score_report(
        template_path,
        reference_ws,
        answers={
            ("reading and writing", "module1", 1): "A",
            ("reading and writing", "module1", 2): "X",  # answered, but wrong
            ("reading and writing", "harder", 1): "C",
            # question 2 of the harder module has no entry -- left blank
        },
        active_variants={"reading and writing": "harder"},
        student_name="Jane Student",
        test_date=dt.date(2026, 3, 8),
        test_code="4",
        section_scores={"reading and writing": 690},
    )

    assert _at(result, "A1") == "Jane Student"
    assert _at(result, "A2") == "2026-03-08"
    # "Score Report"'s own placeholder, regenerated wholesale with the
    # actual test number substituted in -- see _find_test_number_cell's
    # own docstring for why the placeholder isn't trusted either.
    assert _at(result, "X3", sheet="Score Report") == "Digital SAT #4"

    # Module 1 -- correct answer/Domain/Skill from reference_ws, your
    # answer from `answers`.
    assert _at(result, "B6") == "A"
    assert _at(result, "C6") == "A"
    assert _at(result, "E6") == "Craft and Structure"
    assert _at(result, "F6") == "Words in Context"
    assert _at(result, "B7") == "B"
    assert _at(result, "C7") == "X"

    # Module 2 -- title gets the identified difficulty appended.
    assert _at(result, "H4") == "Reading and Writing Module 2 - Higher Difficulty"
    assert _at(result, "I6") == "C"
    assert _at(result, "J6") == "C"
    assert _at(result, "L6") == "Expression of Ideas"
    assert _at(result, "M6") == "Rhetorical Synthesis"
    # Omitted answer -- still an explicit write, just a blank one, same
    # as fill_sat_score_report's own module1 handling.
    assert _at(result, "J7") is None
    assert _at(result, "I7") == "D"  # correct-answer key still shown regardless

    assert _at(result, "AN10") == 690

    # This template never has an inactive occurrence sitting on it, so
    # none of these apply -- unlike fill_sat_score_report's own result.
    assert result.cleared_ranges == ()
    assert result.hidden_column_ranges == ()
    assert result.narrowed_column_ranges == ()
    assert result.header_bar_extension == ()
    assert result.deleted_row_ranges == ()
    # font_size_cells isn't this test's concern -- see the dedicated test
    # below (a bare openpyxl.Workbook(), as _write_reference_template
    # builds, already gives every untouched cell some non-None size of
    # its own, unlike a real Sheets-exported file's, so asserting an
    # empty result here would just be an artifact of that, not a real
    # behavior worth pinning down).


def test_fill_simple_sat_score_report_copies_the_reference_correct_columns_own_font_size(tmp_path):
    """Confirmed against a real template pair: the current-format
    template's own correct_col explicitly overrides its font size, but
    the simplified template's matching cells never got that override --
    see ReferenceQuestion's own docstring. Copied per question, onto
    correct_col specifically -- nowhere else, since nothing else was
    confirmed to differ, and only where the reference itself actually has
    an explicit size to copy."""
    template_path = tmp_path / "simple_template.xlsx"
    _write_simple_template(template_path)
    reference_path = tmp_path / "reference_template.xlsx"
    _write_reference_template(reference_path)
    wb = openpyxl.load_workbook(reference_path)
    ws = wb["Student Responses"]
    # A bare openpyxl.Workbook()'s own default style already gives every
    # cell some non-None size of its own (unlike a real Sheets-exported
    # file's, where an untouched cell's own size reads as None) -- forced
    # to None explicitly here so the untouched cells below actually
    # exercise the "no override, nothing to copy" case, not an artifact
    # of how this fixture happens to be built.
    for row in ws.iter_rows():
        for cell in row:
            cell.font = Font(size=None)
    ws["B6"].font = Font(size=12)  # Module 1 Q1's own correct_col cell
    # Q2 (B7) and Module 2's own Q1 (I6) deliberately left alone --
    # still read as unset, no write for either.
    wb.save(str(reference_path))
    reference_ws = openpyxl.load_workbook(reference_path)["Student Responses"]

    result = fill_simple_sat_score_report(
        template_path,
        reference_ws,
        answers={
            ("reading and writing", "module1", 1): "A",
            ("reading and writing", "module1", 2): "B",
            ("reading and writing", "harder", 1): "C",
        },
        active_variants={"reading and writing": "harder"},
        student_name="Jane Student",
        test_date=dt.date(2026, 3, 8),
        test_code="4",
    )

    # (sheet, 0-indexed row, 0-indexed column, font size) -- the shape
    # set_font_sizes' own Sheets API request needs. B6 -> row 5, col 1.
    assert result.font_size_cells == [("Student Responses", 5, 1, 12.0)]


def test_fill_simple_sat_score_report_regenerates_a_messy_module2_title(tmp_path):
    """Confirmed against a real template: a Module 2 title cell isn't
    reliably blank -- it can carry a placeholder ("R & W Module 2 -
    (Enter Difficulty)") or even a stale, pre-baked difficulty ("Math
    Module 2 - Higher Difficulty") left over from however the template
    was built. Both must still be found (the title match is a prefix
    match, ignoring whatever trails "Module 2") and both must come out
    with a clean, correctly-regenerated title -- not the placeholder,
    and not the stale difficulty left in place or doubled up, even when
    (Math, here) it happens to name the *wrong* difficulty for this
    particular student."""
    template_path = tmp_path / "messy_template.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Student Responses"
    ws["A1"] = "Type name here, date below"
    ws["A2"] = dt.datetime(2024, 1, 1)
    _simple_block(ws, 1, "Reading and Writing Module 1", questions=[1])
    _simple_block(ws, 8, "R & W Module 2 - (Enter Difficulty)", questions=[1])
    _simple_block(ws, 15, "Math Module 1", questions=[1])
    _simple_block(ws, 22, "Math Module 2 - Higher Difficulty", questions=[1])  # stale, and wrong for this student
    wb.create_sheet("Score Report")["X3"] = "Digital SAT (test number)"
    wb.save(str(template_path))

    reference_path = tmp_path / "reference_template.xlsx"
    wb2 = openpyxl.Workbook()
    ws2 = wb2.active
    ws2.title = "Student Responses"
    _reference_block(ws2, 1, "Reading and Writing Module 1", [(1, "A", "Craft and Structure", "Words in Context")])
    _reference_block(
        ws2, 8, "R & W Module 2 - Higher Difficulty", [(1, "C", "Expression of Ideas", "Rhetorical Synthesis")]
    )
    _reference_block(ws2, 15, "Math Module 1", [(1, "B", "Algebra", "Linear Equations")])
    _reference_block(ws2, 22, "Math Module 2 - Lower Difficulty", [(1, "D", "Problem-Solving", "Ratios")])
    wb2.save(str(reference_path))
    reference_ws = openpyxl.load_workbook(reference_path)["Student Responses"]

    result = fill_simple_sat_score_report(
        template_path,
        reference_ws,
        answers={
            ("reading and writing", "harder", 1): "C",
            ("math", "easier", 1): "B",  # Lower -- despite the template's own stale text saying Higher; wrong answer
        },
        active_variants={"reading and writing": "harder", "math": "easier"},
        student_name="Jane Student",
        test_date=dt.date(2026, 3, 8),
        test_code="4",
    )

    # Regenerated from each block's own subject text (preserving the
    # template author's own abbreviation, "R & W") plus the *actual*
    # identified difficulty -- never the placeholder, and never Math's
    # own stale "Higher" text.
    assert _at(result, "H4") == "R & W Module 2 - Higher Difficulty"
    assert _at(result, "V4") == "Math Module 2 - Lower Difficulty"
    # The right answer/reference data landed regardless of which title
    # text happened to be sitting there beforehand.
    assert _at(result, "J6") == "C"
    assert _at(result, "X6") == "B"  # this student's own (wrong) answer
    assert _at(result, "W6") == "D"  # correct-answer key for Math's actual (Lower) variant, from reference_ws


def test_fill_simple_sat_score_report_raises_for_an_inactive_variant_answer(tmp_path):
    template_path = tmp_path / "simple_template.xlsx"
    _write_simple_template(template_path)
    reference_path = tmp_path / "reference_template.xlsx"
    _write_reference_template(reference_path)
    reference_ws = openpyxl.load_workbook(reference_path)["Student Responses"]

    with pytest.raises(ValueError, match="active variant"):
        fill_simple_sat_score_report(
            template_path,
            reference_ws,
            answers={("reading and writing", "easier", 1): "A"},  # harder is active, not easier
            active_variants={"reading and writing": "harder"},
            student_name="Jane Student",
            test_date=dt.date(2026, 3, 8),
            test_code="4",
        )


def test_fill_simple_sat_score_report_raises_without_an_active_variant(tmp_path):
    template_path = tmp_path / "simple_template.xlsx"
    _write_simple_template(template_path)
    reference_path = tmp_path / "reference_template.xlsx"
    _write_reference_template(reference_path)
    reference_ws = openpyxl.load_workbook(reference_path)["Student Responses"]

    with pytest.raises(ValueError, match="[Nn]o active variant"):
        fill_simple_sat_score_report(
            template_path,
            reference_ws,
            answers={},
            active_variants={},  # Module 2 exists on the template, but no variant was identified
            student_name="Jane Student",
            test_date=dt.date(2026, 3, 8),
            test_code="4",
        )


def test_fill_simple_sat_score_report_raises_when_score_report_has_no_placeholder(tmp_path):
    template_path = tmp_path / "simple_template.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Student Responses"
    ws["A1"] = "Type name here, date below"
    ws["A2"] = dt.datetime(2024, 1, 1)
    _simple_block(ws, 1, "Reading and Writing Module 1", questions=[1])
    wb.create_sheet("Score Report")  # no "Digital SAT ..." placeholder cell at all
    wb.save(str(template_path))
    reference_path = tmp_path / "reference_template.xlsx"
    _write_reference_template(reference_path)
    reference_ws = openpyxl.load_workbook(reference_path)["Student Responses"]

    with pytest.raises(ValueError, match="Digital SAT"):
        fill_simple_sat_score_report(
            template_path,
            reference_ws,
            answers={},
            active_variants={},
            student_name="Jane Student",
            test_date=dt.date(2026, 3, 8),
            test_code="4",
        )


def test_fill_simple_sat_score_report_raises_when_reference_ws_lacks_the_block(tmp_path):
    template_path = tmp_path / "simple_template.xlsx"
    _write_simple_template(template_path)
    # An empty reference template -- no Module 1/2 blocks at all.
    reference_path = tmp_path / "reference_template.xlsx"
    wb = openpyxl.Workbook()
    wb.active.title = "Student Responses"
    wb.save(str(reference_path))
    reference_ws = openpyxl.load_workbook(reference_path)["Student Responses"]

    with pytest.raises(ValueError, match="reading and writing"):
        fill_simple_sat_score_report(
            template_path,
            reference_ws,
            answers={},
            active_variants={"reading and writing": "harder"},
            student_name="Jane Student",
            test_date=dt.date(2026, 3, 8),
            test_code="4",
        )
