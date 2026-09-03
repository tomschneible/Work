import json
import os
import time
from unittest.mock import patch

import cv2
import openpyxl
import pytest
from openpyxl import load_workbook

from answer_extractor.auto_compare_cli import _pair_with_pending_drop, main
from answer_extractor.detect import QuestionResult
from answer_extractor.export import write_xlsx
from answer_extractor.pipeline import SheetResult
from answer_extractor.template import Template
from tests.sat_scoresheet_pdf_synth import SatGroup, write_sat_scoresheet_pdf
from tests.scoresheet_pdf_synth import MARK_FONT_PATH, Block, Row, write_scoresheet_pdf
from tests.synth import render_sheet

pdf_pytestmark = pytest.mark.skipif(
    not os.path.exists(MARK_FONT_PATH),
    reason=f"synthetic PDF fixtures need a Unicode-capable font at {MARK_FONT_PATH} (fonts-dejavu-core)",
)

_MODULE = "answer_extractor.auto_compare_cli"

_TEMPLATE_YAML = """
page:
  width: 900
  height: 700
sections:
  - name: English
    columns:
      - {first_question: 1, last_question: 3, x_start: 150, y_start: 100, row_height: 80}
bubble_spacing_x: 60
bubble_radius: 18
choices:
  even: [A, B, C, D]
  odd: [F, G, H, J]
thresholds:
  fill_ratio_min: 0.35
  relative_margin: 0.15
"""


def make_template_yaml(tmp_path):
    path = tmp_path / "template.yaml"
    path.write_text(_TEMPLATE_YAML)
    return path


def write_reference(path, rows):
    """rows: list of (question, correct, your_answer, mark) for an
    "English" block -- same minimal shape as the real vendor file."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "ScoreSheet"
    ws["A1"] = "English"
    ws["B2"], ws["C2"] = "Correct Answer", "Your Answer"
    for i, (q, correct, yours, mark) in enumerate(rows, start=3):
        ws.cell(row=i, column=1, value=q)
        ws.cell(row=i, column=2, value=correct)
        ws.cell(row=i, column=3, value=yours)
        ws.cell(row=i, column=4, value=mark)
    wb.save(str(path))


def test_scan_plus_reference_adds_a_comparison_tab(tmp_path):
    template_path = make_template_yaml(tmp_path)
    template = Template.from_yaml(template_path)

    image_path = tmp_path / "sheet.png"
    cv2.imwrite(str(image_path), render_sheet(template, {1: ["F"], 2: ["A"], 3: []}))

    reference_path = tmp_path / "reference.xlsx"
    write_reference(
        reference_path,
        [
            (1, "F", "F", "✔"),  # matches our F -> match
            (2, "A", "B", "✘"),  # we read A, reference says B -> silent miss
            (3, "C", None, "ø"),  # both blank -> match
        ],
    )

    output_path = tmp_path / "out.xlsx"
    exit_code = main(
        [
            "--input",
            str(image_path),
            str(reference_path),
            "--template",
            str(template_path),
            "--output",
            str(output_path),
        ]
    )

    assert exit_code == 0
    wb = load_workbook(output_path)
    assert "Comparison" in wb.sheetnames
    assert "sheet" in wb.sheetnames  # the ordinary bubble-sheet tab is still there

    comparison_rows = list(wb["Comparison"].iter_rows(min_row=2, values_only=True))
    by_question = {row[1]: row for row in comparison_rows}
    assert by_question[1][4] == "✔"  # Match column
    assert by_question[2][4] == "✘"
    assert not by_question[2][5]  # no flag -- a silent miss
    assert by_question[3][4] == "✔"


def test_no_reference_dropped_behaves_like_a_plain_scan(tmp_path):
    template_path = make_template_yaml(tmp_path)
    template = Template.from_yaml(template_path)
    image_path = tmp_path / "sheet.png"
    cv2.imwrite(str(image_path), render_sheet(template, {1: ["F"]}))

    output_path = tmp_path / "out.xlsx"
    exit_code = main(
        ["--input", str(image_path), "--template", str(template_path), "--output", str(output_path)]
    )

    assert exit_code == 0
    wb = load_workbook(output_path)
    assert "Comparison" not in wb.sheetnames


def test_two_reference_candidates_is_an_error(tmp_path):
    template_path = make_template_yaml(tmp_path)
    template = Template.from_yaml(template_path)
    image_path = tmp_path / "sheet.png"
    cv2.imwrite(str(image_path), render_sheet(template, {1: ["F"]}))

    ref1 = tmp_path / "reference1.xlsx"
    ref2 = tmp_path / "reference2.xlsx"
    write_reference(ref1, [(1, "F", "F", "✔")])
    write_reference(ref2, [(1, "F", "F", "✔")])

    output_path = tmp_path / "out.xlsx"
    exit_code = main(
        [
            "--input",
            str(image_path),
            str(ref1),
            str(ref2),
            "--template",
            str(template_path),
            "--output",
            str(output_path),
        ]
    )

    assert exit_code == 1
    assert not output_path.exists()


def test_reference_without_a_matching_bubble_sheet_skips_comparison_but_still_writes(tmp_path):
    reference_path = tmp_path / "reference.xlsx"
    write_reference(reference_path, [(1, "F", "F", "✔")])

    template_path = make_template_yaml(tmp_path)
    template = Template.from_yaml(template_path)
    image_path_1 = tmp_path / "sheet1.png"
    image_path_2 = tmp_path / "sheet2.png"
    cv2.imwrite(str(image_path_1), render_sheet(template, {1: ["F"]}))
    cv2.imwrite(str(image_path_2), render_sheet(template, {1: ["G"]}))

    output_path = tmp_path / "out.xlsx"
    exit_code = main(
        [
            "--input",
            str(image_path_1),
            str(image_path_2),
            str(reference_path),
            "--template",
            str(template_path),
            "--output",
            str(output_path),
        ]
    )

    assert exit_code == 0
    wb = load_workbook(output_path)
    assert "Comparison" not in wb.sheetnames
    assert {"sheet1", "sheet2"} <= set(wb.sheetnames)


def _write_existing_output(path):
    """An already-exported results file, as if from an earlier run of the
    plain scan droplet -- what "compare only" mode is meant to take
    without re-scanning anything."""
    questions = [
        QuestionResult("English", 1, "F", ["F"], {}, low_confidence=False),
        QuestionResult("English", 2, "A", ["A"], {}, low_confidence=False),  # will mismatch, unflagged
        QuestionResult("English", 3, "", [], {}, low_confidence=False),
    ]
    result = SheetResult(label="prior_scan", source="test", used_contour_alignment=False, questions=questions)
    write_xlsx([result], path)


def test_existing_output_plus_reference_compares_without_rescanning(tmp_path):
    existing_path = tmp_path / "prior_scan_answers.xlsx"
    _write_existing_output(existing_path)

    reference_path = tmp_path / "reference.xlsx"
    write_reference(
        reference_path,
        [
            (1, "F", "F", "✔"),
            (2, "A", "B", "✘"),  # our stored answer was A, reference says B -> silent miss
            (3, "C", None, "ø"),
        ],
    )

    output_path = tmp_path / "out.xlsx"
    exit_code = main(["--input", str(existing_path), str(reference_path), "--output", str(output_path)])

    assert exit_code == 0
    wb = load_workbook(output_path)
    assert "Comparison" in wb.sheetnames
    assert "prior_scan" in wb.sheetnames  # the original tab is carried over untouched

    comparison_rows = list(wb["Comparison"].iter_rows(min_row=2, values_only=True))
    by_question = {row[1]: row for row in comparison_rows}
    assert by_question[1][4] == "✔"
    assert by_question[2][4] == "✘"
    assert not by_question[2][5]  # silent miss -- no flag


def test_existing_output_plus_bubble_sheet_is_ambiguous(tmp_path):
    existing_path = tmp_path / "prior_scan_answers.xlsx"
    _write_existing_output(existing_path)

    template_path = make_template_yaml(tmp_path)
    template = Template.from_yaml(template_path)
    image_path = tmp_path / "sheet.png"
    cv2.imwrite(str(image_path), render_sheet(template, {1: ["F"]}))

    reference_path = tmp_path / "reference.xlsx"
    write_reference(reference_path, [(1, "F", "F", "✔")])

    output_path = tmp_path / "out.xlsx"
    exit_code = main(
        [
            "--input",
            str(existing_path),
            str(image_path),
            str(reference_path),
            "--template",
            str(template_path),
            "--output",
            str(output_path),
        ]
    )

    assert exit_code == 1
    assert not output_path.exists()


def test_existing_output_without_a_reference_is_an_error(tmp_path):
    existing_path = tmp_path / "prior_scan_answers.xlsx"
    _write_existing_output(existing_path)

    output_path = tmp_path / "out.xlsx"
    exit_code = main(["--input", str(existing_path), "--output", str(output_path)])

    assert exit_code == 1
    assert not output_path.exists()


def _write_scoresheet_pdf(path, english_rows):
    """english_rows: list of (question, correct, your_answer) tuples."""
    write_scoresheet_pdf(
        path,
        [
            Block("English", "Math", [
                [Row(q, correct, yours) for q, correct, yours in english_rows],
                [Row(1, "F", "F")],
            ]),
        ],
    )


@pdf_pytestmark
def test_two_scoresheet_pdfs_compare_directly_first_as_ours(tmp_path):
    """No spreadsheet at all -- the two PDFs pair up directly, first given
    on the command line as "ours", second as "reference"."""
    ours_path = tmp_path / "our_report.pdf"
    reference_path = tmp_path / "their_report.pdf"
    _write_scoresheet_pdf(ours_path, [(1, "F", "F"), (2, "A", "B")])  # Q2: silent miss if B differs
    _write_scoresheet_pdf(reference_path, [(1, "F", "F"), (2, "A", "A")])

    output_path = tmp_path / "out.xlsx"
    exit_code = main(["--input", str(ours_path), str(reference_path), "--output", str(output_path)])

    assert exit_code == 0
    wb = load_workbook(output_path)
    assert wb.sheetnames == ["Comparison"]  # no scan tab -- neither side was scanned
    by_key = {(row[0], row[1]): row for row in wb["Comparison"].iter_rows(min_row=2, values_only=True)}
    assert by_key[("English", 1)][4] == "✔"
    assert by_key[("English", 2)][4] == "✘"


def _write_sat_scoresheet_pdf(path, module1_rows):
    """module1_rows: list of (question, correct, your_answer) tuples for
    Math Module 1 -- Module 2 is a single fixed row, same shape
    _write_scoresheet_pdf's own Math filler group has."""
    write_sat_scoresheet_pdf(
        path,
        [
            [
                SatGroup("Math Module 1", [Row(q, correct, yours) for q, correct, yours in module1_rows]),
                SatGroup("Math Module 2", [Row(1, "F", "F")]),
            ],
        ],
    )


@pdf_pytestmark
def test_two_sat_scoresheet_pdfs_compare_directly_first_as_ours(tmp_path):
    """The SAT/DSAT counterpart to test_two_scoresheet_pdfs_compare_directly_
    first_as_ours above -- same dispatch, a different PDF shape
    (sat_score_report_pdf_reader.py, not score_report_pdf_reader.py)."""
    ours_path = tmp_path / "our_report.pdf"
    reference_path = tmp_path / "their_report.pdf"
    _write_sat_scoresheet_pdf(ours_path, [(1, "F", "F"), (2, "A", "B")])  # Q2: silent miss if B differs
    _write_sat_scoresheet_pdf(reference_path, [(1, "F", "F"), (2, "A", "A")])

    output_path = tmp_path / "out.xlsx"
    exit_code = main(["--input", str(ours_path), str(reference_path), "--output", str(output_path)])

    assert exit_code == 0
    wb = load_workbook(output_path)
    assert wb.sheetnames == ["Comparison"]
    by_key = {(row[0], row[1]): row for row in wb["Comparison"].iter_rows(min_row=2, values_only=True)}
    assert by_key[("Math Module 1", 1)][4] == "✔"
    assert by_key[("Math Module 1", 2)][4] == "✘"


@pdf_pytestmark
def test_an_act_pdf_and_a_sat_pdf_can_still_be_compared_against_each_other(tmp_path):
    """Neither PDF reader knows about the other's shape, but
    _is_scoresheet_pdf/_load_pdf_answers try both readers independently
    per file -- so a mismatched pair (comparing the wrong two reports
    together, which is a user mistake this tool has no way to detect) is
    still accepted and compared, not rejected for being "different
    shapes". Confirmed here with an English/Math ACT report against a
    Math-only SAT report -- nothing overlaps, so nothing "matches", but
    not every row comes back the same severity: "ours" (ACT)'s own keys,
    absent from the SAT reference entirely, are "unmatched" ("?"); the
    SAT reference's own keys, which ACT has no answer for at all (not
    even a wrong one), come back "flagged" ("✘") instead -- same
    compare() behavior a same-shape pair would get for a question the
    reference names that "ours" is silently missing altogether."""
    act_path = tmp_path / "act_report.pdf"
    sat_path = tmp_path / "sat_report.pdf"
    _write_scoresheet_pdf(act_path, [(1, "F", "F")])
    _write_sat_scoresheet_pdf(sat_path, [(1, "A", "A")])

    output_path = tmp_path / "out.xlsx"
    exit_code = main(["--input", str(act_path), str(sat_path), "--output", str(output_path)])

    assert exit_code == 0
    wb = load_workbook(output_path)
    by_section = {row[0]: row[4] for row in wb["Comparison"].iter_rows(min_row=2, values_only=True)}
    assert by_section["English"] == "?" and by_section["Mathematics"] == "?"  # ACT's own -- not in the SAT reference
    assert by_section["Math Module 1"] == "✘" and by_section["Math Module 2"] == "✘"  # SAT's -- "ours" has nothing


def test_pair_with_pending_drop_records_the_first_file_and_returns_none(tmp_path):
    marker = tmp_path / "pending.json"
    with patch(f"{_MODULE}._PENDING_COMPARE_MARKER", marker):
        result = _pair_with_pending_drop(tmp_path / "a.pdf")

    assert result is None
    assert json.loads(marker.read_text())["path"] == str(tmp_path / "a.pdf")


def test_pair_with_pending_drop_pairs_a_second_file_with_a_fresh_marker(tmp_path):
    marker = tmp_path / "pending.json"
    with patch(f"{_MODULE}._PENDING_COMPARE_MARKER", marker):
        _pair_with_pending_drop(tmp_path / "a.pdf")
        result = _pair_with_pending_drop(tmp_path / "b.pdf")

    assert result == tmp_path / "a.pdf"
    assert not marker.exists()  # cleared once paired -- doesn't linger to pair a third file too


def test_pair_with_pending_drop_ignores_a_stale_marker(tmp_path):
    from answer_extractor.auto_compare_cli import _PENDING_COMPARE_TIMEOUT_SECONDS

    marker = tmp_path / "pending.json"
    marker.write_text(
        json.dumps({"path": str(tmp_path / "a.pdf"), "timestamp": time.time() - _PENDING_COMPARE_TIMEOUT_SECONDS - 1})
    )

    with patch(f"{_MODULE}._PENDING_COMPARE_MARKER", marker):
        result = _pair_with_pending_drop(tmp_path / "b.pdf")

    assert result is None  # too old to still be "waiting" -- treated as a new first file instead
    assert json.loads(marker.read_text())["path"] == str(tmp_path / "b.pdf")


def test_pair_with_pending_drop_does_not_pair_a_file_with_itself(tmp_path):
    marker = tmp_path / "pending.json"
    with patch(f"{_MODULE}._PENDING_COMPARE_MARKER", marker):
        _pair_with_pending_drop(tmp_path / "a.pdf")
        result = _pair_with_pending_drop(tmp_path / "a.pdf")  # same path again -- e.g. a retried launch

    assert result is None


def test_pair_with_pending_drop_treats_a_corrupt_marker_as_no_marker(tmp_path):
    marker = tmp_path / "pending.json"
    marker.write_text("not valid json at all")

    with patch(f"{_MODULE}._PENDING_COMPARE_MARKER", marker):
        result = _pair_with_pending_drop(tmp_path / "b.pdf")

    assert result is None
    assert json.loads(marker.read_text())["path"] == str(tmp_path / "b.pdf")


@pdf_pytestmark
def test_a_lone_dropped_pdf_waits_for_its_pending_partner_instead_of_failing(tmp_path):
    """See auto_compare_cli.py's own module docstring on why: on a Mac
    where Finder/Automator splits a two-file drop into two separate
    launches of this same script (one file each), the first of the two
    used to be a hard failure ("found a ScoreSheet-shaped PDF but nothing
    to compare it against"). Now it just waits for its pair."""
    lone_path = tmp_path / "lone_report.pdf"
    _write_scoresheet_pdf(lone_path, [(1, "F", "F")])
    output_path = tmp_path / "out.xlsx"
    marker = tmp_path / "pending.json"

    with patch(f"{_MODULE}._PENDING_COMPARE_MARKER", marker):
        exit_code = main(["--input", str(lone_path), "--output", str(output_path)])

    assert exit_code == 0  # not a failure -- this is expected, ordinary progress
    assert not output_path.exists()  # nothing to compare yet -- nothing written
    assert marker.exists()


@pdf_pytestmark
def test_a_second_lone_pdf_dropped_shortly_after_completes_the_comparison(tmp_path):
    """The first file dropped (recorded, then paired) plays "ours"; the
    second (the one that completes the pair) plays "reference" -- same
    "first given is ours" convention a same-drop pair already uses."""
    ours_path = tmp_path / "our_report.pdf"
    reference_path = tmp_path / "their_report.pdf"
    _write_scoresheet_pdf(ours_path, [(1, "F", "F"), (2, "A", "B")])  # Q2: silent miss if B differs
    _write_scoresheet_pdf(reference_path, [(1, "F", "F"), (2, "A", "A")])
    output_path = tmp_path / "out.xlsx"
    marker = tmp_path / "pending.json"

    with patch(f"{_MODULE}._PENDING_COMPARE_MARKER", marker):
        first_exit = main(["--input", str(ours_path), "--output", str(output_path)])
        assert first_exit == 0
        assert not output_path.exists()

        second_exit = main(["--input", str(reference_path), "--output", str(output_path)])

    assert second_exit == 0
    assert not marker.exists()
    wb = load_workbook(output_path)
    by_key = {(row[0], row[1]): row for row in wb["Comparison"].iter_rows(min_row=2, values_only=True)}
    assert by_key[("English", 1)][4] == "✔"
    assert by_key[("English", 2)][4] == "✘"


@pdf_pytestmark
def test_a_stale_pending_partner_is_not_reused_for_a_new_drop(tmp_path):
    from answer_extractor.auto_compare_cli import _PENDING_COMPARE_TIMEOUT_SECONDS

    old_path = tmp_path / "old_report.pdf"
    _write_scoresheet_pdf(old_path, [(1, "F", "F")])
    new_path = tmp_path / "new_report.pdf"
    _write_scoresheet_pdf(new_path, [(1, "F", "F")])
    output_path = tmp_path / "out.xlsx"
    marker = tmp_path / "pending.json"
    marker.write_text(
        json.dumps({"path": str(old_path), "timestamp": time.time() - _PENDING_COMPARE_TIMEOUT_SECONDS - 1})
    )

    with patch(f"{_MODULE}._PENDING_COMPARE_MARKER", marker):
        exit_code = main(["--input", str(new_path), "--output", str(output_path)])

    assert exit_code == 0
    assert not output_path.exists()  # too stale to pair with -- treated as a fresh first file instead
    assert json.loads(marker.read_text())["path"] == str(new_path)


@pdf_pytestmark
def test_existing_output_plus_scoresheet_pdf_compares_without_rescanning(tmp_path):
    """A pre-existing export (always "ours" -- same convention as the
    xlsx-vs-xlsx compare-only mode) paired with a lone ScoreSheet PDF,
    which then plays the reference role by elimination."""
    existing_path = tmp_path / "prior_scan_answers.xlsx"
    _write_existing_output(existing_path)  # English 1=F, 2=A, 3=blank

    reference_path = tmp_path / "reference.pdf"
    _write_scoresheet_pdf(reference_path, [(1, "F", "F"), (2, "A", "B"), (3, "C", None)])

    output_path = tmp_path / "out.xlsx"
    exit_code = main(["--input", str(existing_path), str(reference_path), "--output", str(output_path)])

    assert exit_code == 0
    wb = load_workbook(output_path)
    assert "Comparison" in wb.sheetnames
    assert "prior_scan" in wb.sheetnames  # the original tab is carried over untouched
    by_key = {(row[0], row[1]): row for row in wb["Comparison"].iter_rows(min_row=2, values_only=True)}
    assert by_key[("English", 1)][4] == "✔"
    assert by_key[("English", 2)][4] == "✘"


@pdf_pytestmark
def test_tagged_reference_xlsx_plus_scoresheet_pdf_uses_the_pdf_as_ours(tmp_path):
    """The tagged spreadsheet always plays reference; a lone PDF alongside
    it plays ours by elimination -- no scanning involved."""
    ours_path = tmp_path / "our_report.pdf"
    _write_scoresheet_pdf(ours_path, [(1, "F", "F"), (2, "A", "B")])

    reference_path = tmp_path / "reference.xlsx"
    write_reference(reference_path, [(1, "F", "F", "✔"), (2, "A", "A", "✔")])

    output_path = tmp_path / "out.xlsx"
    exit_code = main(["--input", str(ours_path), str(reference_path), "--output", str(output_path)])

    assert exit_code == 0
    wb = load_workbook(output_path)
    by_key = {(row[0], row[1]): row for row in wb["Comparison"].iter_rows(min_row=2, values_only=True)}
    assert by_key[("English", 1)][4] == "✔"
    assert by_key[("English", 2)][4] == "✘"


@pdf_pytestmark
def test_scan_plus_scoresheet_pdf_reference_adds_a_comparison_tab(tmp_path):
    """A lone ScoreSheet-shaped PDF can serve as the reference for an
    actual scan too, same as a tagged spreadsheet always could."""
    template_path = make_template_yaml(tmp_path)
    template = Template.from_yaml(template_path)
    image_path = tmp_path / "sheet.png"
    cv2.imwrite(str(image_path), render_sheet(template, {1: ["F"], 2: ["A"], 3: []}))

    reference_path = tmp_path / "reference.pdf"
    _write_scoresheet_pdf(reference_path, [(1, "F", "F"), (2, "A", "B"), (3, "C", None)])

    output_path = tmp_path / "out.xlsx"
    exit_code = main(
        [
            "--input",
            str(image_path),
            str(reference_path),
            "--template",
            str(template_path),
            "--output",
            str(output_path),
        ]
    )

    assert exit_code == 0
    wb = load_workbook(output_path)
    assert "Comparison" in wb.sheetnames
    assert "sheet" in wb.sheetnames


@pdf_pytestmark
def test_too_many_comparable_candidates_is_an_error(tmp_path):
    pdf_a = tmp_path / "a.pdf"
    pdf_b = tmp_path / "b.pdf"
    _write_scoresheet_pdf(pdf_a, [(1, "F", "F")])
    _write_scoresheet_pdf(pdf_b, [(1, "F", "F")])
    reference_path = tmp_path / "reference.xlsx"
    write_reference(reference_path, [(1, "F", "F", "✔")])

    output_path = tmp_path / "out.xlsx"
    exit_code = main(
        ["--input", str(pdf_a), str(pdf_b), str(reference_path), "--output", str(output_path)]
    )

    assert exit_code == 1
    assert not output_path.exists()


@pdf_pytestmark
def test_lone_scoresheet_pdf_with_nothing_else_now_waits_instead_of_erroring(tmp_path):
    """Superseded by the pending-pair mechanism (see
    test_a_lone_dropped_pdf_waits_for_its_pending_partner_instead_of_failing
    for the full behavior with a properly isolated marker) -- kept here,
    patching the marker the same way, specifically to document that this
    used to be a hard error and deliberately no longer is."""
    pdf_path = tmp_path / "reference.pdf"
    _write_scoresheet_pdf(pdf_path, [(1, "F", "F")])

    output_path = tmp_path / "out.xlsx"
    with patch(f"{_MODULE}._PENDING_COMPARE_MARKER", tmp_path / "pending.json"):
        exit_code = main(["--input", str(pdf_path), "--output", str(output_path)])

    assert exit_code == 0
    assert not output_path.exists()


@pdf_pytestmark
def test_comparison_pair_plus_something_to_scan_is_ambiguous(tmp_path):
    ours_path = tmp_path / "our_report.pdf"
    reference_path = tmp_path / "their_report.pdf"
    _write_scoresheet_pdf(ours_path, [(1, "F", "F")])
    _write_scoresheet_pdf(reference_path, [(1, "F", "F")])

    template_path = make_template_yaml(tmp_path)
    template = Template.from_yaml(template_path)
    image_path = tmp_path / "sheet.png"
    cv2.imwrite(str(image_path), render_sheet(template, {1: ["F"]}))

    output_path = tmp_path / "out.xlsx"
    exit_code = main(
        [
            "--input",
            str(ours_path),
            str(reference_path),
            str(image_path),
            "--template",
            str(template_path),
            "--output",
            str(output_path),
        ]
    )

    assert exit_code == 1
    assert not output_path.exists()
