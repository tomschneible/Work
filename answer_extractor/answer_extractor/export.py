"""Write sheet results to an .xlsx spreadsheet.

Layout: one tab per scanned sheet, each with a "Question" column and one
column per section (e.g. English, Mathematics, Reading, Science) --
matching the sheet's own layout rather than one column per individual
question. The first sheet's tab is the workbook's active tab, so the
output lands there when opened.
"""
from __future__ import annotations

import re
from pathlib import Path
from typing import Dict, List

from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from .pipeline import SheetResult

BLANK_FILL = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")
MULTIPLE_FILL = PatternFill(start_color="F8D7DA", end_color="F8D7DA", fill_type="solid")
LOW_CONFIDENCE_FONT = Font(italic=True, color="808080")

_INVALID_SHEET_NAME_CHARS = re.compile(r"[:\\/?*\[\]]")


def _section_order(results: List[SheetResult]) -> List[str]:
    """Section names in first-seen order (i.e. the template's declared
    order), taking the union across all sheets in case they were scored
    against slightly different templates."""
    seen: List[str] = []
    for result in results:
        for q in result.questions:
            if q.section not in seen:
                seen.append(q.section)
    return seen


def _safe_sheet_title(label: str, used: Dict[str, int]) -> str:
    title = _INVALID_SHEET_NAME_CHARS.sub("_", label)[:31] or "Sheet"
    if title not in used:
        used[title] = 0
        return title
    used[title] += 1
    suffix = f" ({used[title]})"
    return title[: 31 - len(suffix)] + suffix


def _write_sheet_tab(ws: Worksheet, result: SheetResult, sections: List[str]) -> None:
    header = ["Question"] + sections
    ws.append(header)
    for cell in ws[1]:
        cell.font = Font(bold=True)

    by_section: Dict[str, Dict[int, "object"]] = {name: {} for name in sections}
    for q in result.questions:
        by_section[q.section][q.question] = q

    max_questions = max((max(qs.keys(), default=0) for qs in by_section.values()), default=0)

    for row_num in range(1, max_questions + 1):
        row = [row_num]
        for name in sections:
            q = by_section[name].get(row_num)
            row.append(q.answer if q else "")
        ws.append(row)

        row_index = ws.max_row
        for col_offset, name in enumerate(sections, start=2):
            q = by_section[name].get(row_num)
            if q is None:
                continue
            cell = ws.cell(row=row_index, column=col_offset)
            if q.answer == "MULTIPLE":
                cell.fill = MULTIPLE_FILL
                cell.comment = Comment(", ".join(q.candidates), "answer_extractor")
            elif q.answer == "":
                cell.fill = BLANK_FILL
            if q.low_confidence:
                cell.font = LOW_CONFIDENCE_FONT

    ws.column_dimensions["A"].width = 10
    for col_index in range(2, len(header) + 1):
        ws.column_dimensions[get_column_letter(col_index)].width = 14


def add_bubble_sheet_answers_sheet(wb: Workbook, results: List[SheetResult], title: str = "") -> None:
    """Add one tab per scanned sheet to an existing workbook (used both
    standalone by write_xlsx and alongside a score-report sheet when a
    batch mixes both input types). The first sheet's tab becomes the
    workbook's active tab.

    `title` is accepted for backwards compatibility but no longer names a
    single combined tab, since each sheet now gets its own tab.
    """
    if not results:
        raise ValueError("No results to export")

    sections = _section_order(results)

    used_titles: Dict[str, int] = {}
    sheet_titles = [_safe_sheet_title(r.label, used_titles) for r in results]

    first_index = len(wb.worksheets)
    for result, tab_title in zip(results, sheet_titles):
        ws = wb.create_sheet(title=tab_title)
        _write_sheet_tab(ws, result, sections)

    wb.active = first_index


def write_xlsx(results: List[SheetResult], output_path: str | Path) -> None:
    wb = Workbook()
    del wb["Sheet"]  # remove the default blank sheet; add_bubble_sheet_answers_sheet creates its own
    add_bubble_sheet_answers_sheet(wb, results)
    wb.save(str(output_path))
