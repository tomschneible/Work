"""Figure out which cells of a per-test score-report template's
"ScoreSheet" tab need one student's name, test date, and answers -- and
exactly what to write into them.

The templates themselves (Enhanced ACT, Legacy ACT, SAT -- one file per
specific test administration, e.g. "25MC1") live in the org's Drive and
are duplicated per report rather than edited in place; this module only
figures out where a copy's fields go. It never touches the correct
answers, categories, or any of the template's scoring formulas
(ScoreReport tab's VLOOKUPs, the per-row "match" formulas, etc.) -- those
are already baked into each per-test template and are computed by
whatever eventually opens the sheet (Excel, Google Sheets), not by us.

Only three things ever get written, and all three are located generically
rather than by hardcoded cell reference, so this keeps working against
however many differently-sized sub-templates the org uploads without the
program needing to know about any of them individually:

  - The student's name and test date, via the same "Enter Name/Date on
    'ScoreSheet' Tab" placeholder text every template we've seen uses as
    its own fill-in-here marker.
  - Each question's "Your Answer" cell, located via scoresheet_grid's
    generic block scan -- the same one scoresheet_check.py's reader uses,
    just writing instead of reading.

This scans a *local, read-only* copy of the template (downloaded once via
google_sheets_export.export_xlsx) purely to find where things go --
unlike an earlier version of this module, it no longer edits or returns a
Workbook to be saved and re-uploaded wholesale. See
google_sheets_export.py's own module docstring for why: re-importing an
openpyxl-authored .xlsx turned out not to reconstruct some of a
template's own formatting (a merged, centered title cell on a tab this
module never even touches) with full fidelity. fill_score_report instead
returns the list of individual CellWrite values a caller pushes into the
live Sheet via google_sheets_export.write_cells, touching nothing else on
the file at all.
"""
from __future__ import annotations

import datetime as dt
from pathlib import Path
from typing import List, Mapping

import openpyxl
from openpyxl.worksheet.worksheet import Worksheet

from .google_sheets_export import CellWrite, format_date_for_sheets
from .scoresheet_grid import QuestionKey, iter_block_questions, locate_answer_blocks

_NAME_PLACEHOLDER_PREFIX = "enter name"
_DATE_PLACEHOLDER_PREFIX = "enter date"


def _find_placeholder_cell(ws: Worksheet, prefix: str) -> tuple[int, int]:
    for row in ws.iter_rows():
        for cell in row:
            value = cell.value
            if isinstance(value, str) and value.strip().lower().startswith(prefix):
                return cell.row, cell.column
    raise ValueError(
        f"Could not find a cell starting with {prefix!r} on {ws.title!r} -- "
        "expected this template to have an 'Enter Name/Date on ScoreSheet Tab' "
        "placeholder marking where to fill these in."
    )


def fill_score_report(
    template_path: str | Path,
    answers: Mapping[QuestionKey, str],
    student_name: str,
    test_date: dt.date | str,
    sheet_name: str = "ScoreSheet",
) -> List[CellWrite]:
    """Return every cell write needed to fill `template_path`'s
    `sheet_name` tab in with the student's name, test date, and every
    answer in `answers`. `answers` maps (normalized_section,
    question_number) -> answer letter ("" for omitted), the same shape
    scoresheet_check.parse_reference_scoresheet and
    .scoresheet_grid.normalize_section produce -- feed it this
    pipeline's own per-question results run through the same section
    normalization.

    `test_date` takes a plain `str` as well as a `date`/`datetime`
    deliberately: when the source scan's filename only gave a month/year
    (see scan_filename.ScanFilename.day_known), pass its
    formatted_test_date string ("January 2026") instead of manufacturing
    a specific day nothing in the input confirmed -- a real `date` object
    is for when the day is genuinely known. Either way, see
    google_sheets_export.format_date_for_sheets for how it becomes a
    single cell value.

    A template question with no entry in `answers` is left blank (an
    omitted bubble is a legitimate outcome). An `answers` entry for a
    (section, question) this template has no block for is a real
    mismatch worth failing loudly over -- most likely the wrong template
    was picked for this scan -- so this raises ValueError listing exactly
    which keys went unmatched, rather than silently dropping them.
    """
    wb = openpyxl.load_workbook(template_path, data_only=True)
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"No {sheet_name!r} tab in {template_path} (tabs: {wb.sheetnames})")
    ws = wb[sheet_name]

    name_row, name_col = _find_placeholder_cell(ws, _NAME_PLACEHOLDER_PREFIX)
    date_row, date_col = _find_placeholder_cell(ws, _DATE_PLACEHOLDER_PREFIX)
    writes = [
        CellWrite(sheet_name, name_row, name_col, student_name),
        CellWrite(sheet_name, date_row, date_col, format_date_for_sheets(test_date)),
    ]

    remaining = dict(answers)
    for block in locate_answer_blocks(ws):
        for r, question in iter_block_questions(ws, block):
            key = (block.section, question)
            writes.append(CellWrite(sheet_name, r, block.answer_col, remaining.pop(key, None)))

    if remaining:
        unmatched = ", ".join(f"{section} {question}" for section, question in sorted(remaining))
        raise ValueError(
            f"{template_path}!{sheet_name} has no answer block for: {unmatched} -- "
            "likely the wrong template for this scan."
        )

    return writes
