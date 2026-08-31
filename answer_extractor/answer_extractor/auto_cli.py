"""Combined entry point: auto-detects whether each dropped file is a
scanned bubble sheet or a text-based score-report PDF, and routes it to
the matching pipeline. Both kinds can be mixed in the same run; results
land as separate tabs in one spreadsheet. This is what the macOS droplet
(scripts/mac_droplet.sh) calls -- use answer_extractor.cli or
answer_extractor.score_report_cli directly if you only ever have one kind
of input and want a plain, single-purpose CLI.

Detection: images are always treated as bubble sheets (no text layer to
inspect). PDFs are routed by whether they actually parse as a score
report -- a real score-report PDF always yields at least one answer row
via answer_extractor.score_report.parse_score_report; a scanned/vector-print
bubble sheet PDF has no matching text table and yields none.

Bubble sheets also get their *template* auto-detected per sheet (see
template_detect), rather than assuming every dropped sheet is the same
format -- pass --template to force one fixed template for all of them
instead (e.g. if a sheet's format is genuinely ambiguous to auto-detect,
or for a quick one-off test).

A sheet whose auto-detected template is one of the ACT formats wired to
the org's Drive templates (see score_report_pipeline.should_export_to_sheets)
gets its own individual score-report PDF exported there instead of a tab
in the combined .xlsx -- named after the student and flagged (a " FLAG"
suffix, plus the familiar color-coded .xlsx alongside it) whenever the
sheet has any review items (blank/MULTIPLE/low-confidence/unreadable/
pattern-inferred). A score-report PDF's rows are similarly grouped by
source file (one PDF, one student) and exported the same way (see
sat_score_report_pipeline.export_sat_report) once answer-key
identification succeeds for it -- this is the one export path that
prompts for input mid-run (each subject's scaled section score, via a
native dialog; see gui_prompt.py), since nothing upstream can supply that
value yet. Everything else -- an unidentified score report, an
unrecognized bubble-sheet template, or any bubble sheet when --template
forces a fixed one -- still goes into the combined .xlsx exactly as
before this existed. Anything that fails to export on its own (bad
filename convention, no matching Drive template, an unidentified Module 2
difficulty, a cancelled score prompt, Google auth not set up) falls back
into the combined .xlsx too, with a warning explaining why, rather than
failing the whole batch.
"""
from __future__ import annotations

import argparse
import os
import sys
from pathlib import Path
from typing import List, Tuple

from openpyxl import Workbook

from .answer_keys import annotate_rows, load_answer_keys
from .export import add_bubble_sheet_answers_sheet
from .google_sheets_export import build_services
from .loading import IMAGE_SUFFIXES, PDF_SUFFIXES
from .pipeline import SheetResult, UndetectedSheet, process_paths, process_paths_auto
from .sat_score_report_pipeline import export_sat_report
from .score_report import ScoreReportRow, group_by_source, parse_score_report
from .score_report_export import add_score_report_answers_sheet
from .score_report_pipeline import ExportOutcome, export_sheet_report, should_export_to_sheets
from .template import Template

# "Testmastergrids", the root of the org's Drive score-report template
# tree (see README's "Google Sheets score reports" section) --
# overridable for a different Drive layout without a code change.
_DEFAULT_TEMPLATES_ROOT_FOLDER_ID = "1hzDrOzqBymstYHdTqjdLxKOmdlbKqSSt"
# "Temporary Files", the sibling folder the org already set aside for
# exactly this -- each report's filled-in Sheet copy is kept there
# (by default -- see export_filled_report's keep_working_copy), rather
# than in the same folder as the real templates.
_DEFAULT_TEMP_FOLDER_ID = "1eUp4nToItX0_xtDe4Dt3VDlfCQU3ba_y"


def _iter_files(path: Path) -> List[Path]:
    if path.is_dir():
        files: List[Path] = []
        for child in sorted(path.iterdir()):
            files.extend(_iter_files(child))
        return files
    if path.suffix.lower() in IMAGE_SUFFIXES or path.suffix.lower() in PDF_SUFFIXES:
        return [path]
    return []


def classify_inputs(paths: List[Path]) -> Tuple[List[Path], List[ScoreReportRow]]:
    """Split resolved files into (bubble_sheet_paths, score_report_rows).
    Score-report PDFs are fully parsed here (not just flagged) so callers
    don't need to parse them a second time."""
    bubble_paths: List[Path] = []
    score_rows: List[ScoreReportRow] = []
    for path in paths:
        for file in _iter_files(path):
            if file.suffix.lower() in IMAGE_SUFFIXES:
                bubble_paths.append(file)
                continue
            try:
                rows = parse_score_report(file)
            except Exception:
                rows = []
            if rows:
                score_rows.extend(rows)
            else:
                bubble_paths.append(file)
    return bubble_paths, score_rows


def scan_bubble_sheets(
    bubble_paths: List[Path], template_path: str | None
) -> List[SheetResult]:
    """Scan every bubble-sheet input, either against one fixed template
    (if `template_path` is given) or auto-detecting each sheet's template
    individually (see template_detect). Prints a warning to stderr for any
    sheet auto-detection couldn't confidently match rather than failing
    the whole batch over it -- the rest still get scanned and included."""
    if template_path is not None:
        template = Template.from_yaml(template_path)
        template.validate()
        return process_paths(bubble_paths, template)

    results, undetected = process_paths_auto(bubble_paths)
    for sheet in undetected:
        print(f"Warning: couldn't identify {sheet.label}'s template ({sheet.reason})", file=sys.stderr)
    return results


def template_breakdown(results: List[SheetResult]) -> str:
    """" [name: count, ...]" summarizing which auto-detected template each
    sheet matched, or "" when there's nothing to show (a fixed template
    was used, or there's only one sheet)."""
    counts: dict[str, int] = {}
    for r in results:
        if r.template_name:
            counts[r.template_name] = counts.get(r.template_name, 0) + 1
    if len(counts) < 2:
        return ""
    return " [" + ", ".join(f"{name}: {n}" for name, n in counts.items()) + "]"


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        description=(
            "Extract answers from a mix of scanned bubble sheets and "
            "text-based score-report PDFs, auto-detecting each input's type."
        )
    )
    parser.add_argument(
        "--input", required=True, nargs="+", help="One or more images, PDFs, and/or directories"
    )
    parser.add_argument(
        "--template",
        default=None,
        help=(
            "Template YAML to use for every bubble-sheet input found, skipping auto-detection "
            "(by default, each sheet's template is auto-detected individually -- see "
            "answer_extractor.template_detect -- so a batch can mix sheet formats freely; pass "
            "this to force one fixed template instead)"
        ),
    )
    parser.add_argument(
        "--output", required=True, help="Path to write the combined .xlsx to (sheets exported as their own "
        "Sheets-report PDF don't go in here -- see --report-output-dir)"
    )
    parser.add_argument(
        "--no-refresh-keys",
        action="store_true",
        help=(
            "Skip fetching the latest answer key reference data over the network; use the "
            "last cached copy (or the one bundled in this checkout) instead"
        ),
    )
    parser.add_argument(
        "--report-output-dir",
        default=None,
        help="Where to write per-student Sheets-report PDFs (and any flagged .xlsx alongside them) -- "
        "defaults to the Desktop, or $ANSWER_EXTRACTOR_REPORT_OUTPUT_DIR if set",
    )
    parser.add_argument(
        "--templates-root-folder-id",
        default=None,
        help="Drive folder id of the score-report templates root (see README) -- defaults to this org's "
        "own templates folder, or $ANSWER_EXTRACTOR_TEMPLATES_ROOT_FOLDER_ID if set",
    )
    parser.add_argument(
        "--temp-folder-id",
        default=None,
        help="Drive folder id to place each report's filled-in Sheet copy in (kept there, not deleted "
        "-- see export_filled_report's keep_working_copy), instead of the same folder as the real "
        "template it was copied from -- defaults to this org's \"Temporary Files\" folder, or "
        "$ANSWER_EXTRACTOR_TEMP_FOLDER_ID if set",
    )
    return parser


def main(argv: list[str] | None = None) -> int:
    args = build_parser().parse_args(argv)

    input_paths = [Path(p) for p in args.input]
    missing = [p for p in input_paths if not p.exists()]
    if missing:
        for p in missing:
            print(f"Input path does not exist: {p}", file=sys.stderr)
        return 1

    bubble_paths, score_rows = classify_inputs(input_paths)

    if not bubble_paths and not score_rows:
        print(
            f"No images or PDFs found in: {', '.join(str(p) for p in input_paths)}",
            file=sys.stderr,
        )
        return 1

    wb = Workbook()
    del wb["Sheet"]
    summary_parts = []
    exported: List[ExportOutcome] = []
    exported_sat_paths: List[Path] = []

    # --template forces a fixed template for every bubble sheet, which
    # also means opting out of the per-student Sheets-report path
    # (should_export_to_sheets only ever recognizes an *auto-detected*
    # template name -- see pipeline.SheetResult).
    to_export: List[SheetResult] = []
    to_combine: List[SheetResult] = []
    if bubble_paths:
        for r in scan_bubble_sheets(bubble_paths, args.template):
            (to_export if args.template is None and should_export_to_sheets(r) else to_combine).append(r)

    # A score-report PDF's rows can only be routed to the per-student SAT
    # export path once answer-key identification succeeds for it (that's
    # what determines each Module 2's difficulty, without which there's
    # no way to know which of the SAT template's block-pairs to fill in)
    # -- if identification itself fails, every row falls straight back to
    # the combined .xlsx, same as always, with the one warning below
    # rather than a redundant one per source file.
    sat_row_groups: List[List[ScoreReportRow]] = []
    score_rows_to_combine: List[ScoreReportRow] = []
    if score_rows:
        try:
            library = load_answer_keys(refresh=not args.no_refresh_keys)
            score_rows = annotate_rows(score_rows, library)
        except Exception as exc:  # answer-key identification is a bonus, not required for extraction
            print(f"Warning: answer key identification skipped ({exc})", file=sys.stderr)
            score_rows_to_combine = score_rows
        else:
            sat_row_groups = list(group_by_source(score_rows).values())

    output_dir = None
    templates_root_folder_id = None
    if to_export or sat_row_groups:
        try:
            drive, sheets = build_services()
        except Exception as exc:
            print(
                f"Warning: couldn't set up Google Sheets access ({exc}); "
                f"including these sheet(s)/report(s) in {args.output} instead.",
                file=sys.stderr,
            )
            to_combine.extend(to_export)
            to_export = []
            for group in sat_row_groups:
                score_rows_to_combine.extend(group)
            sat_row_groups = []
        else:
            output_dir = Path(
                args.report_output_dir
                or os.environ.get("ANSWER_EXTRACTOR_REPORT_OUTPUT_DIR")
                or (Path.home() / "Desktop")
            )
            output_dir.mkdir(parents=True, exist_ok=True)
            templates_root_folder_id = (
                args.templates_root_folder_id
                or os.environ.get("ANSWER_EXTRACTOR_TEMPLATES_ROOT_FOLDER_ID")
                or _DEFAULT_TEMPLATES_ROOT_FOLDER_ID
            )
            temp_folder_id = (
                args.temp_folder_id
                or os.environ.get("ANSWER_EXTRACTOR_TEMP_FOLDER_ID")
                or _DEFAULT_TEMP_FOLDER_ID
            )

            for r in to_export:
                try:
                    exported.append(
                        export_sheet_report(
                            drive, sheets, templates_root_folder_id, r, output_dir, temp_folder_id=temp_folder_id
                        )
                    )
                except Exception as exc:
                    print(
                        f"Warning: couldn't export {r.label} to a Sheets report ({exc}); "
                        f"including it in {args.output} instead.",
                        file=sys.stderr,
                    )
                    to_combine.append(r)

            for group in sat_row_groups:
                try:
                    exported_sat_paths.append(
                        export_sat_report(
                            drive, sheets, templates_root_folder_id, group, output_dir, temp_folder_id=temp_folder_id
                        )
                    )
                except Exception as exc:
                    source = group[0].source if group else "(unknown)"
                    print(
                        f"Warning: couldn't export {source} to a Sheets report ({exc}); "
                        f"including it in {args.output} instead.",
                        file=sys.stderr,
                    )
                    score_rows_to_combine.extend(group)

    if to_combine:
        add_bubble_sheet_answers_sheet(wb, to_combine)
        summary_parts.append(f"{len(to_combine)} bubble sheet(s) in {args.output}{template_breakdown(to_combine)}")

    total_exported = len(exported) + len(exported_sat_paths)
    if total_exported:
        flagged = sum(1 for o in exported if o.xlsx_path is not None)
        flagged_note = f", {flagged} flagged for review" if flagged else ""
        summary_parts.append(f"{total_exported} score report(s) exported to {output_dir}{flagged_note}")

    if score_rows_to_combine:
        add_score_report_answers_sheet(wb, score_rows_to_combine)
        num_reports = len({row.source for row in score_rows_to_combine})
        summary_parts.append(
            f"{len(score_rows_to_combine)} score-report question(s) from {num_reports} file(s) in {args.output}"
        )

    if not wb.sheetnames and not exported and not exported_sat_paths:
        print("No answers could be extracted from the given input.", file=sys.stderr)
        return 1

    if wb.sheetnames:
        wb.save(args.output)

    print(f"{'; '.join(summary_parts)}.")
    return 0


if __name__ == "__main__":
    sys.exit(main())
