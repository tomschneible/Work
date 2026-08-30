"""Fill the *simplified* SAT/DSAT score-report template's "Student
Responses" tab -- one Module 2 slot per subject, holding whichever
variant answer_keys.annotate_rows already identified, instead of a
checkbox-selected pair of duplicate blocks. Nothing here ever clears,
hides, or narrows a block the way sat_score_report_writer.py's
fill_sat_score_report does for the current-format template -- there's
never an inactive occurrence sitting on the sheet to begin with (see this
repo's README "Google Sheets score reports" section for the whole
narrow/hide/clear saga this sidesteps entirely by construction, not by
patching around it).

A block's title never carries a difficulty on the blank template itself
-- just "<Subject> Module 2", no "- Higher/Lower Difficulty" suffix, since
that can't be known until a specific student's active variant is. This
writes the identified difficulty onto the end of whatever title text the
template already has there, rather than reconstructing the subject's own
display text from scratch -- its capitalization/abbreviation is the
template author's own choice, not something to guess at here.

Per-question facts this template doesn't carry itself -- a question's
correct answer, Domain, and Skill -- come from
sat_score_report_writer.read_reference_questions instead: read straight
off the *current-format* template for the same test, rather than
duplicated into a second, separately-maintained source (see that
function's own docstring for why a current-format template -- made once
per test regardless, and kept for hand-grading -- is already the one
real source for them).

Unlike the current-format template, this one carries no per-test content
of its own -- its shape is fixed by the exam format's own question
counts, not by which specific test it's grading -- so there's exactly
one of it for now, regardless of test code (see the export/lookup layer
for where that would need to grow if PSAT 10/8-9's own shorter modules
join this pipeline; their per-module question counts differ from the
full digital SAT's).
"""
from __future__ import annotations

import dataclasses
import datetime as dt
import re
from pathlib import Path
from typing import List, Mapping, Optional

import openpyxl
from openpyxl.worksheet.worksheet import Worksheet

from .google_sheets_export import CellWrite, FillResult, format_date_for_sheets
from .sat_score_report_writer import (
    SatKey,
    find_header_row,
    find_name_cell,
    find_score_value_cells,
    normalize_subject,
    read_reference_questions,
)

_TITLE_PATTERN = re.compile(r"^(?P<subject>.+?)\s+Module\s+(?P<module_num>1|2)\s*$", re.IGNORECASE)
# "harder"/"easier" (answer_keys.annotate_rows' own vocabulary, and
# active_variants' own values) -> the suffix appended to a Module 2
# block's own title once its variant is known. Matches the wording
# sat_score_report_writer._DIFFICULTY_TO_SLOT maps *from* on the
# current-format template ("Higher"/"Lower Difficulty"), so a report
# built from either template reads the same way.
_SLOT_TO_DIFFICULTY_SUFFIX = {"harder": " - Higher Difficulty", "easier": " - Lower Difficulty"}


@dataclasses.dataclass(frozen=True)
class SimpleSatBlock:
    """One block on the *simplified* template -- unlike SatBlock, never
    one of several same-(subject, module_slot) duplicates (there's only
    ever one on this template), and module_slot is just
    "module1"/"module2": the template itself never encodes a difficulty,
    since that isn't known until a specific student's active variant is
    (see this module's own docstring)."""

    subject: str
    module_slot: str  # "module1" | "module2"
    title_row: int
    header_row: int
    question_col: int
    correct_col: int
    answer_col: int
    mark_col: int


def locate_simple_sat_blocks(ws: Worksheet) -> List[SimpleSatBlock]:
    """Every block on the simplified template's own "Student Responses"
    tab -- exactly one Module 1 and one Module 2 per subject. Raises
    ValueError (via find_header_row) if a title's own "Your Answer"
    header can't be found nearby."""
    blocks = []
    for row in ws.iter_rows():
        for cell in row:
            value = cell.value
            if not isinstance(value, str):
                continue
            match = _TITLE_PATTERN.match(value.strip())
            if not match:
                continue
            subject = normalize_subject(match.group("subject"))
            module_slot = "module1" if match.group("module_num") == "1" else "module2"
            title_row, title_col = cell.row, cell.column
            header_row = find_header_row(ws, title_row, title_col)
            blocks.append(
                SimpleSatBlock(
                    subject=subject,
                    module_slot=module_slot,
                    title_row=title_row,
                    header_row=header_row,
                    question_col=title_col,
                    correct_col=title_col + 1,
                    answer_col=title_col + 2,
                    mark_col=title_col + 3,
                )
            )
    return blocks


def fill_simple_sat_score_report(
    template_path: str | Path,
    reference_ws: Worksheet,
    answers: Mapping[SatKey, str],
    active_variants: Mapping[str, str],
    student_name: str,
    test_date: dt.date | str,
    section_scores: Optional[Mapping[str, int]] = None,
    sheet_name: str = "Student Responses",
) -> FillResult:
    """Return every cell write needed to fill the simplified template's
    `sheet_name` tab in with the student's name, test date, every answer
    in `answers`, and (if given) each subject's scaled section score --
    see sat_score_report_writer.fill_sat_score_report's own docstring for
    what `active_variants`/`answers`/`section_scores` mean, all shared
    verbatim; the difference here is entirely about *where things go*,
    not what they mean.

    `reference_ws` is the *current-format* template's own "Student
    Responses" tab for the same test, loaded read-only the same way
    `template_path` is -- used via read_reference_questions to source
    each block's own Correct Answer/Domain/Skill (see that function's own
    docstring for why this template doesn't carry them itself).

    Unlike fill_sat_score_report, there is no repositioning, clearing, or
    hiding to do: every block found is, by construction, the one this
    student's report should show, so FillResult's other fields are always
    empty here. `answers` must only contain keys for "module1" or each
    subject's active variant, same as fill_sat_score_report -- an entry
    for the inactive variant raises, since there'd be nowhere on this
    template to put it (there's no "wrong" Module 2 slot left sitting
    around to accidentally write into, unlike the current-format
    template's own duplicates). Raises ValueError if a Module 2 block's
    own subject has no entry in `active_variants`, or if
    read_reference_questions can't find a matching block on
    `reference_ws`, or has no entry for one of this block's own question
    numbers (almost always a template/test mismatch between
    `template_path` and `reference_ws`)."""
    for subject, slot, _question in answers:
        if slot != "module1" and active_variants.get(subject) != slot:
            raise ValueError(
                f"Got an answer for {subject!r} {slot!r}, but the active variant for {subject!r} "
                f"is {active_variants.get(subject)!r}"
            )

    wb = openpyxl.load_workbook(template_path, data_only=True)
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"No {sheet_name!r} tab in {template_path} (tabs: {wb.sheetnames})")
    ws = wb[sheet_name]

    name_row, name_col = find_name_cell(ws)
    writes = [
        CellWrite(sheet_name, name_row, name_col, student_name),
        # date sits directly below name
        CellWrite(sheet_name, name_row + 1, name_col, format_date_for_sheets(test_date)),
    ]

    if section_scores:
        score_cells = find_score_value_cells(ws)
        for subject, score in section_scores.items():
            if subject not in score_cells:
                raise ValueError(
                    f"No score cell found for subject {subject!r} in {template_path}!{sheet_name} "
                    f"(found score cells for: {sorted(score_cells)})"
                )
            row, col = score_cells[subject]
            writes.append(CellWrite(sheet_name, row, col, score))

    remaining = dict(answers)
    for block in locate_simple_sat_blocks(ws):
        if block.module_slot == "module1":
            effective_slot = "module1"
        else:
            effective_slot = active_variants.get(block.subject)
            if effective_slot is None:
                raise ValueError(f"No active variant given for {block.subject!r} -- can't fill its Module 2 block")
            title_text = ws.cell(row=block.title_row, column=block.question_col).value
            suffix = _SLOT_TO_DIFFICULTY_SUFFIX[effective_slot]
            writes.append(CellWrite(sheet_name, block.title_row, block.question_col, f"{title_text}{suffix}"))

        reference_questions = read_reference_questions(reference_ws, block.subject, effective_slot)

        r = block.header_row + 1
        while True:
            question = ws.cell(row=r, column=block.question_col).value
            if question is None:
                break
            question_num = int(question)
            reference = reference_questions.get(question_num)
            if reference is None:
                raise ValueError(
                    f"reference_ws has no {block.subject!r} {effective_slot!r} question {question_num} -- "
                    f"template/reference mismatch"
                )
            answer_value = remaining.pop((block.subject, effective_slot, question_num), None)
            writes.append(CellWrite(sheet_name, r, block.correct_col, reference.correct_answer))
            writes.append(CellWrite(sheet_name, r, block.answer_col, answer_value))
            writes.append(CellWrite(sheet_name, r, block.question_col + 4, reference.domain))
            writes.append(CellWrite(sheet_name, r, block.question_col + 5, reference.skill))
            r += 1

    if remaining:
        unmatched = ", ".join(f"{subject} {slot} {question}" for subject, slot, question in sorted(remaining))
        raise ValueError(f"{template_path}!{sheet_name} has no answer block for: {unmatched}")

    return FillResult(cell_writes=writes)
