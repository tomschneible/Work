"""Drive/Sheets operations behind the Google Sheets score-report export
path: duplicate a template, fill it in via direct Sheets API cell writes,
and export the copy as a PDF -- see README's "Google Sheets score
reports" section for the overall design and why each step exists.

Filling in a report used to round-trip the *entire* workbook through
.xlsx: export_xlsx pulled a copy down locally, a writer module
(score_report_writer.fill_score_report / sat_score_report_writer's
counterpart) edited it with openpyxl, and replace_content pushed the
whole thing back in, letting Drive convert it back to native Sheets
format on upload. That turned out not to be safe: confirmed live,
re-importing an openpyxl-authored .xlsx doesn't reconstruct some of a
template's own formatting with full fidelity to how Google's own native
export/import round-trip would -- specifically, a merged, centered title
cell on a tab this pipeline never even touches (the org's own "Cover
Page") came back with its text no longer filling the cell, even though
the underlying value was intact. Since every tab this pipeline actually
needs to change (a report's own answer grid) works the same as any other
live-Sheet edit -- a value in a specific cell -- there was never a real
need to touch the rest of the workbook at all.

So now: export_xlsx is still used, but only to download a *read-only*
local copy a writer module scans to figure out *where* to write (find the
name/date placeholder cells, each question's answer cell, ...) --
score_report_writer.fill_score_report and its SAT counterpart now return
a plain list of CellWrite instead of an edited Workbook. write_cells then
pushes exactly those cells into the live Sheet via the Sheets API's
values().batchUpdate -- every other tab (Cover Page, Student Report,
Content, ...), most of which are populated by formulas referencing the
answer tab anyway, is never re-converted through .xlsx at all, so nothing
about their own formatting is ever at risk from this pipeline again.

replace_content itself is kept here only as a thin, generically correct
wrapper -- nothing in this codebase calls it any more. A template's own
gridlines get turned off via hide_gridlines instead (a direct Sheets API
metadata change, no file conversion involved at all): confirmed live the
hard way that pointing the xlsx round-trip at a *template* file directly
(via replace_content) is exactly as unsafe as it was for a per-report
copy -- it corrupted the org's own live "ACT 25MC1" template's Cover Page
the one time it was tried, recovered only via Sheets' own version
history. Don't reach for replace_content to fix a template's formatting;
extend hide_gridlines's approach (a targeted batchUpdate request) instead
of reintroducing an xlsx round-trip anywhere in this codebase.

Deliberately thin wrappers around the official googleapiclient calls
rather than a bigger abstraction: there's no meaningful behavior to
capture beyond "make this one Drive/Sheets API call correctly," and
keeping each call in its own small function is what makes it possible to
unit-test the request-shaping (does this pass supportsAllDrives, the
right fields, ...) with a mocked service object instead of live network
access every test run needs. export_pdf is the one exception -- it calls
Sheets' own dedicated (undocumented, no googleapiclient wrapper) export
URL directly rather than a Drive/Sheets API method, since confirmed live
that Drive's generic files.export doesn't reliably apply a sheet's own
print scale settings; see its own docstring.
"""
from __future__ import annotations

import dataclasses
import datetime as dt
import io
from typing import Dict, List, Optional, Sequence, Tuple, Union

from google.auth.transport.requests import AuthorizedSession
from googleapiclient.discovery import Resource, build
from googleapiclient.http import MediaFileUpload, MediaIoBaseDownload
from openpyxl.utils import get_column_letter

from .google_auth import get_credentials

_XLSX_MIME_TYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

CellValue = Union[str, int, float, bool, None]


@dataclasses.dataclass(frozen=True)
class CellWrite:
    """One cell to set in a live Sheet -- `row`/`column` are 1-indexed,
    matching openpyxl (and this pipeline's writer modules locate cells
    with openpyxl in the first place, against a read-only local download
    -- see this module's own docstring)."""

    sheet: str
    row: int
    column: int
    value: CellValue


@dataclasses.dataclass(frozen=True)
class FillResult:
    """What a fill_fn (score_report_writer.fill_score_report,
    sat_score_report_writer.fill_sat_score_report) returns to
    google_report_export_common.export_filled_report: the individual
    cell writes needed, plus any rectangles that should be cleared
    (value and border formatting both) from the exported PDF once those
    writes land, plus any whole columns that should be hidden outright.
    Both are empty for every fill_fn except SAT's -- it uses these to
    remove whichever Module 2 block occurrence (a subject's other
    difficulty, or a duplicate/twin) wasn't actually administered, so
    the report only shows the modules that were actually filled in; see
    sat_score_report_writer.blocks_to_clear for why *both* are needed
    (clearing alone leaves an occurrence's columns blank but still
    taking up print-area width, which forces the exported PDF's own
    "fit to page" scale down far more than the actually-visible content
    needs -- confirmed live, see hide_columns' own docstring).

    `cleared_ranges` entries are (sheet_name, 0-indexed start row,
    0-indexed end row, 0-indexed start column, 0-indexed end column) --
    all ends exclusive, the shape clear_cells' Sheets API request needs.
    `hidden_column_ranges` entries are (sheet_name, 0-indexed start
    column, 0-indexed end column) -- end exclusive, the shape
    hide_columns' Sheets API request needs; whole-column, so no row
    bounds. `deleted_row_ranges` is a *different* fix for a *different*
    problem, not the row-dimension counterpart of hiding: (sheet_name,
    0-indexed start row, 0-indexed end row), end exclusive, for
    delete_rows -- used to remove a sheet's own trailing blank rows
    outright, since merely hiding them (tried first) turned out to have
    no effect at all on the exported PDF's print area, unlike hiding a
    column; see delete_rows' own docstring. `narrowed_column_ranges` is
    yet another different fix, for a third, separate problem: (sheet_name,
    0-indexed start column, 0-indexed end column, shrink factor), end
    column exclusive, for narrow_columns -- used to shrink the answer
    tables' own column widths so "fit to page" doesn't have to shrink the
    whole page's scale nearly as far to keep them within one page's
    width, which -- since that scale applies uniformly -- was leaving
    height under-filled even though height alone had room to spare; see
    sat_score_report_writer.visible_table_columns_to_narrow for the full
    reasoning."""

    cell_writes: List[CellWrite]
    cleared_ranges: Sequence[Tuple[str, int, int, int, int]] = ()
    hidden_column_ranges: Sequence[Tuple[str, int, int]] = ()
    deleted_row_ranges: Sequence[Tuple[str, int, int]] = ()
    narrowed_column_ranges: Sequence[Tuple[str, int, int, float]] = ()


def format_date_for_sheets(value: dt.date | str) -> str:
    """A `test_date` (score_report_writer.fill_score_report and its SAT
    counterpart both take either a real `date` or an already-formatted
    string -- see either one's own docstring) as a string suitable for a
    Sheets API cell write. ISO format (`date.isoformat()`) for a real
    `date`, since with `write_cells`'s USER_ENTERED input option it's
    reliably recognized as a date regardless of the Sheet's own locale
    settings, unlike an ambiguous M/D/Y string; an already-formatted
    string (no specific day known -- e.g. "January 2026") is passed
    through as plain text, same as before."""
    if isinstance(value, dt.date):
        return value.isoformat()
    return value


def build_services() -> tuple[Resource, Resource]:
    """(drive_service, sheets_service), both authenticated the same way --
    see google_auth.get_credentials for what triggers the interactive
    consent flow vs. a silent token refresh."""
    creds = get_credentials()
    drive = build("drive", "v3", credentials=creds)
    sheets = build("sheets", "v4", credentials=creds)
    return drive, sheets


def list_folder(drive: Resource, folder_id: str) -> List[Dict[str, str]]:
    """Every non-trashed file directly inside a Drive folder -- id, name,
    mimeType -- including one that lives in a Shared Drive (see
    supportsAllDrives/includeItemsFromAllDrives/corpora below; omitting
    any of the three silently drops Shared Drive results rather than
    erroring, so all three are always passed together). Used to identify
    a template file's id from a human-readable folder link before it's
    hardcoded anywhere -- see google_sheets_cli's `list-folder` command."""
    files: List[Dict[str, str]] = []
    page_token = None
    while True:
        response = (
            drive.files()
            .list(
                q=f"'{folder_id}' in parents and trashed = false",
                fields="nextPageToken, files(id, name, mimeType)",
                supportsAllDrives=True,
                includeItemsFromAllDrives=True,
                corpora="allDrives",
                pageToken=page_token,
            )
            .execute()
        )
        files.extend(response.get("files", []))
        page_token = response.get("nextPageToken")
        if not page_token:
            break
    return files


def copy_template(
    drive: Resource, template_file_id: str, new_name: str, parent_folder_id: Optional[str] = None
) -> str:
    """Duplicate `template_file_id` (a Sheets file) as `new_name`, returning
    the new copy's own file id. `parent_folder_id`, if given, places the
    copy there instead of Drive's default (the copier's own My Drive
    root) -- also needs supportsAllDrives when that folder is a Shared
    Drive folder."""
    body = {"name": new_name}
    if parent_folder_id:
        body["parents"] = [parent_folder_id]
    result = drive.files().copy(fileId=template_file_id, body=body, supportsAllDrives=True).execute()
    return result["id"]


def _export(drive: Resource, file_id: str, mime_type: str) -> bytes:
    request = drive.files().export_media(fileId=file_id, mimeType=mime_type)
    buffer = io.BytesIO()
    downloader = MediaIoBaseDownload(buffer, request)
    done = False
    while not done:
        _, done = downloader.next_chunk()
    return buffer.getvalue()


def export_pdf(spreadsheet_id: str) -> bytes:
    """The Sheets file at `spreadsheet_id`, rendered to PDF via Sheets'
    own dedicated export endpoint (`docs.google.com/spreadsheets/d/{id}/
    export?format=pdf`) -- the same URL "File > Download > PDF" in the
    Sheets UI itself navigates to, not Drive's generic `files.export`
    (used by every other `_export`-based function in this module, and
    what this used to call too). Switched deliberately: confirmed live
    that Drive's generic export doesn't reliably apply a sheet's own
    "fit to height"/"fit to page" print scale the same way the Sheets UI
    export does, even though the setting itself is genuinely saved on the
    file correctly -- the exported PDF came out at undistorted, un-shrunk
    size and overflowed onto an extra page regardless. This endpoint,
    called with no scale/margin/gid overrides of its own, defers to
    whatever print setup (page range, layout, scale) is already saved on
    the file, the same as Drive's export was meant to -- just apparently
    more faithfully. No `gid` is passed, so this exports the *entire*
    workbook (every visible sheet), matching the Drive-based export it
    replaces.

    Undocumented as a formal Google API (there's no googleapiclient
    wrapper for it) -- authenticated the same way as every other call in
    this module (`google_auth.get_credentials`), just via a raw
    `AuthorizedSession` GET instead of a `Resource` method, since this
    isn't a `sheets`/`drive` API call. Raises RuntimeError if the
    response isn't actually a PDF (e.g. an HTML error/login page, which
    this endpoint can return with a 200 status instead of a clean HTTP
    error for some failure modes)."""
    creds = get_credentials()
    session = AuthorizedSession(creds)
    url = f"https://docs.google.com/spreadsheets/d/{spreadsheet_id}/export"
    response = session.get(url, params={"format": "pdf"})
    response.raise_for_status()
    content_type = response.headers.get("Content-Type", "")
    if not content_type.startswith("application/pdf"):
        raise RuntimeError(
            f"Expected a PDF from Sheets' export endpoint for {spreadsheet_id}, got "
            f"content-type {content_type!r} instead (first 200 bytes: {response.content[:200]!r})"
        )
    return response.content


def export_xlsx(drive: Resource, file_id: str) -> bytes:
    """The Sheets file at `file_id`, converted to .xlsx bytes -- used to
    pull a freshly-duplicated template down locally, *read-only*, so a
    writer module (score_report_writer.fill_score_report, which only
    knows how to scan a local .xlsx via openpyxl, not a live Sheet over
    the API) can figure out where its writes need to go -- see this
    module's own docstring for why the result is never edited or
    re-uploaded wholesale any more, only used to locate cells for
    write_cells."""
    return _export(drive, file_id, _XLSX_MIME_TYPE)


def replace_content(drive: Resource, file_id: str, local_xlsx_path: str) -> None:
    """Overwrite the Sheets file at `file_id` with the contents of a local
    .xlsx -- Drive converts it to native Sheets format on upload, the same
    conversion Google Sheets' own File > Import > Replace spreadsheet
    does. Nothing in this codebase calls this any more (see this
    module's own docstring): confirmed live that pointing it at a
    template file directly doesn't reconstruct that file's own
    formatting with full fidelity, corrupting a real template the one
    time this was tried for template maintenance. Kept only as a thin,
    correct wrapper -- don't reach for this to fix a template's
    formatting; use a targeted Sheets API batchUpdate (see
    hide_gridlines) instead."""
    media = MediaFileUpload(local_xlsx_path, mimetype=_XLSX_MIME_TYPE)
    drive.files().update(fileId=file_id, media_body=media, supportsAllDrives=True).execute()


def delete_file(drive: Resource, file_id: str) -> None:
    """Permanently remove a file (used to clean up a working copy after a
    failed export attempt, or after a successful one when the caller
    asked not to keep it -- see google_report_export_common.
    export_filled_report's keep_working_copy)."""
    drive.files().delete(fileId=file_id, supportsAllDrives=True).execute()


def _a1(row: int, column: int) -> str:
    return f"{get_column_letter(column)}{row}"


def write_cells(sheets: Resource, spreadsheet_id: str, cells: Sequence[CellWrite]) -> None:
    """Write `cells` into the live Sheet at `spreadsheet_id` in one
    `values().batchUpdate` call, each cell addressed by its own
    single-cell A1 range (e.g. `"'ScoreSheet'!C5"`) -- this pipeline's
    writes are scattered across many rows/columns, never one contiguous
    range, so one range per cell rather than trying to batch adjacent
    ones is both simpler and doesn't need write_cells to know or care
    about layout at all. `value_input_option="USER_ENTERED"` -- same as
    typing directly into a cell -- so e.g. a date string from
    format_date_for_sheets is still recognized and formatted as a date,
    matching what the previous xlsx-based writer relied on openpyxl's
    native date cell type for. A cell whose value is None becomes an
    empty string, clearing whatever the template otherwise has there
    (e.g. an omitted question's blank "Your Answer") -- the API has no
    separate "null" value. A no-op (no API call at all) if `cells` is
    empty."""
    if not cells:
        return
    data = [
        {
            "range": f"'{c.sheet}'!{_a1(c.row, c.column)}",
            "values": [["" if c.value is None else c.value]],
        }
        for c in cells
    ]
    sheets.spreadsheets().values().batchUpdate(
        spreadsheetId=spreadsheet_id,
        body={"valueInputOption": "USER_ENTERED", "data": data},
    ).execute()


def clear_cells(
    sheets: Resource, spreadsheet_id: str, ranges: Sequence[Tuple[str, int, int, int, int]]
) -> None:
    """Clear the value, border formatting, *and* data validation of every
    cell in each of `ranges` -- (sheet_name, 0-indexed start row,
    0-indexed end row, 0-indexed start column, 0-indexed end column), all
    ends exclusive, the shape FillResult.cleared_ranges carries -- via one
    Sheets API `batchUpdate` `repeatCell` request per range. A per-report,
    per-copy content change (unrelated to hide_gridlines' template-wide
    metadata fix above) -- used by sat_score_report_writer.
    fill_sat_score_report to remove a Module 2 block occurrence that
    wasn't administered. Data validation is cleared alongside value and
    border deliberately: a boolean-type validation renders as a checkbox
    *widget* independent of the cell's own value, so clearing only the
    value left an empty, unchecked checkbox floating with nothing else
    around it -- confirmed live. One `get` call resolves every sheet name
    to its numeric sheetId first, since the batchUpdate request itself
    only accepts that, not a name. A no-op (no API call at all) if
    `ranges` is empty. Raises ValueError if a range names a sheet this
    spreadsheet doesn't have."""
    if not ranges:
        return
    meta = sheets.spreadsheets().get(spreadsheetId=spreadsheet_id, fields="sheets.properties").execute()
    sheet_id_by_title = {s["properties"]["title"]: s["properties"]["sheetId"] for s in meta.get("sheets", [])}
    requests = []
    for sheet_name, start_row, end_row, start_col, end_col in ranges:
        if sheet_name not in sheet_id_by_title:
            raise ValueError(f"No sheet named {sheet_name!r} in spreadsheet {spreadsheet_id}")
        requests.append(
            {
                "repeatCell": {
                    "range": {
                        "sheetId": sheet_id_by_title[sheet_name],
                        "startRowIndex": start_row,
                        "endRowIndex": end_row,
                        "startColumnIndex": start_col,
                        "endColumnIndex": end_col,
                    },
                    "cell": {},
                    "fields": "userEnteredValue,userEnteredFormat.borders,dataValidation",
                }
            }
        )
    sheets.spreadsheets().batchUpdate(spreadsheetId=spreadsheet_id, body={"requests": requests}).execute()


def narrow_columns(
    sheets: Resource, spreadsheet_id: str, ranges: Sequence[Tuple[str, int, int, float]]
) -> None:
    """Shrink every column in each of `ranges` -- (sheet_name, 0-indexed
    start column, 0-indexed end column [exclusive], shrink factor) -- to
    `factor` times its own *current* width, via one Sheets API `get` call
    that reads each column's actual current `pixelSize`, followed by one
    `updateDimensionProperties` request per column setting its new
    `pixelSize` explicitly (`round(current * factor)`, floored at 1px). A
    no-op (no API call at all) if `ranges` is empty.

    This exists because Sheets' "fit to page" print scale is computed
    from the print area's *natural* (unscaled) size -- confirmed live
    (see sat_score_report_writer.visible_table_columns_to_narrow's own
    docstring for the full reasoning and the real numbers behind it):
    with the Module 2 answer tables' natural column widths as wide as
    they currently are, "fit to page" has to shrink everything down to
    ~55% to keep the *widest* dimension (width, confirmed the binding
    one -- a real export's rendered width already reached the page's
    full available width at that scale, while its rendered height fell
    well short of the page's available height) inside one page -- and
    since scale applies uniformly to both dimensions, that same
    width-driven 55% leaves height under-filled even though height alone
    had plenty of room to spare. Narrowing the columns that make width
    the tighter constraint lets "fit to page" recompute a *larger*
    uniform scale on its own (still never overflowing -- "fit to page"
    always finds whatever scale fits, regardless of how large or small
    the natural size is), which -- being applied uniformly -- makes the
    *unshrunk* rows render taller too, filling more of the page's actual
    height. This changes column *width*, not font size or row height, on
    the theory that width is what's forcing the scale down in the first
    place; nothing about font size is touched here.

    Deliberately reads each column's *actual* current pixel width from
    Sheets itself, rather than trying to convert `openpyxl`'s own
    (character-unit) column width into pixels locally -- confirmed live
    that a generic char-unit-to-pixel formula doesn't reliably match
    what Sheets actually renders a given `width` as; reading the real
    `pixelSize` avoids compounding that guesswork into an already
    uncertain factor.

    Rows and columns are *not* interchangeable for this kind of fix --
    confirmed live, painfully: hiding and deleting a sheet's own trailing
    rows (see delete_rows) had *zero* measurable effect on the exported
    PDF, while hiding columns (hide_columns) measurably changed it. This
    function only ever touches columns for that reason.

    One `get` call resolves every sheet name to its numeric sheetId and
    reads that sheet's current column widths together (multiple `ranges`
    for the same sheet are matched back up by request order -- the
    Sheets API's own `get` response groups `data` entries per sheet, not
    per requested range, so this doesn't just zip `ranges` against the
    response 1:1). Raises ValueError if a range names a sheet this
    spreadsheet doesn't have."""
    if not ranges:
        return
    meta = sheets.spreadsheets().get(spreadsheetId=spreadsheet_id, fields="sheets.properties").execute()
    sheet_id_by_title = {s["properties"]["title"]: s["properties"]["sheetId"] for s in meta.get("sheets", [])}
    for sheet_name, _start, _end, _factor in ranges:
        if sheet_name not in sheet_id_by_title:
            raise ValueError(f"No sheet named {sheet_name!r} in spreadsheet {spreadsheet_id}")

    a1_ranges = [
        f"'{sheet_name}'!{get_column_letter(start_col + 1)}1:{get_column_letter(end_col)}1"
        for sheet_name, start_col, end_col, _factor in ranges
    ]
    widths_response = sheets.spreadsheets().get(
        spreadsheetId=spreadsheet_id,
        ranges=a1_ranges,
        fields="sheets.properties.sheetId,sheets.data.columnMetadata.pixelSize",
    ).execute()
    data_by_sheet_id: Dict[int, List[dict]] = {
        sheet_result["properties"]["sheetId"]: sheet_result.get("data", [])
        for sheet_result in widths_response.get("sheets", [])
    }
    next_data_index: Dict[int, int] = {}

    requests = []
    for sheet_name, start_col, end_col, factor in ranges:
        sheet_id = sheet_id_by_title[sheet_name]
        data_index = next_data_index.get(sheet_id, 0)
        next_data_index[sheet_id] = data_index + 1
        column_metadata = data_by_sheet_id.get(sheet_id, [])[data_index].get("columnMetadata", [])
        for offset, col_meta in enumerate(column_metadata):
            current_width = col_meta.get("pixelSize")
            if current_width is None:
                continue
            new_width = max(1, round(current_width * factor))
            col_index = start_col + offset
            requests.append(
                {
                    "updateDimensionProperties": {
                        "range": {
                            "sheetId": sheet_id,
                            "dimension": "COLUMNS",
                            "startIndex": col_index,
                            "endIndex": col_index + 1,
                        },
                        "properties": {"pixelSize": new_width},
                        "fields": "pixelSize",
                    }
                }
            )
    if requests:
        sheets.spreadsheets().batchUpdate(spreadsheetId=spreadsheet_id, body={"requests": requests}).execute()


def hide_columns(sheets: Resource, spreadsheet_id: str, ranges: Sequence[Tuple[str, int, int]]) -> None:
    """Hide every whole column in each of `ranges` -- (sheet_name,
    0-indexed start column, 0-indexed end column), end exclusive, the
    shape FillResult.hidden_column_ranges carries -- via one Sheets API
    `batchUpdate` `updateDimensionProperties` request per range, the same
    metadata change Sheets' own right-click "Hide column" sets
    (`hiddenByUser`).

    Used alongside clear_cells, not instead of it: clearing a Module 2
    occurrence's cells removes its *content* but leaves its columns
    still fully present (and still full width) in the sheet's print
    area -- confirmed live against a real export, a cleared-but-not-
    hidden occurrence's blank columns were still being counted when
    "fit to page" computed its scale, forcing that scale down far below
    what the actually-visible content needed and leaving the exported
    PDF's real tables squeezed into a fraction of the page with a large
    blank margin around them. Hiding the same columns this clears
    removes them from the print area entirely, letting "fit to page"
    scale to the content that's actually left.

    This used to be how a Module 2 occurrence was hidden at all, before
    fill_sat_score_report started consolidating every subject's real
    answers into one shared column (_canonical_module2_col) -- back then
    hiding was keyed only by which difficulty was "active," which broke
    the moment two subjects administered *different* difficulties (each
    needing a *different* column hidden, but Reading & Writing and Math
    share the same four column positions -- see sat_score_report_writer's
    own module docstring). Now that every subject's real answers always
    land in the same canonical column and every other occurrence is
    always cleared regardless of subject, the same non-canonical columns
    are always the ones being hidden too -- safe again, since nothing
    real is ever left there for *any* subject to lose.

    One `get` call resolves every sheet name to its numeric sheetId
    first, since the batchUpdate request itself only accepts that, not a
    name. A no-op (no API call at all) if `ranges` is empty. Raises
    ValueError if a range names a sheet this spreadsheet doesn't have."""
    if not ranges:
        return
    meta = sheets.spreadsheets().get(spreadsheetId=spreadsheet_id, fields="sheets.properties").execute()
    sheet_id_by_title = {s["properties"]["title"]: s["properties"]["sheetId"] for s in meta.get("sheets", [])}
    requests = []
    for sheet_name, start_col, end_col in ranges:
        if sheet_name not in sheet_id_by_title:
            raise ValueError(f"No sheet named {sheet_name!r} in spreadsheet {spreadsheet_id}")
        requests.append(
            {
                "updateDimensionProperties": {
                    "range": {
                        "sheetId": sheet_id_by_title[sheet_name],
                        "dimension": "COLUMNS",
                        "startIndex": start_col,
                        "endIndex": end_col,
                    },
                    "properties": {"hiddenByUser": True},
                    "fields": "hiddenByUser",
                }
            }
        )
    sheets.spreadsheets().batchUpdate(spreadsheetId=spreadsheet_id, body={"requests": requests}).execute()


def delete_rows(sheets: Resource, spreadsheet_id: str, ranges: Sequence[Tuple[str, int, int]]) -> None:
    """Permanently remove every whole row in each of `ranges` --
    (sheet_name, 0-indexed start row, 0-indexed end row), end exclusive,
    the shape FillResult.deleted_row_ranges carries -- via one Sheets API
    `batchUpdate` `deleteDimension` request per range.

    This used to be `hide_rows` (`updateDimensionProperties`, setting
    `hiddenByUser` -- the exact same *kind* of request hide_columns still
    uses for Module 2's non-canonical columns) -- confirmed live that
    hiding rows this way had *no effect at all* on the exported PDF's
    "fit to page" scale, unlike hiding columns, which measurably fixed
    the analogous width problem: a real export before and after hiding
    the same trailing rows came out pixel-for-pixel identical. Deleting
    the rows outright instead of hiding them does what hiding evidently
    doesn't -- it actually shrinks the sheet's own row count, so there's
    nothing left there at all for Sheets' print-area/used-range
    computation to still be counting.

    Used for a different problem than hide_columns, though: not a Module
    2 occurrence that wasn't administered, but a sheet's own trailing
    rows that were never real content to begin with. Confirmed live
    against the real "Student Responses" tab: it carries formatting
    (row heights, borders) out to row 996 even though its real content
    -- every block, every score cell, the footer -- ends at row 64; see
    sat_score_report_writer.trailing_rows_to_delete for how that boundary
    is found. Safe to delete outright (rather than merely hide) because
    it's *only ever* the sheet's own trailing rows, strictly below every
    row anything real -- a formula, a flag cell, an answer -- lives on:
    deleting rows shifts row numbers for whatever comes *after* the
    deleted range, and nothing does, since this is always the sheet's own
    last rows.

    One `get` call resolves every sheet name to its numeric sheetId
    first, since the batchUpdate request itself only accepts that, not a
    name. A no-op (no API call at all) if `ranges` is empty. Raises
    ValueError if a range names a sheet this spreadsheet doesn't have."""
    if not ranges:
        return
    meta = sheets.spreadsheets().get(spreadsheetId=spreadsheet_id, fields="sheets.properties").execute()
    sheet_id_by_title = {s["properties"]["title"]: s["properties"]["sheetId"] for s in meta.get("sheets", [])}
    requests = []
    for sheet_name, start_row, end_row in ranges:
        if sheet_name not in sheet_id_by_title:
            raise ValueError(f"No sheet named {sheet_name!r} in spreadsheet {spreadsheet_id}")
        requests.append(
            {
                "deleteDimension": {
                    "range": {
                        "sheetId": sheet_id_by_title[sheet_name],
                        "dimension": "ROWS",
                        "startIndex": start_row,
                        "endIndex": end_row,
                    },
                }
            }
        )
    sheets.spreadsheets().batchUpdate(spreadsheetId=spreadsheet_id, body={"requests": requests}).execute()


def hide_gridlines(sheets: Resource, spreadsheet_id: str) -> None:
    """Turn off gridlines on every sheet of `spreadsheet_id` that doesn't
    already have them off, via one Sheets API `batchUpdate` setting each
    sheet's own `gridProperties.hideGridlines` directly -- a pure
    metadata change, no file conversion involved at all (see this
    module's own docstring on why an xlsx round-trip -- download, edit,
    re-upload -- was tried for this instead and confirmed live to
    corrupt a template it was pointed at directly, the same failure mode
    already ruled out for per-report generation).

    Meant for a deliberate one-time fix applied directly to a template
    file itself, via google_sheets_cli's `hide-gridlines` command -- not
    used by per-report generation. A no-op (no API call at all) if every
    sheet already has gridlines hidden.
    """
    meta = sheets.spreadsheets().get(spreadsheetId=spreadsheet_id, fields="sheets.properties").execute()
    requests = []
    for sheet in meta.get("sheets", []):
        properties = sheet["properties"]
        if properties.get("gridProperties", {}).get("hideGridlines"):
            continue
        requests.append(
            {
                "updateSheetProperties": {
                    "properties": {
                        "sheetId": properties["sheetId"],
                        "gridProperties": {"hideGridlines": True},
                    },
                    "fields": "gridProperties.hideGridlines",
                }
            }
        )
    if not requests:
        return
    sheets.spreadsheets().batchUpdate(spreadsheetId=spreadsheet_id, body={"requests": requests}).execute()
