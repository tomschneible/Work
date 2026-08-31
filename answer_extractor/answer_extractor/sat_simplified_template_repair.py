"""One-time repair for the simplified SAT template's own "Calculations"
and "Student Responses" formulas -- not part of the per-student export
pipeline at all: a maintenance operation run directly against the
template file itself, the same category google_sheets_cli.py's
hide-gridlines command already covers for a different template-level
fix (see that module's own docstring).

Why this exists: the current-format template's score-summary formulas
(a subject's total correct/incorrect count on "Student Responses", and
every per-Domain/per-Skill breakdown on "Calculations") were built for
four Module 2 occurrences per subject -- one always-real column, plus
three gated behind a boolean flag cell (`if($O$8=TRUE, ..., 0)` and its
$V$8/$AC$8/$AJ$8 counterparts, one per occurrence -- see
sat_score_report_writer.py's own module docstring on why the flag
scheme exists at all). Deleting the three non-canonical Module 2
occurrences' own columns to build the simplified template (see
sat_simplified_score_report_writer.py) broke every formula that
referenced them: confirmed against a real template, the deleted
columns' cells now read `#REF!`, and since a spreadsheet formula's error
propagates through addition, every summary and Domain/Skill count built
on top of them comes out blank -- while the exported PDF's own
question-level page (which reads "Student Responses" directly, not
through these formulas) still shows real, correct data. This is what
that looks like from the fill side: the answers land correctly, but
"none of the calculations get pulled over."

`repaired_formula` turns a *working* current-format formula (read from a
real current-format template, never from the broken simplified one) into
the simplified template's own equivalent: the three dead
$V$8/$AC$8/$AJ$8-gated branches are dropped entirely (their columns
don't exist on the simplified template to reference), and the one
remaining `if($O$8=TRUE, INNER, 0)` is unwrapped down to just `INNER` --
the simplified template's own Module 2 slot has no flag cell at all
(there's only ever one occurrence, always administered, nothing to
gate), so there's nothing left to check.

`repair_calculations_writes` finds every formula worth restoring in a
*reference* workbook (any real current-format template -- this scheme is
the same across every one, not test-specific), not hardcoded to specific
rows, and returns the repaired version of each as a CellWrite at the
*same* (sheet, row, column) -- the simplified template's own Domain/Skill
label columns are confirmed identical in position to the current-format
template's (both are College Board's own fixed content taxonomy, not
per-test content), so a straight positional copy is safe. Its own
docstring covers exactly what counts as "worth restoring" and why that's
deliberately a *wider* net on "Calculations" than everywhere else: a
first version of this only looked for the flag-gated shape specifically
(confirmed against a real template, 92 cells that way), which missed
"Calculations"' own non-flag-gated formulas (e.g. "% of Section",
`=C2/54`) that got cleared in the very same pass but never referenced a
flag cell to be caught by that narrower search -- confirmed live, this
is what a real export's own "0% of section" (instead of the real
percentage) turned out to be, even after the first 92 were confirmed
fixed. Push the result via google_sheets_export.write_cells (its own
USER_ENTERED value input option means a string starting with "=" lands
as a live formula, not literal text) directly against the simplified
template file -- never through an .xlsx download/re-upload round trip
(see google_sheets_export.py's own module docstring, and hide_gridlines'
own, for why that already corrupted a template once).

Run this once against the simplified template. If a new current-format
template ever changes this formula scheme, or the simplified template's
own Domain/Skill rows are ever reordered relative to a reference
template's, re-run it against a matching pair.
"""
from __future__ import annotations

import re
from typing import List

from openpyxl.workbook import Workbook

from .google_sheets_export import CellWrite

# Matches only the pattern this repair understands -- a whole-cell scan
# for candidates uses this same regex (via .search) before
# repaired_formula is ever called, so this stays in sync with what
# counts as "affected." A plain substring check on e.g. "O8=TRUE" would
# miss the "$O$8=TRUE" shape (the "$" signs break contiguity) that
# "Student Responses"' own formulas actually use, unlike "Calculations"'
# cross-sheet ones -- confirmed live, this cost a round of debugging once
# already; kept as a regex specifically so that mistake can't recur.
_FLAG_GATE_MARKER = re.compile(r"O\$?8=TRUE")

# The three dead branches: $V$8/$AC$8/$AJ$8 (each optionally sheet-
# qualified, e.g. "'Student Responses'!$V$8" from a different sheet, or
# bare "$V$8"/"V8" from a formula already on "Student Responses" itself)
# gating a COUNTIF/COUNTIFS call that reads a now-deleted column. Matched
# non-greedily up to its own closing ",0)" -- confirmed against every
# real formula this was built from that no inner COUNTIF/COUNTIFS call's
# own arguments contain that literal substring, so this doesn't stop
# short at the wrong paren. `\s*` before the leading `+` too, not just
# after -- confirmed live a formula's own "...0) + \nif(...)" left a
# stray space behind otherwise (the space sits *before* the "+" this
# match starts at, so it's outside the match unless explicitly included).
_DEAD_BRANCH = re.compile(r"\s*\+\s*\n?if\((?:'Student Responses'!)?\$?(?:V|AC|AJ)\$?8=TRUE,.*?,0\)")
# The one remaining $O$8 gate, unwrapped down to its own inner formula --
# same optional-sheet-qualifier, non-greedy, and leading-whitespace
# reasoning as above.
_FLAG_GATE = re.compile(r"\s*\+?\s*\n?if\((?:'Student Responses'!)?\$?O\$?8=TRUE,(.*?),0\)")


def repaired_formula(formula: str) -> str:
    """The simplified template's own version of a working current-format
    formula -- see this module's own docstring for what changes and why.
    Raises ValueError if the result still references the deleted
    columns' flag cells, still has an unbalanced paren count, or doesn't
    still read as a formula -- a sign this formula's own shape doesn't
    match what this repair was built to understand, worth a human
    looking at rather than writing a guess into a live template."""
    repaired = _DEAD_BRANCH.sub("", formula)
    repaired = _FLAG_GATE.sub(lambda m: "+" + m.group(1), repaired)

    if not repaired.startswith("="):
        raise ValueError(f"Repaired formula doesn't start with '=': {formula!r} -> {repaired!r}")
    if repaired.count("(") != repaired.count(")"):
        raise ValueError(f"Repaired formula has unbalanced parens: {formula!r} -> {repaired!r}")
    stripped = repaired.replace("$", "")
    if any(tok in stripped for tok in ("V8", "AC8", "AJ8", "O8")):
        raise ValueError(f"Repaired formula still references a flag cell: {formula!r} -> {repaired!r}")
    return repaired


def repair_calculations_writes(reference_wb: Workbook) -> List[CellWrite]:
    """Every CellWrite needed to repair the simplified template, derived
    from `reference_wb` (a real, working current-format template loaded
    read-only via openpyxl -- e.g. `openpyxl.load_workbook(path,
    data_only=False)`; `data_only=False` matters here, unlike most other
    readers in this package, since this needs each cell's own formula
    text, not its last-cached value).

    Two different scopes, deliberately not the same one everywhere:

    - On "Calculations": *every* formula cell, not just ones matching
      the flag-gated shape. Confirmed against a real template that
      whoever built the simplified one cleared this whole sheet's own
      B:E and K:N ranges in one pass, not just the specific cells that
      referenced the deleted Module 2 columns -- e.g. "% of Section"
      (`=C2/54`) never referenced a flag cell at all, so the first
      version of this repair (which only found cells matching the
      flag-gated shape) never touched it, and it stayed blank -- which
      is what a real export's own "0% of section" (instead of the real
      percentage) turned out to be. Safe to restore unconditionally:
      confirmed this sheet holds no per-student data of its own at all,
      only formulas and the same fixed Domain/Skill labels confirmed
      identical in position to the current-format template's (College
      Board's own content taxonomy, not per-test content) -- and
      repaired_formula() is a no-op on a formula that was never
      flag-gated to begin with (nothing for either regex to match),
      not just on ones that are.
    - Everywhere else (e.g. "Student Responses"): only formulas matching
      the flag-gated shape -- deliberately still scoped this narrowly,
      unlike "Calculations". "Student Responses" holds a real student's
      own actual answers in `reference_wb`, and a *different* set of
      formula cells that were never affected by the Module 2 column
      deletions at all (each active block's own mark_col, e.g. D9:D35 --
      confirmed still present and correct on the simplified template,
      since narrowing this repair's own scope everywhere but
      "Calculations" is exactly what keeps this from ever touching
      them). The three non-canonical occurrences' own mark_col formulas
      (Y/AF/AM, confirmed present in `reference_wb` too) are the one
      category genuinely missing on the simplified template beyond what
      this restores, and deliberately left alone: those columns' own
      block no longer exists there at all, nothing on "Calculations" or
      "Score Report" reads them any more (confirmed: neither has any
      reference left to Y/AF/AM after repaired_formula's own dead-branch
      removal), so restoring them would just be writing dead formulas
      into columns nothing prints.

    Raises ValueError (via repaired_formula) if any matching formula
    doesn't repair cleanly."""
    writes: List[CellWrite] = []
    for sheet_name in reference_wb.sheetnames:
        ws = reference_wb[sheet_name]
        for row in ws.iter_rows():
            for cell in row:
                value = cell.value
                if not (isinstance(value, str) and value.startswith("=")):
                    continue
                if sheet_name != "Calculations" and not _FLAG_GATE_MARKER.search(value):
                    continue
                writes.append(CellWrite(sheet_name, cell.row, cell.column, repaired_formula(value)))
    return writes
