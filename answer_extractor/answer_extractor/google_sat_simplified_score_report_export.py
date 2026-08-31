"""Produce a filled-in *simplified* DSAT score-report PDF for one
student, end to end -- shares google_report_export_common.export_filled_report
with google_sat_score_report_export.py (the current-format path), but
supplies sat_simplified_score_report_writer.fill_simple_sat_score_report
as its fill step instead, and finds its own template by a fixed name
rather than by test code (see that module's own docstring for why: it
carries no per-test content of its own to make a new copy of for every
test -- there's exactly one of it, for now, regardless of test code).

Unlike every other export path in this package, this one also has to
find and download a *second*, entirely different template -- the
current-format one for the same test -- read-only, purely as a reference
source for each question's Correct Answer/Domain/Skill (see
sat_score_report_writer.read_reference_questions's own docstring for
why). `test_code` here is only ever used to find that reference
template; the simplified template being duplicated and filled is always
the same file no matter which test_code is passed.
"""
from __future__ import annotations

import datetime as dt
import os
import tempfile
from typing import Mapping, Optional

import openpyxl
from googleapiclient.discovery import Resource
from openpyxl.worksheet.worksheet import Worksheet

from .google_report_export_common import export_filled_report
from .google_sheets_export import export_xlsx
from .sat_score_report_writer import SatKey
from .sat_simplified_score_report_writer import fill_simple_sat_score_report
from .template_lookup import find_file_by_exact_name, find_template_file, resolve_template_folder

# The one simplified template's own exact file name in Drive -- see this
# module's own docstring for why there's only one, not one per test code.
SIMPLIFIED_TEMPLATE_NAME = "DSAT TEMPLATE"
# A folder of its own directly under the templates root, a sibling of
# "SAT" (not a subfolder of it) -- so find_template_file's own
# substring-against-test-code matching inside "SAT" itself is never at
# risk of also matching this file by accident.
SIMPLIFIED_TEMPLATE_CATEGORY_PATH = ["SAT Template"]


def _load_reference_worksheet(
    drive: Resource, templates_root_folder_id: str, test_code: str, sheet_name: str
) -> Worksheet:
    """Download the *current-format* SAT template for `test_code`,
    read-only, and return its `sheet_name` tab -- purely a reference
    source (see this module's own docstring); this never duplicates or
    writes to it, unlike every use of find_template_file elsewhere in
    this package."""
    folder_id = resolve_template_folder(drive, templates_root_folder_id, ["SAT"])
    template = find_template_file(drive, folder_id, test_code)
    fd, tmp_path = tempfile.mkstemp(suffix=".xlsx")
    os.close(fd)
    try:
        with open(tmp_path, "wb") as f:
            f.write(export_xlsx(drive, template["id"]))
        wb = openpyxl.load_workbook(tmp_path, data_only=True)
    finally:
        os.unlink(tmp_path)
    if sheet_name not in wb.sheetnames:
        raise ValueError(
            f"No {sheet_name!r} tab in the current-format template for test {test_code!r} (tabs: {wb.sheetnames})"
        )
    return wb[sheet_name]


def export_simple_sat_score_report(
    drive: Resource,
    sheets: Resource,
    templates_root_folder_id: str,
    test_code: str,
    answers: Mapping[SatKey, str],
    active_variants: Mapping[str, str],
    student_name: str,
    test_date: dt.date | str,
    section_scores: Mapping[str, int],
    output_name: str,
    temp_folder_id: Optional[str] = None,
    sheet_name: str = "Student Responses",
) -> bytes:
    """Return the filled simplified report's PDF bytes -- see
    google_report_export_common.export_filled_report for
    `temp_folder_id`/cleanup semantics, and
    sat_simplified_score_report_writer.fill_simple_sat_score_report for
    what every other argument means.

    `test_code` is used only to find the current-format template to read
    reference data from (via _load_reference_worksheet) -- the
    simplified template itself is always SIMPLIFIED_TEMPLATE_NAME,
    regardless of test_code (see this module's own docstring).

    Passes `fit_to_page=True` through to export_filled_report/export_pdf
    -- confirmed live the simplified template's own Cover Page splits
    across two PDF pages despite "Fit to page" already being its own
    saved setting; forcing that scale explicitly on export is a next
    attempt at closing that gap (see export_pdf's own docstring for the
    full reasoning, its own not-yet-confirmed-live status, and the risk
    it carries: this is a workbook-wide override, so it also overrides
    "Student Responses"' own fixed 54% scale, not just Cover Page's)."""
    reference_ws = _load_reference_worksheet(drive, templates_root_folder_id, test_code, sheet_name)

    simplified_folder_id = resolve_template_folder(
        drive, templates_root_folder_id, SIMPLIFIED_TEMPLATE_CATEGORY_PATH
    )
    simplified_template = find_file_by_exact_name(drive, simplified_folder_id, SIMPLIFIED_TEMPLATE_NAME)

    return export_filled_report(
        drive,
        sheets,
        templates_root_folder_id=None,
        category_path=None,
        test_code=None,
        output_name=output_name,
        fill_fn=lambda tmp_path: fill_simple_sat_score_report(
            tmp_path,
            reference_ws,
            answers,
            active_variants,
            student_name,
            test_date,
            test_code,
            section_scores=section_scores,
            sheet_name=sheet_name,
        ),
        temp_folder_id=temp_folder_id,
        template_id=simplified_template["id"],
        fit_to_page=True,
    )
