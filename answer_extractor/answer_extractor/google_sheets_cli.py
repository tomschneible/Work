"""One-off setup/maintenance utility for the Google Sheets score-report
export path -- not part of the main scan-to-spreadsheet pipeline. Three
jobs, deliberately combined into one command you run locally:

  1. Complete the interactive Google OAuth consent (see google_auth) --
     this has to happen wherever a real browser is available, which is
     never the environment this code gets written/reviewed in. Run it
     logged into the browser as the org's dedicated account for this
     program, not your own -- whichever Google account approves the
     consent screen is the one every future run acts as.
  2. List a Drive folder's contents by id, so a template file's own id
     can be identified (and hardcoded into config, once that config
     exists) from nothing more than the folder link you already have.
     Also doubles as a check that the account actually has access to the
     folder (an empty result usually means it hasn't been shared yet).

    python -m answer_extractor.google_sheets_cli list-folder --folder-id 1BqZHAVfpbHMW-g0HB5sn10GV3g3eCihK

  3. Turn off a template file's own gridlines directly, via
     google_sheets_export.hide_gridlines -- a pure Sheets API metadata
     change (no file conversion, nothing else on the file touched at
     all) -- a rare, deliberate one-time maintenance operation on the
     template itself, not something the per-report export pipeline does
     any more. An earlier version of this command instead downloaded the
     template as .xlsx, edited it locally, and re-uploaded the whole
     thing -- confirmed live that this corrupted the one real template it
     was tried against (recovered only via Sheets' own version history)
     -- so this never touches .xlsx at all any more. Run this once per
     template that needs it -- and once more against a master template
     before duplicating it for a new test code, so every future
     duplicate inherits the fix; an already-duplicated template still
     needs its own run. Takes one or more --file-id values so a whole
     batch of templates can be fixed in one command instead of editing
     and re-running this once per file; one file's failure (e.g. a typo'd
     id) is reported and skipped rather than stopping the rest.

    python -m answer_extractor.google_sheets_cli hide-gridlines --file-id 1BqZHAVfpbHMW-g0HB5sn10GV3g3eCihK
    python -m answer_extractor.google_sheets_cli hide-gridlines --file-id ID_ONE ID_TWO ID_THREE

  4. Repair the simplified SAT template's own "Student Responses"/
     "Calculations" formulas, via sat_simplified_template_repair.py --
     the same category of one-time, Sheets-API-only template fix as
     hide-gridlines, for a different problem (see that module's own
     docstring for the full why: the simplified template's Module 2
     column deletions broke every score-summary/Domain/Skill-breakdown
     formula that used to also count the three difficulty pairs that
     template no longer has). `--reference-file-id` is any real,
     working *current-format* template (its formula scheme is the same
     across every one, not test-specific) -- read-only, never modified;
     `--target-file-id` is the simplified template actually being fixed.

     Writes in chunks of _WRITE_CHUNK_SIZE cells per batchUpdate call --
     not one huge all-cells call, and not one call per cell either, a
     middle ground forced by two different failure modes, both confirmed
     live against the same real template. A single call covering every
     cell fails *entirely* (a 400, "trying to edit a protected cell") the
     moment even one of them is protected, silently hiding whether
     anything else in the same call would also be blocked. But swinging
     all the way to one call per cell instead -- tried first, since it
     does fix that -- runs one write request per cell against Sheets' own
     60-writes-per-minute-per-user quota (WriteRequestsPerMinutePerUser)
     with nothing pacing them, and for 216 cells reliably exceeds it well
     before finishing: confirmed live, a real run got 61/216 done before
     every remaining write started coming back 429 RATE_LIMIT_EXCEEDED.
     Chunking keeps the common case (nothing protected) to a handful of
     requests total -- comfortably under that quota with no pacing needed
     at all -- while still falling back to one write per cell, but *only*
     for whichever chunk itself failed, to isolate exactly which cell(s)
     in it are the problem -- the same guarantee the all-individual
     approach existed to provide, at a fraction of the request count. A
     429 hit at either level (a chunk, or an individual fallback write) is
     retried automatically with a cooldown a little over Sheets' own
     60-second quota window, up to a few times, before being reported as
     still rate-limited -- kept distinct from a genuine protected-cell
     report in the final summary, since the fix for each is different:
     just re-run the command for the former, change protection in Sheets
     for the latter.

    python -m answer_extractor.google_sheets_cli repair-simplified-calculations \\
      --reference-file-id 1BqZHAVfpbHMW-g0HB5sn10GV3g3eCihK --target-file-id 1AbCdEfGhIjKlMnOpQrS

The folder/file id is the long token in a Drive URL:
https://drive.google.com/drive/folders/<this part>?usp=drive_link
https://docs.google.com/spreadsheets/d/<this part>/edit
"""
from __future__ import annotations

import argparse
import io
import sys
import time
from typing import List, Optional, Sequence

import openpyxl
from googleapiclient.discovery import Resource
from googleapiclient.errors import HttpError
from openpyxl.utils import get_column_letter

from .google_sheets_export import CellWrite, build_services, export_xlsx, hide_gridlines, list_folder, write_cells
from .sat_simplified_template_repair import repair_calculations_writes

# repair-simplified-calculations' own write chunking/retry -- see
# _write_with_rate_limit_retry and _is_rate_limit_error below, and this
# module's own docstring (command 4) for the full reasoning.
_WRITE_CHUNK_SIZE = 20  # 216 cells / 20 -> 11 requests in the common (nothing protected) case --
# comfortably under the 60-writes-per-minute-per-user quota with no pacing needed at all; even a
# single bad chunk's own fallback (up to 20 more individual requests, see main() below) keeps a
# whole run's total request count well clear of that quota too.
_RATE_LIMIT_MAX_RETRIES = 3
_RATE_LIMIT_RETRY_SECONDS = 65  # a little over the 60-second window WriteRequestsPerMinutePerUser is measured over


def _is_rate_limit_error(exc: Exception) -> bool:
    """True for a Google API 429 ("RATE_LIMIT_EXCEEDED") -- confirmed live
    against a real repair-simplified-calculations run that (before
    chunking, see _WRITE_CHUNK_SIZE above) issued one write_cells call
    per cell and blew through Sheets' own 60-writes-per-minute-per-user
    quota well before 216 cells were done. Checked the same way
    google_report_export_common._cleanup_delete_is_actionable already
    checks its own different status code (404): `exc.status_code`, a
    convenience property googleapiclient's own HttpError exposes as
    `self.resp.status` -- not by parsing `reason`/`quota_metric` out of
    the error body, since status is the stable part of this across
    client versions. Deliberately narrow: a 400 "trying to edit a
    protected cell" (the other failure this command has hit live) is a
    different, *permanent* kind of failure that retrying would just
    reproduce identically every time -- see _write_with_rate_limit_retry
    for why that distinction matters."""
    return isinstance(exc, HttpError) and exc.status_code == 429


def _write_with_rate_limit_retry(sheets: Resource, target_file_id: str, cells: Sequence[CellWrite]) -> None:
    """write_cells, retrying only a transient 429 (_is_rate_limit_error)
    -- up to _RATE_LIMIT_MAX_RETRIES times, sleeping
    _RATE_LIMIT_RETRY_SECONDS between attempts. Anything else -- a
    protected-cell 400, or a 429 that still hasn't cleared after every
    retry -- propagates unchanged, so a caller sees exactly the same
    exception either way; this only ever adds retries around a transient
    failure, never changes what a non-retryable or exhausted one looks
    like."""
    attempts = 0
    while True:
        try:
            write_cells(sheets, target_file_id, cells)
            return
        except Exception as exc:
            attempts += 1
            if not _is_rate_limit_error(exc) or attempts > _RATE_LIMIT_MAX_RETRIES:
                raise
            print(
                f"  Rate-limited by Sheets (60 writes/minute) -- waiting {_RATE_LIMIT_RETRY_SECONDS}s "
                f"before retrying (attempt {attempts}/{_RATE_LIMIT_MAX_RETRIES})...",
                file=sys.stderr,
            )
            time.sleep(_RATE_LIMIT_RETRY_SECONDS)


def main(argv: Optional[List[str]] = None) -> int:
    parser = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    subparsers = parser.add_subparsers(dest="command", required=True)

    list_parser = subparsers.add_parser("list-folder", help="List every file in a Drive folder")
    list_parser.add_argument("--folder-id", required=True, help="Drive folder id, from its URL")

    hide_gridlines_parser = subparsers.add_parser(
        "hide-gridlines",
        help="Turn off a template file's gridlines directly, via a Sheets API metadata change only",
    )
    hide_gridlines_parser.add_argument(
        "--file-id",
        required=True,
        nargs="+",
        help="One or more Drive file ids of the templates to fix directly (not copies)",
    )

    repair_parser = subparsers.add_parser(
        "repair-simplified-calculations",
        help="Repair the simplified SAT template's own score-summary/Domain/Skill formulas",
    )
    repair_parser.add_argument(
        "--reference-file-id",
        required=True,
        help="A real, working current-format SAT template -- read-only, never modified",
    )
    repair_parser.add_argument(
        "--target-file-id",
        required=True,
        help="The simplified SAT template to fix directly (not a copy)",
    )

    args = parser.parse_args(argv)

    if args.command == "list-folder":
        drive, _sheets = build_services()
        files = list_folder(drive, args.folder_id)
        if not files:
            print("No files found -- either the folder is empty, or this Google account doesn't have access to it.")
            return 0
        for f in files:
            print(f"{f['id']}  {f['mimeType']:<45}  {f['name']}")
        return 0

    if args.command == "hide-gridlines":
        _drive, sheets = build_services()
        failures = 0
        for file_id in args.file_id:
            try:
                hide_gridlines(sheets, file_id)
            except Exception as exc:
                failures += 1
                print(f"Warning: couldn't hide gridlines on {file_id}: {exc}", file=sys.stderr)
            else:
                print(f"Hid gridlines on every sheet of {file_id}.")
        if len(args.file_id) > 1:
            print(f"{len(args.file_id) - failures}/{len(args.file_id)} succeeded.")
        return 1 if failures else 0

    if args.command == "repair-simplified-calculations":
        drive, sheets = build_services()
        reference_wb = openpyxl.load_workbook(
            io.BytesIO(export_xlsx(drive, args.reference_file_id)), data_only=False
        )
        writes = repair_calculations_writes(reference_wb)
        if not writes:
            print(
                f"No matching formulas found in {args.reference_file_id} -- "
                "double check it's a real current-format template, not the simplified one."
            )
            return 1
        failures = 0
        rate_limit_failures = 0
        by_sheet: dict[str, int] = {}
        for start in range(0, len(writes), _WRITE_CHUNK_SIZE):
            chunk = writes[start : start + _WRITE_CHUNK_SIZE]
            try:
                _write_with_rate_limit_retry(sheets, args.target_file_id, chunk)
            except Exception:
                # This chunk itself failed -- one blocked or still-rate-
                # limited cell fails the whole batched call (see this
                # command's own help text), hiding whether anything else
                # in it would have succeeded. Fall back to one write per
                # cell, only for this chunk, to find out which.
                for w in chunk:
                    coord = f"{w.sheet}!{get_column_letter(w.column)}{w.row}"
                    try:
                        _write_with_rate_limit_retry(sheets, args.target_file_id, [w])
                    except Exception as exc:
                        failures += 1
                        if _is_rate_limit_error(exc):
                            rate_limit_failures += 1
                            print(f"Warning: {coord} is still rate-limited after retrying.", file=sys.stderr)
                        else:
                            print(f"Warning: couldn't repair {coord}: {exc}", file=sys.stderr)
                    else:
                        by_sheet[w.sheet] = by_sheet.get(w.sheet, 0) + 1
            else:
                for w in chunk:
                    by_sheet[w.sheet] = by_sheet.get(w.sheet, 0) + 1
        succeeded = len(writes) - failures
        breakdown = ", ".join(f"{n} on {sheet!r}" for sheet, n in by_sheet.items()) or "none"
        print(f"Repaired {succeeded}/{len(writes)} formulas on {args.target_file_id} ({breakdown}).")
        if rate_limit_failures:
            print(
                f"{rate_limit_failures} cell(s) are still rate-limited after retrying -- Sheets' own "
                "60-writes-per-minute-per-user quota. Just re-run the exact same command again in a "
                "minute or two; already-repaired cells are harmless to write again, so this picks up "
                "the rest.",
                file=sys.stderr,
            )
        if failures - rate_limit_failures:
            print(
                f"{failures - rate_limit_failures} cell(s) couldn't be written -- likely a protected "
                "range on the target (Data -> Protected sheets and ranges in Sheets). Fix protection "
                "for the cells listed above and re-run; already-repaired cells are harmless to write "
                "again.",
                file=sys.stderr,
            )
        return 1 if failures else 0

    return 1


if __name__ == "__main__":
    sys.exit(main())
