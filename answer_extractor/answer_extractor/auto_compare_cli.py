"""Combined entry point for the "scan + compare" macOS droplet
(scripts/mac_droplet_compare.sh). Three ways to use it, auto-detected from
whatever you drop:

1. Scan + compare: a scanned bubble sheet (and/or a text-based
   score-report PDF) together with a reference report that has an
   independently-scored answer key -- either a spreadsheet with a
   "ScoreSheet" tab, or a rendered ScoreSheet-style PDF (see #3 below).
   Scans the sheet as normal, then adds a "Comparison" tab checking the
   answers against the reference.

2. Compare only: a spreadsheet this tool *already* exported (e.g. from an
   earlier run of the plain scan droplet) together with a reference report
   -- no re-scanning, since the answers are already sitting in that file.
   Just appends the "Comparison" tab to a copy of it.

3. Direct comparison: two already-finished score reports, checked directly
   against each other with nothing to scan at all -- e.g. this pipeline's
   own generated PDF report against a report you already had for that
   student. Each side can independently be a rendered ACT ScoreSheet-style
   PDF (this pipeline's own export, or any other report using the same
   repeated Question/Correct Answer/Your Answer/Category column-group
   layout -- see score_report_pdf_reader.py), a rendered SAT/DSAT "Your
   Question-Level Feedback" PDF (same idea, different column-group shape
   -- see sat_score_report_pdf_reader.py), or a spreadsheet with a
   "ScoreSheet" tab; a spreadsheet always plays the reference role (same
   as modes 1-2), a PDF plays whichever role isn't already taken by a
   spreadsheet, and with two PDFs and nothing else to disambiguate, the
   first one given on the command line (leftmost on the Automator drop,
   in whatever order Finder passes dropped files) is treated as "ours" and
   the second as "reference" -- the printed summary always names which
   file played which role, so it's never a silent guess. The two PDFs
   don't need to be the same shape (an ACT report against a SAT/DSAT one
   is accepted the same as any other pair) -- comparing a matching pair is
   what makes the result meaningful, not something this tool enforces.

    python -m answer_extractor.auto_compare_cli \\
        --input sheet.pdf reference.xlsx --template ... --output out.xlsx
    python -m answer_extractor.auto_compare_cli \\
        --input previous_answers.xlsx reference.xlsx --output out.xlsx
    python -m answer_extractor.auto_compare_cli \\
        --input our_report.pdf their_report.pdf --output out.xlsx

Whichever dropped .xlsx/.xlsm file contains the reference tab (default
"ScoreSheet") is treated as the reference; any *other* .xlsx/.xlsm is
treated as a pre-existing output to compare without scanning; a .pdf that
parses as a ScoreSheet-shaped grid (score_report_pdf_reader.py) is a
comparable report in its own right (see mode 3); everything else is
routed the same way answer_extractor.auto_cli routes it (images/PDFs
auto-detected as bubble sheets vs. text-based score reports to scan).
This is meant for the common one-student-at-a-time case, not batch
comparison -- more comparable candidates than a mode above can pair up,
or mixing a comparison pair with something to actually scan, is a clear
error rather than a guess at which files go together.

Mode 3's own two-PDF case tolerates one more real-world wrinkle: dropping
two files together (one Finder multi-selection, dragged in a single
motion) onto an Automator droplet built around this script can still
launch the underlying app once *per file* instead of once with both --
confirmed live on a real Mac (macOS Tahoe), and not specific to this
script (the unrelated plain scan droplet split the same way on the same
machine, and rebuilding the droplet from scratch didn't change it) --
so it's a real OS/Finder/Automator behavior this code has no way to fix
directly, not a bug in the comparison logic itself. Since a droplet
delivering the same two files as separate launches would otherwise leave
mode 3 permanently unusable on a Mac with this behavior (each launch only
ever sees one file, so "found a PDF but nothing to compare it against"
every time), a lone ScoreSheet-shaped PDF is instead remembered (see
_PENDING_COMPARE_MARKER) and automatically paired with the next one
dropped shortly after -- same result as if Finder had delivered both
together.
"""
from __future__ import annotations

import argparse
import json
import sys
import tempfile
import time
from pathlib import Path
from typing import List, Optional, Tuple

import openpyxl
from openpyxl import Workbook

from .answer_keys import annotate_rows, load_answer_keys
from .auto_cli import classify_inputs, scan_bubble_sheets, template_breakdown
from .export import add_bubble_sheet_answers_sheet
from .pipeline import SheetResult
from .sat_score_report_pdf_reader import parse_sat_score_report_pdf
from .score_report_export import add_score_report_answers_sheet
from .score_report_pdf_reader import parse_scoresheet_pdf
from .scoresheet_check import (
    add_comparison_sheet,
    compare,
    load_our_answers,
    load_reference_answers,
    ours_from_results,
    summarize,
)

_XLSX_SUFFIXES = {".xlsx", ".xlsm"}

# Confirmed live on a real Mac (macOS Tahoe): dropping two files together
# (selected as one Finder multi-selection, dragged in one motion) onto an
# Automator-built droplet app -- this one *and* the unrelated plain scan
# droplet, so it isn't anything about this script -- can still launch the
# app once *per file* instead of once with both, with no way found to stop
# it (a from-scratch rebuild of the droplet didn't help either). Since mode
# 3's own lone-PDF case used to just fail outright ("found a ScoreSheet-
# shaped PDF but nothing to compare it against") whenever that happens, two
# people can never actually use a direct-PDF-comparison droplet on a Mac
# with this behavior at all -- so instead of depending on a Finder/
# Automator fix neither this code nor its user has any control over, a
# lone comparable PDF is remembered here (a small marker file, not
# in-memory -- each split launch is its own separate process, so nothing
# in memory would survive between them) and automatically paired with the
# *next* one dropped within _PENDING_COMPARE_TIMEOUT_SECONDS -- the same
# result as if Finder had delivered both together, just built to tolerate
# it not doing that. See _pair_with_pending_drop.
_PENDING_COMPARE_MARKER = Path(tempfile.gettempdir()) / "answer_extractor_pending_compare.json"
_PENDING_COMPARE_TIMEOUT_SECONDS = 90


def _is_scoresheet_pdf(path: Path) -> bool:
    """Whether `path` is a rendered score-report grid this tool can
    compare directly -- ACT's own ScoreSheet-style layout
    (score_report_pdf_reader.py) or SAT/DSAT's own "Your Question-Level
    Feedback" layout (sat_score_report_pdf_reader.py), each this
    pipeline's own PDF export or any other report using the same
    column-group layout -- as opposed to a text-based score-report PDF
    meant to be scanned (e.g. a raw "Score Details" PDF, handled by
    auto_cli.classify_inputs instead), or not a PDF this tool understands
    at all."""
    if path.suffix.lower() != ".pdf":
        return False
    try:
        parse_scoresheet_pdf(path)
        return True
    except Exception:
        pass
    try:
        parse_sat_score_report_pdf(path)
        return True
    except Exception:
        return False


def _classify_report_sources(
    paths: List[Path], reference_tab: str
) -> Tuple[Optional[Path], List[Path], List[Path], List[Path]]:
    """Split `paths` into:

      - the one .xlsx/.xlsm with `reference_tab` (or None) -- always plays
        the reference role (modes 1-3);
      - every .pdf that parses as a ScoreSheet-shaped grid -- a comparable
        report in its own right (mode 3), playing whichever role (ours or
        reference) isn't already taken -- see _assign_comparison_roles;
      - every *other* .xlsx/.xlsm -- a pre-existing output of this tool's
        own, always plays the ours role (modes 1-2);
      - everything else (images, text-based score-report PDFs, etc.),
        handled by auto_cli.classify_inputs downstream.

    Raises ValueError if more than one file has `reference_tab` -- that's
    always ambiguous, regardless of what else was dropped.
    """
    tagged_references: List[Path] = []
    scoresheet_pdfs: List[Path] = []
    existing_outputs: List[Path] = []
    rest: List[Path] = []
    for p in paths:
        if p.suffix.lower() in _XLSX_SUFFIXES:
            try:
                wb = openpyxl.load_workbook(p, read_only=True)
            except Exception:
                rest.append(p)  # not a spreadsheet we can actually read -- let it fall through unchanged
                continue
            if reference_tab in wb.sheetnames:
                tagged_references.append(p)
            else:
                existing_outputs.append(p)
        elif _is_scoresheet_pdf(p):
            scoresheet_pdfs.append(p)
        else:
            rest.append(p)

    if len(tagged_references) > 1:
        raise ValueError(
            f"Found {len(tagged_references)} spreadsheets with a {reference_tab!r} tab "
            f"({', '.join(p.name for p in tagged_references)}) -- drop one reference at a time."
        )
    reference = tagged_references[0] if tagged_references else None
    return reference, scoresheet_pdfs, existing_outputs, rest


def _assign_comparison_roles(
    tagged_reference: Optional[Path], scoresheet_pdfs: List[Path], existing_outputs: List[Path]
) -> Optional[Tuple[Path, Path]]:
    """Decide which two files (if any) form a direct ours/reference pair
    that needs no scanning at all -- returns (ours_path, reference_path),
    or None if there isn't exactly one such pair. A spreadsheet's role is
    fixed by which bucket it's already in (a `reference_tab` spreadsheet
    is always the reference, any other spreadsheet is always ours -- same
    convention modes 1-2 have always used); a ScoreSheet-shaped PDF takes
    whichever role isn't already spoken for. Raises ValueError if there
    are more comparable candidates than a single pair can resolve (e.g.
    two PDFs *and* a tagged reference spreadsheet)."""
    candidates = ([tagged_reference] if tagged_reference else []) + scoresheet_pdfs + existing_outputs
    if len(candidates) > 2:
        raise ValueError(
            f"Found {len(candidates)} comparable score reports among the dropped files "
            f"({', '.join(p.name for p in candidates)}) -- drop at most two (an \"ours\" side "
            "and a reference) at a time, so it's clear which two go together."
        )
    if len(candidates) < 2:
        return None

    if tagged_reference and existing_outputs:
        return existing_outputs[0], tagged_reference
    if tagged_reference and scoresheet_pdfs:
        return scoresheet_pdfs[0], tagged_reference
    if existing_outputs and scoresheet_pdfs:
        return existing_outputs[0], scoresheet_pdfs[0]
    # Only remaining way to reach exactly 2 candidates: two ScoreSheet PDFs
    # and nothing else -- order given on the command line decides, and the
    # caller always prints which file played which role.
    return scoresheet_pdfs[0], scoresheet_pdfs[1]


def _pair_with_pending_drop(candidate: Path) -> Optional[Path]:
    """See _PENDING_COMPARE_MARKER's own comment for why this exists at
    all. If a still-fresh marker from an earlier, separate launch exists
    (and doesn't just name this same file again -- a launch retried or
    somehow duplicated shouldn't pair a file with itself), clears it and
    returns that earlier file, so the caller can run the comparison
    exactly as if both had arrived in one drop. Otherwise records
    `candidate` as the new pending file (overwriting whatever was there
    -- an unpaired file from over _PENDING_COMPARE_TIMEOUT_SECONDS ago is
    stale, not still waiting) and returns None.

    A corrupt or unreadable marker is treated the same as no marker at
    all -- this is a convenience for a real, confirmed OS quirk, not
    something worth ever failing a comparison over."""
    try:
        if _PENDING_COMPARE_MARKER.exists():
            recorded = json.loads(_PENDING_COMPARE_MARKER.read_text())
            recorded_path = Path(recorded["path"])
            age_seconds = time.time() - recorded["timestamp"]
            if age_seconds <= _PENDING_COMPARE_TIMEOUT_SECONDS and recorded_path != candidate:
                _PENDING_COMPARE_MARKER.unlink(missing_ok=True)
                return recorded_path
    except Exception:
        pass
    _PENDING_COMPARE_MARKER.write_text(json.dumps({"path": str(candidate), "timestamp": time.time()}))
    return None


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        description=(
            "Extract answers from a scanned bubble sheet and/or score-report PDFs (or reuse a "
            "spreadsheet this tool already exported), and if a reference report was dropped too "
            "(a spreadsheet or a rendered ScoreSheet-style PDF), compare the answers against it."
        )
    )
    parser.add_argument(
        "--input", required=True, nargs="+", help="One or more images, PDFs, spreadsheets, and/or directories"
    )
    parser.add_argument(
        "--template",
        default=None,
        help=(
            "Template YAML to use for every bubble-sheet input found, skipping auto-detection "
            "(by default, each sheet's template is auto-detected individually -- see "
            "answer_extractor.template_detect; unused if no bubble sheets are present)"
        ),
    )
    parser.add_argument("--output", required=True, help="Path to write the combined .xlsx to")
    parser.add_argument(
        "--reference-tab",
        default="ScoreSheet",
        help="Tab name that marks a dropped spreadsheet as the reference to compare against (default: ScoreSheet)",
    )
    parser.add_argument(
        "--no-refresh-keys",
        action="store_true",
        help=(
            "Skip fetching the latest answer key reference data over the network; use the "
            "last cached copy (or the one bundled in this checkout) instead"
        ),
    )
    return parser


def main(argv: list[str] | None = None) -> int:
    args = build_parser().parse_args(argv)

    input_paths = [Path(p) for p in args.input]
    missing = [p for p in input_paths if not p.exists()]
    if missing:
        for p in missing:
            print(f"Input path does not exist: {p}", file=sys.stderr)
        return 1

    try:
        reference_path, scoresheet_pdf_paths, existing_output_paths, remaining_paths = _classify_report_sources(
            input_paths, args.reference_tab
        )
        direct_pair = _assign_comparison_roles(reference_path, scoresheet_pdf_paths, existing_output_paths)
    except ValueError as exc:
        print(str(exc), file=sys.stderr)
        return 1
    bubble_paths, score_rows = classify_inputs(remaining_paths)

    if direct_pair and (bubble_paths or score_rows):
        ours_name, reference_name = direct_pair[0].name, direct_pair[1].name
        print(
            f"Found both something to scan and a ready-to-compare pair ({ours_name} vs "
            f"{reference_name}) among the dropped files -- drop one or the other, not both, "
            "so it's clear whether to scan or just compare.",
            file=sys.stderr,
        )
        return 1

    if direct_pair:
        # Mode 2/3: nothing to scan -- both sides are already-finished
        # reports (spreadsheets and/or ScoreSheet-shaped PDFs), compared
        # directly. load_our_answers/load_reference_answers each pick the
        # right parser by file extension (scoresheet_check.py).
        ours_path, reference_path_final = direct_pair
        reference = load_reference_answers(reference_path_final, sheet_name=args.reference_tab)
        ours = load_our_answers(ours_path)
        rows = compare(reference, ours)

        if ours_path.suffix.lower() in _XLSX_SUFFIXES:
            # Mode 2: the "ours" side is a pre-existing export of this
            # tool's own -- append the Comparison tab to a copy of it, so
            # its own answer tab(s) are carried over untouched.
            wb = openpyxl.load_workbook(ours_path)
        else:
            # Mode 3: the "ours" side is a PDF, with no answer tab of its
            # own to carry over -- the output is just the Comparison tab.
            wb = Workbook()
            del wb["Sheet"]
        add_comparison_sheet(wb, rows)
        wb.save(args.output)
        print(
            f"Wrote {args.output}: compared {ours_path.name} (ours) against "
            f"{reference_path_final.name} (reference)."
        )
        print(summarize(rows))
        return 0

    if existing_output_paths and (bubble_paths or score_rows):
        print(
            "Found both something to scan and an existing results spreadsheet "
            f"({', '.join(p.name for p in existing_output_paths)}) among the dropped files -- "
            "drop one or the other, not both, so it's clear whether to scan or just compare.",
            file=sys.stderr,
        )
        return 1

    if existing_output_paths:
        print(
            f"Found {existing_output_paths[0].name} but no reference report (a spreadsheet with a "
            f"{args.reference_tab!r} tab, or a ScoreSheet-shaped PDF) to compare it against.",
            file=sys.stderr,
        )
        return 1

    if scoresheet_pdf_paths and not (bubble_paths or score_rows):
        if len(scoresheet_pdf_paths) == 1:
            # Exactly one ScoreSheet-shaped PDF and nothing else -- see
            # _PENDING_COMPARE_MARKER's own comment for why this might be
            # one half of a two-file drop that got split into separate
            # launches, not really "nothing to compare against" at all.
            lone = scoresheet_pdf_paths[0]
            partner = _pair_with_pending_drop(lone)
            if partner is not None:
                # `partner` was dropped first (it's the one _pair_with_
                # pending_drop already had recorded) and `lone` second --
                # same "first given is ours, second is reference"
                # convention _assign_comparison_roles already uses when
                # both arrive in one drop.
                reference = load_reference_answers(lone, sheet_name=args.reference_tab)
                ours = load_our_answers(partner)
                rows = compare(reference, ours)
                wb = Workbook()
                del wb["Sheet"]
                add_comparison_sheet(wb, rows)
                wb.save(args.output)
                print(
                    f"Wrote {args.output}: compared {partner.name} (ours) against {lone.name} "
                    "(reference) -- dropped separately, paired automatically."
                )
                print(summarize(rows))
                return 0
            print(
                f"Got {lone.name} -- drop the file to compare it against within "
                f"{_PENDING_COMPARE_TIMEOUT_SECONDS} seconds (this Mac's own Finder/Automator is "
                "splitting the two-file drop into two separate launches instead of delivering them "
                "together -- see this module's own docstring). Nothing written yet.",
                file=sys.stderr,
            )
            return 0
        names = ", ".join(p.name for p in scoresheet_pdf_paths)
        print(
            f"Found a ScoreSheet-shaped PDF ({names}) but nothing to compare it against.",
            file=sys.stderr,
        )
        return 1

    if not bubble_paths and not score_rows:
        print(
            f"No images or PDFs found in: {', '.join(str(p) for p in input_paths)}",
            file=sys.stderr,
        )
        return 1

    wb = Workbook()
    del wb["Sheet"]
    summary_parts = []
    bubble_results: List[SheetResult] = []

    if bubble_paths:
        bubble_results = scan_bubble_sheets(bubble_paths, args.template)
        if bubble_results:
            add_bubble_sheet_answers_sheet(wb, bubble_results)
            summary_parts.append(f"{len(bubble_results)} bubble sheet(s){template_breakdown(bubble_results)}")

    if score_rows:
        try:
            library = load_answer_keys(refresh=not args.no_refresh_keys)
            score_rows = annotate_rows(score_rows, library)
        except Exception as exc:  # answer-key identification is a bonus, not required for extraction
            print(f"Warning: answer key identification skipped ({exc})", file=sys.stderr)
        add_score_report_answers_sheet(wb, score_rows)
        num_reports = len({row.source for row in score_rows})
        summary_parts.append(f"{len(score_rows)} score-report question(s) from {num_reports} file(s)")

    if not wb.sheetnames:
        print("No answers could be extracted from the given input.", file=sys.stderr)
        return 1

    # Mode 1: scan + compare -- reference_path (a tagged spreadsheet) or a
    # lone ScoreSheet-shaped PDF can each serve as the reference here.
    reference_side = reference_path or (scoresheet_pdf_paths[0] if scoresheet_pdf_paths else None)
    comparison_summary = None
    if reference_side is not None:
        if len(bubble_results) != 1:
            print(
                f"Found a reference report ({reference_side.name}) but "
                f"{len(bubble_results)} scanned bubble sheet(s) -- drop exactly one bubble "
                "sheet alongside a reference so the pairing isn't ambiguous. Wrote the scan(s) "
                "without a comparison.",
                file=sys.stderr,
            )
        else:
            reference = load_reference_answers(reference_side, sheet_name=args.reference_tab)
            ours = ours_from_results(bubble_results[0].questions)
            rows = compare(reference, ours)
            add_comparison_sheet(wb, rows)
            comparison_summary = summarize(rows)
            summary_parts.append(f"compared against {reference_side.name}")

    wb.save(args.output)
    print(f"Wrote {args.output}: {'; '.join(summary_parts)}.")
    if comparison_summary:
        print(comparison_summary)
    return 0


if __name__ == "__main__":
    sys.exit(main())
