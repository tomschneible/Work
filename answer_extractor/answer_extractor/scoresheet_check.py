"""Compare this tool's extracted answers against an independently-scored
reference "ScoreSheet" spreadsheet (e.g. a test-prep vendor's own scoring
export) -- an accuracy check external to the pipeline itself, working
purely off two .xlsx files on disk (no need to re-run the pipeline or keep
its in-memory results around).

Two different spreadsheet shapes are involved:

  - Our own output (answer_extractor.export.write_xlsx): one tab per
    sheet, "Question" + one column per section, one row per question.
    Flags (blank/MULTIPLE/pattern-inferred/unreadable/low-confidence) were
    encoded as cell fill color + italic font, not as data -- read back
    here via the exact same constants export.py wrote them with, so
    there's a single source of truth for what each color/style means.

  - The reference "ScoreSheet" tab: a vendor-specific grid of repeated
    (Question, Correct Answer, Your Answer, mark, Category) column-groups,
    several per section (each section's questions are split across two
    side-by-side blocks, e.g. English 1-35 and English 36-75), with the
    section name as a small standalone title cell a couple of rows above
    the leftmost block's header row -- one title covers every block to its
    right up to the next title. Parsed generically by locating every
    "Your Answer" header cell and walking down beneath it, rather than
    hardcoding row/column numbers, since block sizes differ by section
    (75 questions for English, 60 for Math, 40 each for Reading/Science)
    and this needs to keep working if a future export shifts row numbers.

The two sheets don't even agree on section names (the reference calls it
"Math", our template calls it "Mathematics") -- see
scoresheet_grid.normalize_section. Block-locating itself (the title/header
scan above) lives in scoresheet_grid.py, shared with score_report_writer.py,
which fills in a template's own "Your Answer" cells rather than reading a
vendor's back out.

A third shape can stand in for either side: a rendered score-report PDF
(this pipeline's own export, or a report someone already has) -- either
ACT's own ScoreSheet-style grid (score_report_pdf_reader.py) or SAT/DSAT's
own "Your Question-Level Feedback" grid (sat_score_report_pdf_reader.py),
each parsed into the same {(section, question): answer} shape
parse_reference_scoresheet produces from a .xlsx -- see
load_reference_answers/load_our_answers, which pick the right parser by
file extension (and, for a .pdf, by trying each reader in turn -- see
_load_pdf_answers) so callers (e.g. compare_cli) don't need to care
whether a given side is a .xlsx, an ACT PDF, or a SAT/DSAT PDF. A PDF
side carries no flag/low-confidence data (nothing in a finished report
says which answers the pipeline itself was unsure of), so it's always
treated as unflagged -- any mismatch against it comes out as
"silent_miss", same as it would for an unflagged xlsx cell.

A SAT/DSAT PDF's own section keys carry both subject *and* module (e.g.
"reading and writing module 1") -- see sat_score_report_pdf_reader.py's
own module docstring for why. _SECTION_ORDER lists those alongside ACT's
own four so compare()'s own row ordering groups a SAT/DSAT comparison by
section too, not just by raw question number across every section at
once.
"""
from __future__ import annotations

import dataclasses
from pathlib import Path
from typing import Dict, List, Optional

import openpyxl
from openpyxl.styles import Font, PatternFill
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from .detect import QuestionResult
from .export import BLANK_FILL, MULTIPLE_FILL, PATTERN_INFERRED_FILL, UNREADABLE_FILL, flag_for
from .sat_score_report_pdf_reader import parse_sat_score_report_pdf
from .score_report_pdf_reader import parse_scoresheet_pdf
from .scoresheet_grid import QuestionKey, iter_block_questions, locate_answer_blocks, normalize_section

_OMITTED_MARKS = {"ø", "o", "omitted", "-"}

_SECTION_ORDER = [
    "english",
    "mathematics",
    "reading",
    "science",
    "reading and writing module 1",
    "reading and writing module 2",
    "math module 1",
    "math module 2",
]


def parse_reference_scoresheet(
    path: str | Path, sheet_name: str = "ScoreSheet"
) -> Dict[QuestionKey, str]:
    """Parse a vendor "ScoreSheet" tab into {(normalized_section,
    question): your_answer}, with "" for an omitted question. Raises
    ValueError if the tab is missing, a header's section can't be
    determined, or two blocks disagree about the same question."""
    wb = openpyxl.load_workbook(path, data_only=True)
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"No {sheet_name!r} tab in {path} (tabs: {wb.sheetnames})")
    ws = wb[sheet_name]

    result: Dict[QuestionKey, str] = {}
    for block in locate_answer_blocks(ws):
        for r, question in iter_block_questions(ws, block):
            mark = ws.cell(row=r, column=block.mark_col).value
            your_answer = ws.cell(row=r, column=block.answer_col).value
            omitted = your_answer is None or (
                isinstance(mark, str) and mark.strip().lower() in _OMITTED_MARKS
            )
            answer = "" if omitted else str(your_answer).strip().upper()
            key = (block.section, question)
            if key in result and result[key] != answer:
                raise ValueError(
                    f"Conflicting entries for {key} in {sheet_name}: "
                    f"{result[key]!r} vs {answer!r}"
                )
            result[key] = answer
    return result


@dataclasses.dataclass(frozen=True)
class OurAnswer:
    answer: str
    flag: Optional[str]
    low_confidence: bool


def _flag_by_color() -> Dict[str, str]:
    return {
        MULTIPLE_FILL.start_color.rgb: "MULTIPLE",
        UNREADABLE_FILL.start_color.rgb: "unreadable",
        PATTERN_INFERRED_FILL.start_color.rgb: "pattern_inferred",
        BLANK_FILL.start_color.rgb: "blank",
    }


def parse_program_output(
    path: str | Path, tab_name: Optional[str] = None
) -> Dict[QuestionKey, OurAnswer]:
    """Parse one tab of this tool's own exported .xlsx (see export.py)
    back into {(normalized_section, question): OurAnswer}, reading the
    flag (blank/MULTIPLE/pattern_inferred/unreadable) back off the same
    cell fill colors export.py wrote, and low-confidence off the same
    italic font -- there's no other copy of that information once results
    are on disk instead of in memory. `tab_name` defaults to the
    workbook's first (active) tab.
    """
    wb = openpyxl.load_workbook(path)
    ws = wb[tab_name] if tab_name else wb.worksheets[0]
    flag_by_color = _flag_by_color()

    header = [cell.value for cell in ws[1]]
    sections = header[1:]

    result: Dict[QuestionKey, OurAnswer] = {}
    for row in ws.iter_rows(min_row=2):
        question = row[0].value
        if question is None:
            continue
        for col_idx, section in enumerate(sections, start=1):
            if section is None or col_idx >= len(row):
                continue
            cell = row[col_idx]
            fill_color = cell.fill.start_color.rgb if cell.fill else None
            flag = flag_by_color.get(fill_color)
            if cell.value is None and flag is None:
                continue  # not a real question for this section -- table padding
            answer = cell.value or ""
            low_confidence = bool(cell.font and cell.font.italic)
            result[(normalize_section(section), int(question))] = OurAnswer(
                answer, flag, low_confidence
            )
    return result


def _is_pdf(path: str | Path) -> bool:
    return str(path).lower().endswith(".pdf")


def _load_pdf_answers(path: str | Path) -> Dict[QuestionKey, str]:
    """Try the ACT-shaped reader first, then the SAT/DSAT-shaped one --
    whichever actually matches this PDF's own layout, so
    load_reference_answers/load_our_answers don't need to know in advance
    which kind of report a given .pdf is. Raises ValueError naming both
    readers' own failure reasons if neither recognizes it, rather than
    picking one arbitrarily to surface (a bare "not ACT-shaped" would be
    just as unhelpful as a bare "not SAT/DSAT-shaped" for a PDF that's
    actually neither, or genuinely not a score report at all)."""
    try:
        return parse_scoresheet_pdf(path)
    except Exception as act_exc:
        try:
            return parse_sat_score_report_pdf(path)
        except Exception as sat_exc:
            raise ValueError(
                f"{path} doesn't match either PDF shape this tool understands "
                f"(ACT ScoreSheet: {act_exc}; SAT/DSAT Question-Level Feedback: {sat_exc})"
            ) from sat_exc


def load_reference_answers(path: str | Path, sheet_name: str = "ScoreSheet") -> Dict[QuestionKey, str]:
    """parse_reference_scoresheet for a .xlsx, or _load_pdf_answers (an
    ACT or SAT/DSAT-shaped reader, whichever matches) for a .pdf -- picked
    by file extension so callers can take either kind of reference report
    without knowing in advance which one they have. `sheet_name` is
    ignored for a .pdf."""
    if _is_pdf(path):
        return _load_pdf_answers(path)
    return parse_reference_scoresheet(path, sheet_name=sheet_name)


def load_our_answers(path: str | Path, tab_name: Optional[str] = None) -> Dict[QuestionKey, OurAnswer]:
    """parse_program_output for a .xlsx, or _load_pdf_answers (wrapped as
    unflagged OurAnswers -- see module docstring) for a .pdf. `tab_name`
    is ignored for a .pdf."""
    if _is_pdf(path):
        return {
            key: OurAnswer(answer=answer, flag=None, low_confidence=False)
            for key, answer in _load_pdf_answers(path).items()
        }
    return parse_program_output(path, tab_name=tab_name)


def ours_from_results(questions: List[QuestionResult]) -> Dict[QuestionKey, OurAnswer]:
    """Build the same {(normalized_section, question): OurAnswer} shape as
    parse_program_output, but directly from this run's own QuestionResult
    objects rather than round-tripping through a written-out .xlsx's cell
    colors -- for callers (e.g. auto_compare_cli) that have both the scan
    and the reference available in the same process and don't need to
    write, close, and reopen a file just to compare them. Uses
    export.flag_for so the flag classification can't drift from what
    export.py would actually put in the cell."""
    return {
        (normalize_section(q.section), q.question): OurAnswer(q.answer, flag_for(q), q.low_confidence)
        for q in questions
    }


@dataclasses.dataclass(frozen=True)
class ComparisonRow:
    section: str
    question: int
    our_answer: str
    our_flag: Optional[str]
    our_low_confidence: bool
    reference_answer: Optional[str]
    # "match": answers agree.
    # "silent_miss": answers disagree and we gave no flag at all -- a
    #   confident wrong answer, the worst case (see module docstring in
    #   detect.py on false positives being worse than flagged blanks).
    # "flagged": answers disagree but we already flagged this one for
    #   review (blank/MULTIPLE/pattern_inferred/unreadable/low_confidence).
    # "unmatched": this question wasn't present in the reference sheet at
    #   all, so there's nothing to compare against.
    severity: str


def compare(
    reference: Dict[QuestionKey, str], ours: Dict[QuestionKey, OurAnswer]
) -> List[ComparisonRow]:
    keys = sorted(
        set(reference) | set(ours),
        key=lambda k: (_SECTION_ORDER.index(k[0]) if k[0] in _SECTION_ORDER else 99, k[1]),
    )
    rows: List[ComparisonRow] = []
    for key in keys:
        section, question = key
        ref_answer = reference.get(key)
        our = ours.get(key)
        our_answer = our.answer if our else ""
        our_flag = our.flag if our else "missing"
        our_low_confidence = our.low_confidence if our else False

        if ref_answer is None:
            severity = "unmatched"
        elif our_answer == ref_answer:
            severity = "match"
        elif our_flag:
            severity = "flagged"
        else:
            severity = "silent_miss"

        rows.append(
            ComparisonRow(
                section=section,
                question=question,
                our_answer=our_answer,
                our_flag=our_flag if our_flag != "missing" else None,
                our_low_confidence=our_low_confidence,
                reference_answer=ref_answer,
                severity=severity,
            )
        )
    return rows


_SEVERITY_FILL = {
    "silent_miss": PatternFill(start_color="F8D7DA", end_color="F8D7DA", fill_type="solid"),
    "flagged": PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid"),
    "unmatched": PatternFill(start_color="E2E3E5", end_color="E2E3E5", fill_type="solid"),
}
_MATCH_SYMBOL = {"match": "✔", "silent_miss": "✘", "flagged": "✘", "unmatched": "?"}


def add_comparison_sheet(wb: Workbook, rows: List[ComparisonRow], title: str = "Comparison") -> Worksheet:
    """Add a color-coded comparison tab to an existing workbook (used both
    standalone by write_comparison_report and alongside a sheet's own
    answer tabs by auto_compare_cli) and return it."""
    ws = wb.create_sheet(title=title)
    header = [
        "Section",
        "Question",
        "Our Answer",
        "Reference Answer",
        "Match",
        "Our Flag",
        "Low Confidence",
    ]
    ws.append(header)
    for cell in ws[1]:
        cell.font = Font(bold=True)

    for row in rows:
        ws.append(
            [
                row.section.title(),
                row.question,
                row.our_answer,
                row.reference_answer if row.reference_answer is not None else "(not in reference)",
                _MATCH_SYMBOL[row.severity],
                row.our_flag or "",
                "yes" if row.our_low_confidence else "",
            ]
        )
        fill = _SEVERITY_FILL.get(row.severity)
        if fill:
            for cell in ws[ws.max_row]:
                cell.fill = fill

    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["D"].width = 18
    ws.column_dimensions["F"].width = 18
    ws.freeze_panes = "A2"
    return ws


def write_comparison_report(rows: List[ComparisonRow], output_path: str | Path) -> None:
    wb = Workbook()
    del wb["Sheet"]  # remove the default blank sheet; add_comparison_sheet creates its own
    add_comparison_sheet(wb, rows)
    wb.save(str(output_path))


def summarize(rows: List[ComparisonRow]) -> str:
    total = len(rows)
    matches = sum(1 for r in rows if r.severity == "match")
    flagged = sum(1 for r in rows if r.severity == "flagged")
    silent = sum(1 for r in rows if r.severity == "silent_miss")
    unmatched = sum(1 for r in rows if r.severity == "unmatched")
    lines = [
        f"{total} questions compared: {matches} match, {flagged} flagged mismatch, "
        f"{silent} silent miss (unflagged wrong answer), {unmatched} not in reference.",
    ]
    if silent:
        worst = [f"{r.section.title()} {r.question} (ours={r.our_answer or 'blank'!r}, "
                 f"reference={r.reference_answer!r})" for r in rows if r.severity == "silent_miss"]
        lines.append("Silent misses (worth investigating first): " + "; ".join(worst))
    return "\n".join(lines)
