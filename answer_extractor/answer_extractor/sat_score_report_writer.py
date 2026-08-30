"""Fill a per-test SAT/DSAT score-report template's "Student Responses"
tab with one student's name, test date, and answers.

Structurally different from the ACT templates (see score_report_writer.py
and scoresheet_grid.py) in three ways, all confirmed against a real blank
template and its filled counterpart:

  - Every block has its own title directly above it, rather than one
    shared title governing several -- e.g. "R & W Module 2 - Higher
    Difficulty" titles just that one block, not a run of blocks to its
    right the way ACT's "English" does. Handled here by exploiting a
    layout fact that made this simpler than porting scoresheet_grid's
    title-attribution logic: a title's own column always equals its
    block's question-number column, so a title unambiguously identifies
    exactly one block by column alone.
  - Module 1 and Module 2 both number their questions starting at 1, so a
    block can't be identified by (subject, question) the way ACT's
    (section, question) works -- answers are keyed by
    (subject, module_slot, question) instead, module_slot being
    "module1", "easier", or "harder".
  - Each subject has two *pairs* of Module-2 blocks -- two "Higher
    Difficulty" blocks, two "Lower Difficulty" ones -- not one of each.
    Confirmed against the blank template: within a pair, the correct-
    answer key is byte-identical, so which twin is used is arbitrary;
    this always uses the leftmost and never touches its duplicate.

    A boolean flag cell above the block's own question-number column
    tells the sheet's own score formulas whether that column-group
    counts -- but there's only ONE row of these flags on the whole
    sheet (confirmed against the blank template: every subject's score
    formulas reference the exact same handful of cells, e.g. Math's own
    "correct count" formula reads $O$8/$V$8/$AC$8/$AJ$8, not a
    Math-specific row), reused by column position across every subject
    stacked underneath it. This is exactly why fill_sat_score_report
    consolidates *every* subject's real answers into one single column
    (_canonical_module2_col) rather than leaving each subject's own
    difficulty wherever it naturally sits: confirmed live that when
    Reading & Writing and Math administer *different* difficulties, their
    real answers land in different columns, leaving the two subjects'
    Module 2 tables visibly offset on the exported report -- and since
    every subject's score-count formula reads the *same* four flag cells
    by fixed address regardless of which subject or difficulty, there's
    no way to set a *different* subject-specific flag safely anyway.
    Consolidating means only the canonical column's own flag is ever set
    True; every other Module 2 occurrence is left untouched here and
    cleared and hidden instead (see blocks_to_clear and columns_to_hide
    -- both are needed, not just clearing: a cleared-but-not-hidden
    occurrence's blank columns still count toward the exported PDF's
    print area, forcing its "fit to page" scale down far more than the
    content actually left needs). Module 1 has no flag cell at
    all (everyone takes it, nothing to disambiguate). Flag cells are
    located here via one sheet-wide scan for boolean-valued cells, keyed
    by column -- not by searching near any one block's own header row,
    which is what a later-appearing subject's blocks actually need.

Which difficulty (easier/harder) was actually administered isn't
something this module figures out -- that identification already exists
in answer_keys.annotate_rows, matching a score report's own "Correct
Answer" column against a reference key. Callers pass the result in via
`active_variants`.

Like score_report_writer.py's ACT counterpart, this scans a *local,
read-only* copy of the template purely to find where things go and
returns the list of individual CellWrite values a caller pushes into the
live Sheet via google_sheets_export.write_cells -- it no longer edits or
returns a Workbook to be saved and re-uploaded wholesale (see
google_sheets_export.py's own module docstring for why).
"""
from __future__ import annotations

import dataclasses
import datetime as dt
import re
from pathlib import Path
from typing import Dict, List, Mapping, Optional, Tuple

import openpyxl
from openpyxl.worksheet.worksheet import Worksheet

from .google_sheets_export import CellWrite, FillResult, format_date_for_sheets

SatKey = Tuple[str, str, int]  # (subject, module_slot, question)

_NAME_PLACEHOLDER_PREFIXES = ("enter name", "type name here")

_TITLE_PATTERN = re.compile(
    r"^(?P<subject>.+?)\s+Module\s+(?P<module_num>1|2)"
    r"(?:\s*-\s*(?P<difficulty>Higher|Lower)\s+Difficulty)?\s*$",
    re.IGNORECASE,
)
# A block's full column span for clearing purposes: question, correct-answer,
# your-answer, mark -- SatBlock's own four columns -- plus Domain and Skill,
# two more columns immediately after mark_col that SatBlock doesn't track
# (nothing here ever reads or writes them) but that a cleared block still
# needs cleared too, or its Domain/Skill columns are left behind looking
# like an orphaned, unlabeled leftover even though every actual answer cell
# is gone. See blocks_to_clear, the only place this is used.
_CLEAR_BLOCK_WIDTH = 6
# How much narrower the visible answer-table columns (Module 1 through
# the canonical Module 2 block) get made on the working copy -- see
# visible_table_columns_to_narrow's own docstring for the full derivation.
# The first value tried, 0.75, was derived from: at "fit to page" (~55%
# scale, width the binding dimension -- its own rendered width already
# reached the page's full available width, while rendered height fell
# well short of the page's available height), reaching the target used
# by a real, well-filling reference example (~681pt of vertical space,
# matching this sheet's own real row-height sum of 1020pt scaled to
# ~67%) needs `0.55 / 0.667 =~ 0.824` for the *whole* visible print area
# -- but since the sidebar (Score Summary/Key/Domains) is deliberately
# left unnarrowed here (narrowing it risks clipping/misaligning a
# graphic element none of this ever otherwise touches), the tables alone
# have to shrink further to produce the same overall page-width
# reduction, working out to ~0.75 by the same math.
#
# Confirmed live that 0.75 overshot: real export at 0.75 filled the page
# far better (font size measurably bigger, matching the reference
# example's own fill level) but pushed a subject's last couple of
# questions onto a nearly-blank extra page, and exposed a template
# inconsistency (see allow_text_overflow) that truncated a block title
# outright once its column got that narrow. 0.82 (a pull back toward the
# original, whole-print-area estimate of 0.824) *also* overshot once the
# print area's own centering was separately fixed (hidden_columns_to_shrink
# closing the gap columns_to_hide's own spacer columns left open) --
# confirmed live, that fix alone raised the achievable scale enough that
# 0.82 pushed a subject's last several questions onto a mostly-blank
# extra page, a bigger overflow than 0.75 caused on its own. 0.90 pulled
# back further still, and got close: confirmed live, only the Math
# tables' last two rows (of 22 each; the Reading & Writing tables, with
# more rows but no multi-line wrapped answer cells inflating a couple of
# their row heights, fit in full) spilled onto an otherwise-empty extra
# page, with the printed content ending only ~26pt short of the page's
# own bottom margin (which mirrors its top margin almost exactly, so
# that 26pt is real overflow, not unused page space). 0.95 pulled back
# again -- confirmed live, down to only the single tallest row (the
# multi-line wrapped answer cell) still spilling over.
#
# At that point, switched from eyeballing overflow amounts to a
# scale-independent signal: a real export's own dominant body font size
# (measured directly off the rendered PDF -- font size scales linearly
# with the actual "fit to page" percentage regardless of how much
# content there is, unlike row/overflow counts) compared against the
# same measurement on a real, confirmed-good reference export of the
# same report (a different round of the same underlying test, never
# overflowing, its own Question-Level Feedback page filling all the way
# to a bottom margin matching its top margin almost exactly -- i.e. as
# tightly filled as this page is meant to get). 0.90 measured 6.31pt,
# 0.95 measured 6.11pt -- a consistent, linear -4.0pt of font size per
# +1.0 of factor across the only two live data points gathered since the
# centering fix (both above). The reference export measured 5.92pt.
# Solving that line for 5.92 gives `f =~ 0.9975` -- i.e., once the
# centering fix was in place, the table columns barely need narrowing at
# all to reach the same fill level as the reference; 1.0's own factor
# means the block "shrinks" every column to 1.0x its own current width,
# a pure no-op left in place (rather than special-cased away) so this
# constant stays the single tuning knob if a live export at 1.0 still
# doesn't quite match. Still not confirmed against a live export at this
# specific value -- may need pulling back slightly if it turns out this
# specific reference wasn't as tightly filled as it looked, since 1.0
# leaves no further room to reduce scale without this constant going
# above 1.0 (widening columns past their own original width, which
# narrow_columns supports mechanically but no case here has ever
# needed).
#
# Confirmed live at 1.0: font size landed at 5.93pt, matching the
# reference's 5.92pt almost exactly -- the font-size-matching approach
# was right. But it still wasn't quite enough: not a table row this
# time, just the page's own trailing footer line (a single row,
# directly below one blank spacer row -- nothing structural, no manual
# page break) spilled onto its own near-empty extra page. Measuring
# precisely *why*, rather than by eye, this time: the page's actual
# usable bottom edge is ~736pt (mirroring its own ~56pt top margin, and
# matching where the reference's own footer sits, right at 737pt); at
# 1.0, this export's last drawn content (a spacer row's own background
# fill, drawn even though the row itself is empty) ended at 729pt --
# only ~7pt of slack, for a footer line that itself needs ~7pt. Genuinely
# that close, not some other overlooked factor: matching font size means
# matching per-row height, so the ~7pt gap is really just accumulated
# rounding/measurement slop over ~65 rows' worth of content, not a real
# structural difference from the reference. 0.98 (confirmed live: font
# size 5.99pt) was a small enough pull back to absorb that gap -- until
# columns_to_hide's own missing-spacer fix (see its docstring) changed
# what this factor needed to be, below.
#
# That fix removed one column's worth of previously-un-narrowed,
# un-hidden natural width from the print area -- exactly the kind of
# width this factor's own derivation above has always had to compensate
# for, just not this particular piece of it (it was never counted in
# any of the six values tried above; it simply always happened to be
# there, quietly propping up the scale headroom every one of those
# values was calibrated against). Removing it lets "fit to page" compute
# a *larger* scale than before at the *same* factor -- confirmed live: a
# real export at the same 0.98, taken after that fix, measured 6.32pt,
# not 5.99pt -- overshooting the 5.92pt target upward instead of
# undershooting it, and spilling *more* content onto the extra page than
# 0.98 did before the fix, not less. Not a new problem, and not this
# factor being wrong in isolation -- the fix was correct on its own
# terms (it genuinely removed dead width the print area no longer needs
# to carry); this factor was simply calibrated against a baseline that
# included that width, and now needs recalibrating against one that
# doesn't.
#
# Redoing the same font-size-matching fit against that new baseline:
# the three pre-fix values above (0.90/6.31pt, 0.95/6.11pt, 1.00/5.93pt)
# fit a line in 1/font vs. factor almost exactly (predicted values
# within 0.02pt of each measured one), of the form `1/font = intercept +
# slope * factor` -- expected, since natural width is itself linear in
# this factor and font size scales linearly with "fit to page"'s own
# scale, i.e. inversely with natural width. Removing a fixed amount of
# width (this fix) shifts that line's intercept down by a constant
# (less natural width at every factor means a bigger 1/font at every
# factor) without changing its slope (the *narrowing* mechanism itself
# is unchanged) -- so the post-fix line's slope is the same one already
# fit from the three pre-fix points, and only its intercept needs
# re-anchoring, using the one post-fix data point available (0.98 ->
# 6.32pt) to solve for it. Doing that and solving the resulting line for
# 5.92pt gives `f =~ 1.085` -- i.e., with the dead spacer gone, these
# columns now need *widening* slightly past their own natural width to
# reach the same fill level as the reference, not narrowing at all.
# Widening past 1.0 is mechanically identical to narrowing as far as
# narrow_columns is concerned (see its own docstring) -- just never
# needed before this. Rounded to 1.09; not yet confirmed live at this
# specific value, unlike the three pre-fix points this line was fit
# from, and it extrapolates past the 0.90-1.00 range every one of those
# came from, so treat it as the next data point to confirm, the same as
# every value before it.
_TABLE_COLUMN_NARROW_FACTOR = 1.09
# How much the *hidden* non-canonical Module 2 columns (columns_to_hide)
# get narrowed too, on top of being marked hidden -- see
# hidden_columns_to_shrink's own docstring for why hiding alone wasn't
# enough. Small enough to floor at 1px (narrow_columns' own minimum)
# for any realistic starting width on this sheet.
_HIDDEN_COLUMN_SHRINK_FACTOR = 0.01
# Written into every active block's own mark_col header cell (blank in
# the template -- there's no label over the ✔/✘ column, unlike every
# neighboring one) so it's no longer blank -- see
# fill_sat_score_report's own docstring for why this, rather than any
# column-width fix, is what actually stops that cell from losing its own
# left border in the exported PDF. A zero-width space: real content as
# far as a cell's own blank/non-blank state is concerned, but nothing
# visibly renders in the cell itself.
_MARK_HEADER_NON_BLANK = "​"  # U+200B ZERO WIDTH SPACE
_SUBJECT_ALIASES = {
    "r & w": "reading and writing",
    "r and w": "reading and writing",
    "reading & writing": "reading and writing",
    "reading and writing": "reading and writing",
    "math": "math",
}
_DIFFICULTY_TO_SLOT = {"higher": "harder", "lower": "easier"}
_HEADER_SEARCH_ROWS = 5  # how far below a title to look for its "Your Answer" header

_SCORE_LABEL_PATTERN = re.compile(r"^(?P<subject>.+?)\s*score\s*$", re.IGNORECASE)
_SCORE_VALUE_SEARCH_ROWS = 5  # how far above a "<Subject> Score" label to look for its value cell


def normalize_subject(raw: str) -> str:
    key = re.sub(r"\s+", " ", raw.strip().lower())
    if key not in _SUBJECT_ALIASES:
        raise ValueError(f"Unrecognized SAT subject title {raw!r}")
    return _SUBJECT_ALIASES[key]


@dataclasses.dataclass(frozen=True)
class SatBlock:
    subject: str
    module_slot: str  # "module1" | "easier" | "harder"
    title_row: int
    header_row: int
    question_col: int
    correct_col: int
    answer_col: int
    mark_col: int
    flag_cell: Optional[Tuple[int, int]]  # (row, col); None for module1, which needs no flag


def _scan_raw_titles(ws: Worksheet) -> List[Tuple[int, int, str, str]]:
    """Every block title found anywhere on `ws`, as (row, col, subject,
    module_slot) -- *not* deduplicated (see locate_sat_blocks, which keeps
    only the leftmost per (subject, module_slot); blocks_to_clear needs
    every occurrence's own column, including duplicates/twins, to know
    what to clear). Raises ValueError if a Module 2 title is missing its
    Higher/Lower difficulty."""
    raw_titles: List[Tuple[int, int, str, str]] = []
    for row in ws.iter_rows():
        for cell in row:
            value = cell.value
            if not isinstance(value, str):
                continue
            match = _TITLE_PATTERN.match(value.strip())
            if not match:
                continue
            subject = normalize_subject(match.group("subject"))
            if match.group("module_num") == "1":
                module_slot = "module1"
            else:
                difficulty = match.group("difficulty")
                if difficulty is None:
                    raise ValueError(f"Module 2 title missing a Higher/Lower difficulty: {value!r}")
                module_slot = _DIFFICULTY_TO_SLOT[difficulty.lower()]
            raw_titles.append((cell.row, cell.column, subject, module_slot))
    return raw_titles


def _find_header_row(ws: Worksheet, title_row: int, title_col: int) -> int:
    """The row within _HEADER_SEARCH_ROWS of `title_row` whose
    `title_col + 2` cell reads "Your Answer" -- used by locate_sat_blocks
    to find one block's own header row. Raises ValueError if it can't be
    found."""
    for candidate_row in range(title_row, title_row + _HEADER_SEARCH_ROWS + 1):
        if ws.cell(row=candidate_row, column=title_col + 2).value == "Your Answer":
            return candidate_row
    raise ValueError(f"Could not find a 'Your Answer' header below the title at row {title_row}, column {title_col}")


def locate_sat_blocks(ws: Worksheet) -> List[SatBlock]:
    """Scan `ws` for every SAT block, deduplicated to exactly one per
    (subject, module_slot) -- see module docstring on why a subject's two
    same-difficulty blocks are interchangeable, and why only the leftmost
    is kept. Raises ValueError if a title's own "Your Answer" header can't
    be found nearby, or a Module 2 title is missing its difficulty."""
    raw_titles = _scan_raw_titles(ws)

    # One sheet-wide scan for the flag cells, keyed by column -- not
    # searched relative to any one block's own header row, since a
    # later-appearing subject's blocks reuse an earlier subject's flags
    # rather than having their own (see module docstring). Keeps the
    # topmost boolean found per column, matching what every subject's
    # score formulas actually reference.
    flag_cell_by_col: Dict[int, Tuple[int, int]] = {}
    for row in ws.iter_rows():
        for cell in row:
            if isinstance(cell.value, bool) and cell.column not in flag_cell_by_col:
                flag_cell_by_col[cell.column] = (cell.row, cell.column)

    blocks_by_key: Dict[Tuple[str, str], SatBlock] = {}
    for title_row, title_col, subject, module_slot in raw_titles:
        key = (subject, module_slot)
        if key in blocks_by_key and title_col >= blocks_by_key[key].question_col:
            continue  # a duplicate block further right -- keep the leftmost one already found

        header_row = _find_header_row(ws, title_row, title_col)
        flag_cell = flag_cell_by_col.get(title_col) if module_slot != "module1" else None

        blocks_by_key[key] = SatBlock(
            subject=subject,
            module_slot=module_slot,
            title_row=title_row,
            header_row=header_row,
            question_col=title_col,
            correct_col=title_col + 1,
            answer_col=title_col + 2,
            mark_col=title_col + 3,
            flag_cell=flag_cell,
        )
    return list(blocks_by_key.values())


def _canonical_module2_col(ws: Worksheet) -> Optional[int]:
    """The single column every subject's real Module 2 answers get
    consolidated into, regardless of which difficulty was actually
    administered -- always the leftmost Module 2 title column found (this
    template's own layout puts a subject's Higher-difficulty canonical
    block first; see fill_sat_score_report's own docstring for why every
    subject's real pick has to land in the same column position). None if
    the template has no Module 2 blocks at all."""
    cols = [block.question_col for block in locate_sat_blocks(ws) if block.module_slot != "module1"]
    return min(cols) if cols else None


def _non_canonical_module2_cols(ws: Worksheet) -> List[int]:
    """Every Module 2 title column found anywhere on `ws` (1-indexed)
    except _canonical_module2_col -- shared by blocks_to_clear and
    columns_to_hide, since both need to act on exactly the same set of
    occurrences (see columns_to_hide's own docstring for why a cleared
    occurrence's columns also need hiding, not just clearing)."""
    canonical_col = _canonical_module2_col(ws)
    return sorted(
        {col for _row, col, _subject, module_slot in _scan_raw_titles(ws) if module_slot != "module1"}
        - {canonical_col}
    )


def blocks_to_clear(ws: Worksheet) -> List[Tuple[int, int, int, int]]:
    """0-indexed (start_row, end_row, start_col, end_col) rectangles (end
    exclusive -- the shape a Sheets API GridRange needs) covering every
    Module 2 column position that isn't _canonical_module2_col -- every
    subject's own non-canonical difficulty, and every duplicate/twin --
    for the sheet's *entire* height (row 0 through `ws.max_row`), not just
    the rows a block occurrence's own questions happen to occupy.

    Whole-column, not per-occurrence, deliberately: an earlier version of
    this scoped each rectangle to one occurrence's own row range, because
    at the time a non-canonical column could still legitimately hold a
    *different* subject's real answers (this was before
    fill_sat_score_report consolidated every subject's real data into one
    column -- see its own docstring). Now that nothing real is ever left
    at a non-canonical column for *any* subject, there's nothing left to
    protect by scoping to rows -- confirmed live against a real template:
    a non-canonical column still held a handful of things outside any
    block occurrence's own question rows entirely (a boolean-valued flag
    cell, whose checkbox *widget* persisted even after its value was
    cleared -- see google_sheets_export.clear_cells's own docstring on why
    that needed a data-validation clear too; and a repeated "Page X of Y"
    footer label sitting well below the last question row), both of which
    a row-scoped rectangle silently left behind, still occupying enough
    of the sheet to force the exported PDF to scale down to fit them.
    Clearing the whole column removes all of it in one pass, whatever it
    turns out to be, without needing to enumerate each kind by hand.

    Each rectangle spans the full 6-column block width (title, correct
    answer, your answer, mark, Domain, Skill -- see _CLEAR_BLOCK_WIDTH).
    Module 1 is never included -- every student takes it, and it's never
    duplicated, so its own column is always canonical-equivalent (nothing
    else there needs to be canonicalized in the first place).
    """
    non_canonical_cols = _non_canonical_module2_cols(ws)
    return [(0, ws.max_row, col - 1, col - 1 + _CLEAR_BLOCK_WIDTH) for col in non_canonical_cols]


def columns_to_hide(ws: Worksheet) -> List[Tuple[int, int]]:
    """0-indexed (start_col, end_col) column ranges (end exclusive --
    the shape a Sheets API dimension range needs) covering every
    non-canonical Module 2 column *and* the spacer columns between and
    after them -- one single contiguous range starting right where the
    canonical block's own last column ends (visible_table_columns_to_narrow's
    own end, so the spacer between the two abuts it with no gap) through
    the rightmost occurrence's own last column, empty if there are no
    non-canonical occurrences at all.

    The range starts at the *canonical* block's own end, not the first
    non-canonical occurrence's own title column (an earlier version of
    this): that left exactly one column -- the spacer between the
    canonical block and the first non-canonical occurrence -- outside
    both this range (which started one column later) and
    visible_table_columns_to_narrow's own range (which stops at the
    canonical block's own last column), unlike every *other*
    inter-occurrence spacer and the trailing one, which this range's own
    contiguous span already swept up (see below). Structurally identical
    to those other spacers -- confirmed against a real template's own
    column layout -- so there's no reason for this one alone to be
    exempt; it was just outside where either neighboring range happened
    to start/end. Left at full natural width, unlike everything either
    side of it, it inflated the print area's own natural size the same
    way the spacers this function already sweeps up used to before this
    function existed -- diagnosed from a real export's own rendered
    geometry (extended header bar and real table content both ending
    ~27-30pt short of the print area's actual right edge, symmetric on
    both sides from Sheets' own horizontal centering) rather than a live
    A/B export comparison the way the other lessons in this module were;
    still worth confirming against one.

    Needed *in addition to* blocks_to_clear, not instead of it:
    clearing removes an occurrence's own cell values, but its columns
    are still fully present -- and still full width -- in the sheet's
    print area, since clearing never touches column width or
    visibility. Confirmed live against a real export: with every
    non-canonical occurrence cleared but not hidden, the exported PDF's
    "fit to page" scale was still being computed against a print area
    almost four times as wide as the one Module 2 block per subject
    that's actually left with content, forcing that scale down far more
    than the real content needed and leaving it squeezed into a small
    corner of the page with a large blank margin around it. Hiding
    these same columns removes them from the print area entirely, so
    "fit to page" scales to what's actually left to show.

    Contiguous rather than one range per occurrence (an earlier version
    of this returned three separate _CLEAR_BLOCK_WIDTH-wide ranges,
    leaving the single blank spacer column between each pair of
    occurrences untouched at its own full width) because that gap turned
    out to matter for a *different* reason than clearing/hiding was
    originally about: confirmed live, with `printOptions
    horizontalCentered="1"` already set on the real template (this
    pipeline never sets it, so it was always meant to center the print
    area horizontally), those leftover full-width spacer columns were
    still counted toward whatever Sheets centers against, visibly
    pushing the real content left of center. A single contiguous range
    removes them too, without needing to enumerate every spacer by hand.

    Safe to do unconditionally (every non-canonical column, for every
    subject) for the same reason blocks_to_clear now is: since
    fill_sat_score_report consolidates every subject's real answers
    into _canonical_module2_col, nothing real is ever left in any other
    Module 2 column position for any subject -- see hide_columns' own
    docstring for why this wasn't true, and this approach wasn't safe,
    before that consolidation existed. The spacer columns swept up along
    the way never held anything to begin with, in any state.
    """
    non_canonical_cols = _non_canonical_module2_cols(ws)
    if not non_canonical_cols:
        return []
    canonical_col = _canonical_module2_col(ws)
    assert canonical_col is not None  # a non-canonical occurrence exists, so canonical must too
    start = canonical_col - 1 + _CLEAR_BLOCK_WIDTH
    end = max(non_canonical_cols) - 1 + _CLEAR_BLOCK_WIDTH
    return [(start, end)]


def hidden_columns_to_shrink(ws: Worksheet) -> List[Tuple[int, int, float]]:
    """0-indexed (start_col, end_col, factor) ranges (end exclusive) --
    the exact same non-canonical Module 2 columns columns_to_hide hides,
    narrowed to _HIDDEN_COLUMN_SHRINK_FACTOR (near-zero) via
    narrow_columns *in addition to* being marked hidden.

    Exists because marking a column `hiddenByUser` turned out not to be
    enough on its own for a *different* reason than either of the row
    lessons (trailing_rows_to_delete/delete_rows) or the width-vs-height
    scale lesson (visible_table_columns_to_narrow): confirmed live
    against a real export, with `printOptions horizontalCentered="1"`
    already set on this sheet (confirmed in the real template's own
    saved xlsx -- this pipeline never sets it, so it was always meant to
    center the print area horizontally), the *visible* answer tables
    rendered pushed to the left of center, with a large blank gap on the
    right roughly matching where the hidden Module 2 occurrences used to
    sit. Hiding a column removes it from what gets *drawn*, but
    apparently not from what Sheets still counts as the print area's own
    width when centering it -- the same kind of gap between "hidden" and
    "actually gone" already confirmed for trailing rows. Giving the same
    hidden columns an explicit near-zero width too closes that gap
    without needing to touch anything about the print setup itself.
    """
    return [
        (start, end, _HIDDEN_COLUMN_SHRINK_FACTOR)
        for start, end in columns_to_hide(ws)
    ]


def header_bar_extension(ws: Worksheet) -> Optional[Tuple[int, str, int, int]]:
    """(row 0-indexed, hex fill color, start_col 0-indexed, end_col
    0-indexed exclusive) for extending a decorative full-row fill (the
    accent bar under "Your Question-Level Feedback") to cover the same
    visible answer-table width visible_table_columns_to_narrow narrows
    -- via extend_fill. None if row 1 has no such fill to extend at all
    (a template that doesn't use this design).

    Exists because narrowing the answer tables' own columns
    (visible_table_columns_to_narrow) has a side effect on anything else
    that happens to be sized by those *same* columns: confirmed live
    against a real export, this sheet's own row-1 accent bar -- a solid
    fill spanning a fixed range of columns, confirmed against the real
    template file, from column A through N -- is exactly as wide as the
    sum of those columns' widths, so narrowing the table columns inside
    that same range (Module 1's own, from column H on) shrank the bar
    along with them, leaving it visibly short of the width it used to
    reach while the page title's own text sat above it unchanged.
    Extending the fill to cover the rest of the narrowed table's own
    width (through the canonical Module 2 block's own last column, not
    just where the bar originally stopped at Module 1's own end) makes
    it span the same width as the table content it sits above, which
    reads as more consistent than trying to reproduce its original
    (now-narrower) absolute pixel width exactly.

    The color is read directly from the sheet's own row-1 fill (via
    `openpyxl`, the same local, read-only copy every other scan here
    uses) rather than hardcoded, so this doesn't assume any specific
    template's own color choice."""
    last_col = 0
    color: Optional[str] = None
    for col in range(1, ws.max_column + 1):
        cell = ws.cell(row=1, column=col)
        fill = cell.fill
        fg = fill.fgColor if fill else None
        if fill and fill.patternType == "solid" and fg and fg.rgb not in (None, "00000000", "FFFFFFFF"):
            last_col = col
            color = fg.rgb
    if color is None:
        return None

    canonical_col = _canonical_module2_col(ws)
    blocks = locate_sat_blocks(ws)
    module1 = next((b for b in blocks if b.module_slot == "module1"), None)
    if module1 is None:
        return None
    target_end = (canonical_col + _CLEAR_BLOCK_WIDTH - 1) if canonical_col is not None else module1.question_col
    if target_end <= last_col:
        return None  # already covers (or exceeds) the narrowed table's own width
    # `last_col` (1-indexed) is numerically identical to the 0-indexed
    # start of the column right after it -- e.g. the bar's last filled
    # column is N (14, 1-indexed); O, the first one still needing fill,
    # is 15 (1-indexed) = 14 (0-indexed). Extending from `last_col` as a
    # bare 0-indexed start is this identity, not a coincidence to
    # untangle further.
    return (0, str(color), last_col, target_end)


def visible_table_columns_to_narrow(ws: Worksheet) -> List[Tuple[int, int, float]]:
    """0-indexed (start_col, end_col, factor) ranges (end exclusive -- the
    shape a Sheets API dimension range needs, plus narrow_columns' own
    shrink factor) spanning every *visible* answer-table column on `ws`
    -- Module 1 through the canonical Module 2 block's own last column
    (Domain/Skill) -- to shrink via narrow_columns. Empty if `ws` has no
    Module 1 block at all (nothing to narrow).

    This exists because hiding/deleting non-canonical Module 2 columns
    and a sheet's own trailing blank rows (columns_to_hide,
    trailing_rows_to_delete) turned out not to be the whole story:
    confirmed live, a real export with both of those already applied
    still left the answer tables confined to roughly the top ~86% of the
    page's usable height, the leftover space stranded below rather than
    filling it. The cause: Sheets' "fit to page" print scale is computed
    from the print area's *natural* (unscaled) size, and applies
    *uniformly* to both width and height -- confirmed live that width was
    the tighter of the two even after the fixes above (the real export's
    rendered table width already reached the page's full available
    width at "fit to page"'s own ~55% scale, while its rendered height
    fell well short of the page's available height at that same scale),
    so the scale needed to keep width inside the page was smaller than
    height alone would have required, and that same undersized scale
    then left height under-filled too.

    Rows and columns are *not* interchangeable for a fix like this --
    confirmed live, the hard way: hiding and then deleting a sheet's own
    trailing rows (see trailing_rows_to_delete/delete_rows) had *zero*
    measurable effect on the exported PDF's scale, while hiding columns
    measurably did. Narrowing the answer tables' own columns further
    (rather than touching row heights or font sizes directly) targets
    the dimension already confirmed to move the needle, on the theory
    that it's still what's forcing the scale down: shrinking these
    columns lets "fit to page" recompute a *larger* uniform scale on its
    own (still guaranteed not to overflow -- "fit to page" always finds
    whatever scale fits, however large or small the natural size is),
    which then renders the *untouched* rows taller too, filling more of
    the page's actual height as a side effect -- without narrowing
    anything, adjusting any print setting, or resizing a single font.

    Each active block's own `mark_col` (the ✔/✘ column, already one of
    the narrowest on the sheet -- it holds a single glyph) is
    deliberately *excluded* from every range here, unlike every other
    sub-column of the block -- confirmed live that its own header cell
    (blank -- there's no label over the mark column, unlike every
    neighboring one) can lose its left border entirely in the exported
    PDF, while the *data* rows below it (never blank -- always holding a
    ✔ or ✘) never do. Excluding it from narrowing costs very little of
    the overall size gain (it only ever holds one glyph) on the chance
    that width is part of what triggers this.

    Width alone isn't the actual fix, though -- confirmed live, across
    two different attempts (excluding this column from narrowing at all,
    then actively *widening* it well past its own original size): the
    border loss persisted regardless, and even showed up on Module 1's
    own mark_col at one point, a column never touched by any narrowing
    at any point in this whole investigation. Whatever real threshold
    (if any) governs this isn't reliably tied to column width the way
    the rest of this module's own fixes are. The fix that actually
    worked is in fill_sat_score_report itself: writing a real,
    zero-width character into each active block's own mark_col header
    cell so it's no longer *blank* -- see _MARK_HEADER_NON_BLANK's own
    comment. This function still excludes the column from narrowing
    anyway, since there's no cost to doing so and it removes one more
    variable from the picture.

    The sidebar (Score Summary/Key/Domains) is deliberately excluded
    too -- only Module 1's own title column through the canonical
    Module 2 block's own last column narrows. See
    _TABLE_COLUMN_NARROW_FACTOR's own comment for where that factor's
    value comes from and how confident it actually is (a first estimate,
    not one confirmed against a live export)."""
    blocks = locate_sat_blocks(ws)
    module1 = next((b for b in blocks if b.module_slot == "module1"), None)
    if module1 is None:
        return []
    canonical_col = _canonical_module2_col(ws)

    ranges: List[Tuple[int, int, float]] = []
    # Module 1: title through answer (3 cols), skip mark_col, then domain
    # through skill *and* the spacer column(s) up to wherever Module 2's
    # own canonical block starts (or just domain/skill if there's no
    # Module 2 at all on this sheet).
    m1 = module1.question_col
    ranges.append((m1 - 1, m1 - 1 + 3, _TABLE_COLUMN_NARROW_FACTOR))
    domain_end = (canonical_col - 1) if canonical_col is not None else (m1 - 1 + _CLEAR_BLOCK_WIDTH)
    ranges.append((m1 - 1 + 4, domain_end, _TABLE_COLUMN_NARROW_FACTOR))

    if canonical_col is not None:
        ranges.append((canonical_col - 1, canonical_col - 1 + 3, _TABLE_COLUMN_NARROW_FACTOR))
        ranges.append(
            (canonical_col - 1 + 4, canonical_col - 1 + _CLEAR_BLOCK_WIDTH, _TABLE_COLUMN_NARROW_FACTOR)
        )
    return ranges


def trailing_rows_to_delete(ws: Worksheet) -> List[Tuple[int, int]]:
    """0-indexed (start_row, end_row) row range (end exclusive -- the
    shape a Sheets API dimension range needs) covering every row below
    `ws`'s own last real content -- empty if there's no such gap.

    A completely different problem than blocks_to_clear/columns_to_hide:
    not a Module 2 occurrence that wasn't administered, but the sheet's
    own trailing rows that were never real content in the first place.
    Confirmed against the real template: "Student Responses" carries
    formatting (row heights, borders) all the way out to row 996, even
    though its actual content -- every block, every score cell, the
    footer -- ends at row 64. With no print area explicitly set on the
    file (see delete_rows' own docstring on why this pipeline doesn't set
    one directly), Sheets' PDF export falls back to the sheet's full used
    range for that tab, so those ~930 empty rows were still being
    counted when "fit to page" computed its scale -- confirmed live: the
    real tables were confined to the top ~86% of the page's usable
    height instead of filling it, the same class of problem
    columns_to_hide fixes for width, just for height instead and
    unrelated to Module 2 at all (this still matters even for a subject
    combination that never needed a single non-canonical column hidden).

    These rows get *deleted* outright, not hidden -- a first version of
    this fix hid them the same way columns_to_hide hides a column
    (google_sheets_export.hide_rows, since removed), and confirmed live
    that doing so had *no effect whatsoever* on the exported PDF: a
    before/after export came out pixel-for-pixel identical, unlike hiding
    a column, which measurably fixed the analogous width problem. Only
    actually removing the rows shrinks what Sheets' print-area
    computation still counts.

    "Last real content" is found by scanning every cell on the sheet for
    a non-None value, deliberately not using `ws.max_row`/`ws.dimensions`
    the way an earlier, now-abandoned approach did (_tighten_print_areas,
    see this repo's history) -- those reflect exactly the stray
    formatting this needs to look *past*, not the real content boundary
    itself.
    """
    last_content_row = 0
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is not None:
                last_content_row = max(last_content_row, cell.row)
    if ws.max_row <= last_content_row:
        return []
    return [(last_content_row, ws.max_row)]


def _find_score_value_cells(ws: Worksheet) -> Dict[str, Tuple[int, int]]:
    """{subject: (row, col)} for every "<Subject> Score" label found (e.g.
    "Reading\n& Writing\nScore", "Math\nScore") -- these templates put the
    label a few rows *below* its own value cell (confirmed against a real
    template: "Total Score" is likewise labeled below the cell that sums
    it), so this searches upward from each label for the nearest cell in
    the same column that already holds a number -- the static placeholder
    value (e.g. 200) every blank template ships with in that slot."""
    result: Dict[str, Tuple[int, int]] = {}
    for row in ws.iter_rows():
        for cell in row:
            value = cell.value
            if not isinstance(value, str):
                continue
            normalized = re.sub(r"\s+", " ", value.strip())
            match = _SCORE_LABEL_PATTERN.match(normalized)
            if not match:
                continue
            try:
                subject = normalize_subject(match.group("subject"))
            except ValueError:
                continue  # e.g. "Total Score" -- not a subject this module knows
            for candidate_row in range(cell.row - 1, cell.row - _SCORE_VALUE_SEARCH_ROWS - 1, -1):
                candidate_value = ws.cell(row=candidate_row, column=cell.column).value
                # bool is technically an int subclass -- excluded explicitly
                # so a stray flag cell in the search window is never mistaken
                # for the score value.
                if isinstance(candidate_value, (int, float)) and not isinstance(candidate_value, bool):
                    result[subject] = (candidate_row, cell.column)
                    break
    return result


def _find_name_cell(ws: Worksheet) -> Tuple[int, int]:
    for row in ws.iter_rows():
        for cell in row:
            value = cell.value
            if isinstance(value, str) and value.strip().lower().startswith(_NAME_PLACEHOLDER_PREFIXES):
                return cell.row, cell.column
    raise ValueError(
        f"Could not find a name placeholder cell (looked for a value starting with "
        f"one of {_NAME_PLACEHOLDER_PREFIXES!r})"
    )


def fill_sat_score_report(
    template_path: str | Path,
    answers: Mapping[SatKey, str],
    active_variants: Mapping[str, str],
    student_name: str,
    test_date: dt.date | str,
    section_scores: Optional[Mapping[str, int]] = None,
    sheet_name: str = "Student Responses",
) -> FillResult:
    """Return every cell write needed to fill `template_path`'s
    `sheet_name` tab in with the student's name, test date, every answer
    in `answers`, and (if given) each subject's scaled section score --
    plus, in the same FillResult, the rectangles to clear and the whole
    columns to hide for every Module 2 block occurrence that isn't the
    one column every subject's real answers get consolidated into (see
    blocks_to_clear and columns_to_hide), so the exported report only
    shows the modules that were actually filled in -- both in content
    and in the exported PDF's own sizing -- rather than every
    duplicate/twin block the template ships with; plus the whole rows
    below `sheet_name`'s own real content to delete too (see
    trailing_rows_to_delete), an unrelated problem (the sheet's own
    trailing blank-but-formatted rows, nothing to do with Module 2) that
    inflates the same exported PDF's print area regardless.

    `active_variants` maps subject -> "easier"/"harder", the Module 2
    difficulty actually administered for that subject (from
    answer_keys.annotate_rows) -- this is what decides which of each
    subject's own two same-difficulty block-pairs its answers get read
    from. Regardless of which difficulty that turns out to be, every
    subject's real answers -- title, correct-answer key, your answer,
    Domain, Skill -- are written into the *same* column position on the
    sheet (_canonical_module2_col, always Higher's own canonical block):
    if a subject's active variant already lives there, its cells are
    written in place exactly as before; otherwise those same values are
    copied in from wherever that subject's real block actually sits, and
    that column's own single flag cell is the only one this ever sets
    True. Every other Module 2 occurrence -- a subject's own non-matching
    difficulty, and every duplicate/twin -- is left completely untouched
    here and cleared and hidden instead (see FillResult.cleared_ranges
    and .hidden_column_ranges).

    This exists because Reading & Writing and Math share the exact same
    four column positions for their own Module 2 blocks (see this
    module's own docstring on flag cells) -- when they administer
    *different* difficulties (confirmed live: R&W Higher, Math Lower),
    each subject's real answers naturally sit in different columns, which
    left their two Module 2 tables visibly offset from each other on the
    exported report (confirmed against the org's own reference example,
    where both sit at identical column positions instead) -- a whole-
    column hide/clear keyed only by which difficulty is "active" can
    never fix that, since it never moves anything, only hides/shows it in
    place. Consolidating into one column removes the offset entirely, and
    -- as a side effect -- also means the sheet's own score-total
    formulas (which read a fixed handful of cells by absolute address,
    not by which subject or difficulty) only ever need to see one flag
    true at a time.

    `answers` must only contain keys for "module1" or each subject's
    active variant -- an entry for the *inactive* variant raises, since
    writing it would silently go nowhere the sheet's own formulas count
    (its flag stays False). A template question with no entry in
    `answers` is left blank (a legitimate omitted bubble).

    `section_scores` maps subject -> scaled score (e.g. {"math": 620}) --
    unlike every other field here, this isn't something this pipeline can
    derive from the scan/report itself (see this module's own history in
    the repo for why: no formula computes it in this template, and
    nothing upstream currently extracts it either), so it's expected to
    come from wherever the caller sourced it -- e.g. a value a person
    typed into a prompt. Omit a subject (or the whole mapping) to leave
    its score cell at the template's own default, unchanged.
    """
    for subject, slot, _question in answers:
        if slot != "module1" and active_variants.get(subject) != slot:
            raise ValueError(
                f"Got an answer for {subject!r} {slot!r}, but the active variant for {subject!r} "
                f"is {active_variants.get(subject)!r}"
            )

    wb = openpyxl.load_workbook(template_path, data_only=True)
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"No {sheet_name!r} tab in {template_path} (tabs: {wb.sheetnames})")
    ws = wb[sheet_name]

    name_row, name_col = _find_name_cell(ws)
    writes = [
        CellWrite(sheet_name, name_row, name_col, student_name),
        # date sits directly below name
        CellWrite(sheet_name, name_row + 1, name_col, format_date_for_sheets(test_date)),
    ]

    if section_scores:
        score_cells = _find_score_value_cells(ws)
        for subject, score in section_scores.items():
            if subject not in score_cells:
                raise ValueError(
                    f"No score cell found for subject {subject!r} in {template_path}!{sheet_name} "
                    f"(found score cells for: {sorted(score_cells)})"
                )
            row, col = score_cells[subject]
            writes.append(CellWrite(sheet_name, row, col, score))

    canonical_col = _canonical_module2_col(ws)
    canonical_flag_cell = next(
        (b.flag_cell for b in locate_sat_blocks(ws) if b.question_col == canonical_col), None
    )

    remaining = dict(answers)
    overflow_title_cells: List[Tuple[str, int, int]] = []
    for block in locate_sat_blocks(ws):
        is_active = block.module_slot == "module1" or active_variants.get(block.subject) == block.module_slot
        if not is_active:
            continue  # a non-administered block -- leave completely untouched, cleared instead

        if block.module_slot == "module1":
            overflow_title_cells.append((sheet_name, block.title_row - 1, block.question_col - 1))
            writes.append(CellWrite(sheet_name, block.header_row, block.mark_col, _MARK_HEADER_NON_BLANK))
            r = block.header_row + 1
            while True:
                question = ws.cell(row=r, column=block.question_col).value
                if question is None:
                    break
                key = (block.subject, block.module_slot, int(question))
                writes.append(CellWrite(sheet_name, r, block.answer_col, remaining.pop(key, None)))
                r += 1
            continue

        # Module 2: consolidate into canonical_col regardless of which
        # difficulty this subject actually administered (see this
        # function's own docstring for why).
        if canonical_flag_cell is not None:
            frow, fcol = canonical_flag_cell
            writes.append(CellWrite(sheet_name, frow, fcol, True))

        repositioning = canonical_col is not None and block.question_col != canonical_col
        target_col = canonical_col if repositioning else block.question_col
        overflow_title_cells.append((sheet_name, block.title_row - 1, target_col - 1))
        writes.append(CellWrite(sheet_name, block.header_row, target_col + 3, _MARK_HEADER_NON_BLANK))
        if repositioning:
            title_text = ws.cell(row=block.title_row, column=block.question_col).value
            writes.append(CellWrite(sheet_name, block.title_row, target_col, title_text))

        r = block.header_row + 1
        while True:
            question = ws.cell(row=r, column=block.question_col).value
            if question is None:
                break
            key = (block.subject, block.module_slot, int(question))
            answer_value = remaining.pop(key, None)
            if repositioning:
                # Every other cell the report shows for this question --
                # the correct-answer key and Domain/Skill -- has to move
                # too, or it'd still read as the wrong (or no) difficulty
                # once the title above it changes. mark_col needs no
                # write of its own: it's a formula already sitting at
                # target_col comparing correct-answer against your-answer
                # at that same position, so it recomputes on its own once
                # both of those land there.
                correct_value = ws.cell(row=r, column=block.correct_col).value
                domain_value = ws.cell(row=r, column=block.question_col + 4).value
                skill_value = ws.cell(row=r, column=block.question_col + 5).value
                writes.append(CellWrite(sheet_name, r, target_col + 1, correct_value))
                writes.append(CellWrite(sheet_name, r, target_col + 2, answer_value))
                writes.append(CellWrite(sheet_name, r, target_col + 4, domain_value))
                writes.append(CellWrite(sheet_name, r, target_col + 5, skill_value))
            else:
                writes.append(CellWrite(sheet_name, r, block.answer_col, answer_value))
            r += 1

    if remaining:
        unmatched = ", ".join(f"{subject} {slot} {question}" for subject, slot, question in sorted(remaining))
        raise ValueError(f"{template_path}!{sheet_name} has no answer block for: {unmatched}")

    cleared_ranges = [(sheet_name, top, bottom, left, right) for top, bottom, left, right in blocks_to_clear(ws)]
    hidden_column_ranges = [(sheet_name, left, right) for left, right in columns_to_hide(ws)]
    deleted_row_ranges = [(sheet_name, top, bottom) for top, bottom in trailing_rows_to_delete(ws)]
    narrowed_column_ranges = [
        (sheet_name, left, right, factor)
        for left, right, factor in (visible_table_columns_to_narrow(ws) + hidden_columns_to_shrink(ws))
    ]
    bar = header_bar_extension(ws)
    header_bar_extension_ranges = [(sheet_name,) + bar] if bar is not None else []
    return FillResult(
        cell_writes=writes,
        cleared_ranges=cleared_ranges,
        hidden_column_ranges=hidden_column_ranges,
        deleted_row_ranges=deleted_row_ranges,
        narrowed_column_ranges=narrowed_column_ranges,
        header_bar_extension=header_bar_extension_ranges,
        overflow_title_cells=overflow_title_cells,
    )
